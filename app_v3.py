import os, io, re, json, socket, calendar
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo


# ── ENV FILE LOADING ────────────────────────────────────────────────────────
# Load credentials from a `.env` file at the project root, if present.
# This is the recommended location for secrets (gitignored by default) and
# keeps the WSGI config file free of inline credentials.
#
# Format (KEY=VALUE per line, blank lines and `#` comments ignored,
# optional surrounding quotes on the value are stripped):
#
#     SECRET_KEY=abc123...
#     CARTRACK_USERNAME=...
#     CARTRACK_PASSWORD=...
#     JOBORDERS_TOKEN=...
#
# Already-set process env vars are NEVER overwritten — .env only fills
# in missing keys. So values set via PA Web tab env vars, the WSGI
# config's `os.environ[...]` lines, or the shell continue to take
# precedence. This means deployments can be migrated to .env at their
# own pace without breaking existing setups.
#
# Mirrors the pattern used by cartrack_poll.py's _bootstrap_env_from_wsgi,
# so the polling worker and the Flask app read from the same .env file
# without further coordination.
def _load_env_file():
    here = os.path.dirname(os.path.abspath(__file__))
    env_path = os.path.join(here, '.env')
    if not os.path.exists(env_path):
        return
    try:
        with open(env_path, encoding='utf-8') as f:
            for line in f:
                s = line.strip()
                if not s or s.startswith('#') or '=' not in s:
                    continue
                key, _, val = s.partition('=')
                key = key.strip()
                val = val.strip()
                # Strip surrounding quotes if present.
                if len(val) >= 2 and val[0] == val[-1] and val[0] in ('"', "'"):
                    val = val[1:-1]
                if key and key not in os.environ:
                    os.environ[key] = val
    except Exception as _e:
        # Never let .env parsing kill startup — fall through and let the
        # downstream code (SECRET_KEY check, etc.) report missing keys.
        print(f'[env] WARNING: failed to parse .env: {_e}')


_load_env_file()


# ── FIREBASE SETUP ─────────────────────────────────────────────────────────
try:
    import firebase_admin
    from firebase_admin import credentials as fb_credentials, db as fb_db
    _fb_key = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           'dispatch-scheduler-ce39f-firebase-adminsdk-fbsvc-affa3531e0.json')
    if not firebase_admin._apps and os.path.exists(_fb_key):
        firebase_admin.initialize_app(
            fb_credentials.Certificate(_fb_key),
            {'databaseURL': 'https://dispatch-scheduler-ce39f-default-rtdb.firebaseio.com/'}
        )
    FIREBASE_OK = True
except Exception as _fb_err:
    FIREBASE_OK = False
    print(f"[Firebase] init skipped: {_fb_err}")

def firebase_notify(event='update'):
    """Push a tiny timestamp to Firebase so all browsers know to refresh."""
    if not FIREBASE_OK:
        return
    try:
        fb_db.reference('dispatch_updates').set({
            'ts': int(utc_now().timestamp() * 1000),
            'event': event
        })
    except Exception as e:
        print(f"[Firebase] notify failed: {e}")

PH_TZ  = ZoneInfo('Asia/Manila')
UTC_TZ = ZoneInfo('UTC')


def pht_filter_to_utc(date_str, plus_days=0):
    """Convert a PHT calendar-date filter string (YYYY-MM-DD) to the
    naive-UTC datetime bound used for querying UTC-stored timestamps
    (CartrackEvent.created_at, TruckCycle.started_at, SiteVisit
    enter/exit). PHT midnight = 16:00 UTC of the previous day; raw
    comparison without this shift made every date filter miss the
    00:00-07:59 AM PHT window — the busiest hours of the fleet."""
    d = datetime.strptime(date_str, '%Y-%m-%d') + timedelta(days=plus_days)
    return d.replace(tzinfo=PH_TZ).astimezone(UTC_TZ).replace(tzinfo=None)


def ph_now():
    """Current datetime in Philippine time (tz-aware)."""
    return datetime.now(PH_TZ)

def ph_today():
    """Current date in Philippine time."""
    return ph_now().date()


def iso_ph(dt):
    """Serialize a datetime (or date) to ISO-8601 for API responses.

    For datetime objects, emits a PHT-aware ISO-8601 string with
    +08:00 offset, e.g., '2026-05-21T08:09:50+08:00'. The DB stores
    naive UTC; this converts to PHT so JS new Date() parses correctly.

    For date objects, emits the calendar date as-is (e.g.,
    '2026-05-21'). date has no time component and no timezone, so
    no conversion is meaningful — Wave.date is a business date in
    PHT already.

    Behavior:
      - None             -> None (frontend renders '—')
      - date (not dt)    -> 'YYYY-MM-DD'
      - naive datetime   -> assume UTC, convert to PHT, +08:00 ISO
      - aware datetime   -> convert to PHT, +08:00 ISO
    """
    if dt is None:
        return None
    # date is the parent class of datetime, so we test datetime FIRST
    # before the date branch. (isinstance(x, date) is True for x a
    # datetime, but isinstance(x, datetime) is True ONLY for datetimes.)
    from datetime import date as _date, datetime as _datetime
    if isinstance(dt, _datetime):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UTC_TZ)
        return dt.astimezone(PH_TZ).isoformat()
    if isinstance(dt, _date):
        return dt.isoformat()
    # Anything else — best effort string conversion (shouldn't happen)
    return str(dt)
from flask import (Flask, render_template, request, redirect, url_for,
                   jsonify, session, flash, send_file)
from models_v2 import (db, utc_now, TruckTypeDef, Wave, TripRecord,
                       Driver, Helper, Product, Client, Dispatcher, Plate,
                       ChangeLog, Attendance, HelperAttendance, BreakdownLog, AppSetting,
                       CartrackTruckState, CartrackEvent,
                       CartrackGeofence, SiteVisit, TruckCycle,
                       TRUCK_TYPES_SEED, STATUSES,
                       ATTENDANCE_STATUSES, BREAKDOWN_STATUSES, TRIP_TYPES,
                       DOC_HEADER_DEFAULTS, SHEETS_WEBHOOK_KEY, SHEETS_WEBHOOK_DEFAULT,
                       FULL_DAY_PRODUCTS_SEED, FULL_DAY_KEYWORDS, is_product_full_day)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Support configurable DB path for cloud deployment (e.g. Render persistent disk at /var/data)
DB_PATH = os.environ.get('DB_PATH', os.path.join(BASE_DIR, 'dispatch.db'))

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{DB_PATH}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# ── Security: SECRET_KEY ────────────────────────────────────────────────────
# Used by Flask to sign session cookies. MUST be set as an environment
# variable in production — if missing, refuse to start rather than fall
# back to a hard-coded value (the previous default 'dispatch-scheduler-2026'
# was visible in the public repo, which would let anyone forge sessions).
#
# To run locally without setting SECRET_KEY in the environment, set
# FLASK_DEV_INSECURE=1 to use an ephemeral random key — sessions reset
# on every restart, but no hard-coded value ever leaks.
SECRET_KEY = os.environ.get('SECRET_KEY')
if not SECRET_KEY:
    if os.environ.get('FLASK_DEV_INSECURE') == '1':
        import secrets
        SECRET_KEY = secrets.token_hex(32)
        print('[security] WARNING: using ephemeral dev SECRET_KEY '
              '(FLASK_DEV_INSECURE=1). Sessions will reset on restart.')
    else:
        raise RuntimeError(
            'SECRET_KEY environment variable is required. '
            'Set it in the PythonAnywhere Web tab (or in a .env file for '
            'local dev). For ephemeral local-dev use, set '
            'FLASK_DEV_INSECURE=1 instead.'
        )
app.config['SECRET_KEY'] = SECRET_KEY

# ── Security: session cookie hardening ──────────────────────────────────────
# HTTPONLY  — prevents JS from reading the cookie (XSS mitigation).
# SECURE    — only sent over HTTPS. PythonAnywhere serves HTTPS by default,
#             so this is safe in production. Set FLASK_DEV_INSECURE=1 to
#             relax for plain-HTTP local dev.
# SAMESITE  — 'Lax' blocks cookies on cross-site POSTs (CSRF mitigation),
#             while still allowing top-level navigation links to work.
# LIFETIME  — sessions auto-expire after 12 hours of session.permanent=True.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=(os.environ.get('FLASK_DEV_INSECURE') != '1'),
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
)

# ── Security: ProxyFix middleware ───────────────────────────────────────────
# PythonAnywhere terminates HTTPS at the front-end proxy and forwards the
# request to the WSGI app over plain HTTP. Without ProxyFix the app would
# see request.scheme = 'http' and url_for(_external=True) would generate
# http:// URLs even though the browser is on https://. ProxyFix reads the
# X-Forwarded-Proto / X-Forwarded-For headers PA sets and corrects the
# perceived scheme / host / client IP. Safe to enable behind any single
# reverse proxy.
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

db.init_app(app)

# ── AUTH BLUEPRINT ─────────────────────────────────────────────────────────
from auth import auth_bp                              # noqa: E402
from auth.routes import login_required, admin_required, check_can_delete  # noqa: E402
app.register_blueprint(auth_bp)


# ── HELPERS ────────────────────────────────────────────────────────────────
def parse_date(s):
    try:    return date.fromisoformat(s)
    except: return ph_today()

def get_user():
    return session.get('user_name', 'Dispatcher')

def log_change(action, entity='trip'):
    db.session.add(ChangeLog(user_name=get_user(), action=action, entity=entity))
    firebase_notify(entity)

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return '127.0.0.1'

def master_lists():
    """Return all master data lists for template use."""
    return dict(
        all_drivers     = Driver.query.filter_by(active=True).order_by(Driver.name).all(),
        all_helpers     = Helper.query.filter_by(active=True).order_by(Helper.name).all(),
        all_products    = Product.query.filter_by(active=True).order_by(Product.name).all(),
        all_clients     = Client.query.filter_by(active=True).order_by(Client.name).all(),
        all_dispatchers = Dispatcher.query.filter_by(active=True).order_by(Dispatcher.name).all(),
        all_plates      = Plate.query.filter_by(active=True).order_by(Plate.plate_no).all(),
        truck_types     = TruckTypeDef.query.order_by(TruckTypeDef.sort_order).all(),
    )

@app.context_processor
def inject_globals():
    doc = {k: AppSetting.get(k, v) for k, v in DOC_HEADER_DEFAULTS.items()}
    return dict(
        now=ph_now(),
        today=ph_today(),
        timedelta=timedelta,
        statuses=STATUSES,
        current_user=get_user(),
        current_user_role=session.get('user_role', ''),
        local_ip=get_local_ip(),
        doc=doc,
    )


# ── INDEX ──────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return redirect(url_for('schedule', date_str=ph_today().isoformat()))


# ── SCHEDULE ───────────────────────────────────────────────────────────────
@app.route('/schedule/<date_str>')
@login_required
def schedule(date_str):
    d = parse_date(date_str)
    truck_types = TruckTypeDef.query.order_by(TruckTypeDef.sort_order).all()

    # Build schedule map: {truck_type_code: [Wave, ...]}
    schedule_map = {}
    for tt in truck_types:
        waves = (Wave.query
                 .filter_by(date=d, truck_type_id=tt.id)
                 .order_by(Wave.wave_number).all())
        schedule_map[tt.code] = waves

    # Counts for tab badges
    counts = {}
    for tt in truck_types:
        cnt = (db.session.query(db.func.count(TripRecord.id))
               .join(Wave)
               .filter(Wave.date == d, Wave.truck_type_id == tt.id)
               .scalar() or 0)
        counts[tt.code] = cnt

    doc = {k: AppSetting.get(k, v) for k, v in DOC_HEADER_DEFAULTS.items()}
    return render_template('schedule/daily.html',
        d=d, schedule_map=schedule_map, counts=counts,
        trip_types=TRIP_TYPES, doc=doc,
        **master_lists())


# ── API: WAVES ─────────────────────────────────────────────────────────────
@app.route('/api/wave/add', methods=['POST'])
@login_required
def api_wave_add():
    data          = request.get_json()
    date_str      = data.get('date')
    truck_code    = data.get('truck_code')
    d             = parse_date(date_str)
    tt            = TruckTypeDef.query.filter_by(code=truck_code).first_or_404()

    # next wave number
    last = (Wave.query.filter_by(date=d, truck_type_id=tt.id)
            .order_by(Wave.wave_number.desc()).first())
    next_num = (last.wave_number + 1) if last else 1

    wave = Wave(date=d, truck_type_id=tt.id, wave_number=next_num)
    db.session.add(wave)
    db.session.commit()
    log_change(f"Added {wave.label} for {tt.name} on {d}", 'wave')
    db.session.commit()

    return jsonify({'wave_id': wave.id, 'wave_number': next_num, 'label': wave.label})


@app.route('/api/wave/<int:wid>/delete', methods=['POST'])
@login_required
def api_wave_delete(wid):
    if not check_can_delete():
        return jsonify({'error': 'You do not have permission to delete.'}), 403
    wave = Wave.query.get_or_404(wid)
    info = f"{wave.label} ({wave.truck_type.name}) on {wave.date}"
    db.session.delete(wave)
    log_change(f"Deleted {info}", 'wave')
    db.session.commit()
    return jsonify({'ok': True})


# ── API: TRIPS ──────────────────────────────────────────────────────────────
@app.route('/api/trip/save', methods=['POST'])
@login_required
def api_trip_save():
    data     = request.get_json()
    trip_id  = data.get('trip_id')
    wave_id  = data.get('wave_id')

    if trip_id:
        trip = TripRecord.query.get_or_404(trip_id)
    else:
        # New row — the wave must still exist. SQLite runs without FK
        # enforcement, so inserting against a deleted wave (stale page
        # after a revert/cleanup removed it) would persist an orphan
        # trip that no schedule query can display or delete.
        if not db.session.get(Wave, wave_id):
            return jsonify({'error': 'This wave no longer exists — '
                            'reload the page.'}), 409
        last = (TripRecord.query.filter_by(wave_id=wave_id)
                .order_by(TripRecord.trip_number.desc()).first())
        trip = TripRecord(wave_id=wave_id,
                          trip_number=(last.trip_number + 1 if last else 1))
        db.session.add(trip)

    # Apply all provided fields
    field_map = {
        'driver_id': ('driver_id', int),
        'helper_id': ('helper_id', int),
        'plate_id':  ('plate_id',  int),
        'product_id':('product_id',int),
        'client_id': ('client_id', int),
        'dispatcher_id':('dispatcher_id',int),
        'trip_type':       ('trip_type',       str),
        'rs_no':           ('rs_no',           str),
        'po_no':           ('po_no',           str),
        'reference':       ('reference',       str),
        'dr_no':           ('dr_no',           str),
        'volume':          ('volume',          str),
        'status':          ('status',          str),
        'toll_fee':        ('toll_fee',        float),
        'toll_expressway': ('toll_expressway', str),
        'toll_entry':      ('toll_entry',      str),
        'toll_exit':       ('toll_exit',       str),
        'toll_class':      ('toll_class',      str),
        'notes':           ('notes',           str),
    }
    for key, (attr, cast) in field_map.items():
        if key in data:
            val = data[key]
            if val == '' or val is None:
                setattr(trip, attr, None)
            else:
                try:
                    setattr(trip, attr, cast(val))
                except (ValueError, TypeError):
                    setattr(trip, attr, None)

    trip.updated_by = get_user()
    trip.updated_at = utc_now()
    db.session.commit()

    wave = Wave.query.get(trip.wave_id)
    log_change(
        f"Saved trip #{trip.trip_number} in {wave.label} "
        f"({wave.truck_type.name}) on {wave.date}", 'trip')
    db.session.commit()

    return jsonify(trip.to_dict())


@app.route('/api/trip/<int:tid>/delete', methods=['POST'])
@login_required
def api_trip_delete(tid):
    if not check_can_delete():
        return jsonify({'error': 'You do not have permission to delete.'}), 403
    trip = TripRecord.query.get_or_404(tid)
    wave = trip.wave
    info = f"trip #{trip.trip_number} in {wave.label} ({wave.truck_type.name}) on {wave.date}"
    db.session.delete(trip)
    log_change(f"Deleted {info}", 'trip')
    db.session.commit()
    return jsonify({'ok': True})


# ── API: SCHEDULE IMPORT (monthly monitoring workbook) ─────────────────────
def _imp_norm(s):
    """Lowercase alphanumeric-only key for tolerant name matching
    ('NIZ 1044' == 'niz1044', 'A. Eusebio' == 'A EUSEBIO')."""
    return ''.join(c for c in str(s or '').lower() if c.isalnum())


def _imp_header_key(v):
    """Collapse a header cell to a comparable key: newlines and runs of
    spaces become one space, punctuation dropped, lowercased."""
    s = ' '.join(str(v or '').split()).lower()
    s = s.replace('.', '').replace(',', '').replace('/', ' ')
    return ' '.join(s.split())


# Only these tabs of the monitoring workbook hold trip records
# (agreed with operations, July 2026). Matched on the normalised
# sheet name so '2.1 Data_Rental' and 'Data Rental' both qualify.
_IMPORT_SHEET_KEYWORDS = ('data input', 'data rental', 'data cps',
                          'waste input')

# Header-cell text (in _imp_header_key form) -> field name. Matched
# EXACTLY so 'source dr no' can never claim the 'dr no' column. Each
# sheet names its columns slightly differently, hence the aliases:
#   2.0 Data Input   - Delivery Date / Product Description / RS No.
#   2.1 Data_Rental  - Date of Rental / Product / Service Category
#   2.2 Data_CPS     - Delivery Date / Status
#   6.0 Waste Input  - Date / Waste Materials / Destination / HRF No.
_IMPORT_HEADERS = {
    'delivery date':             'date',
    'date of rental':            'date',
    'date':                      'date',
    'client name':               'client',
    'destination':               'client',      # waste runs: haul-to site
    'product description':       'product',
    'product service category':  'product',
    'waste materials':           'product',
    'truck plate no & id':       'plate',
    'driver name':               'driver',
    'helper name':               'helper',
    'helper 1 name':             'helper',
    'dispatcher':                'dispatcher',
    'delivery status':           'status',
    'status':                    'status',
    'rs no':                     'rs_no',
    'hrf no':                    'rs_no',
    'client po no':              'po_no',
    'dr no':                     'dr_no',
    # Volume must be the PER-TRIP load, never the PO total ("PO Volume
    # Aggregates" is the whole order and repeats on every row of that
    # PO). Leftmost matching column wins per sheet, so 2.0 uses the
    # source-loaded volume (col 15) with the received volume as backup.
    'source aggregates volume m3': 'volume',
    'agg received volume m3':      'volume',
    'received quantity':           'volume',
    'volume m3':                   'volume',
}

# Import batches are journaled here so a bad import can be reverted.
_IMPORT_BATCH_FILE = os.path.join(BASE_DIR, 'instance',
                                  'schedule_import_batches.json')


def _load_import_batches():
    try:
        with open(_IMPORT_BATCH_FILE, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []


def _save_import_batches(batches):
    os.makedirs(os.path.dirname(_IMPORT_BATCH_FILE), exist_ok=True)
    with open(_IMPORT_BATCH_FILE, 'w', encoding='utf-8') as f:
        json.dump(batches[-10:], f, indent=1)

_IMPORT_STATUS_MAP = {
    'completed': 'Delivered',
    'delivered': 'Delivered',
    'cancelled': 'Canceled',
    'canceled':  'Canceled',
    'intransit': 'In Transit',
}

# Trip-type + hauling rules (per operations, July 2026):
#   - Internal material transfers (client is the LOG/RMP stockpile, the
#     RMC plant, or RMP) are BACK LOADS; every other client is a FRONT
#     LOAD. Matched on whole word tokens so e.g. 'Ready Mix Concrete
#     Corp.' does NOT trigger the 'RMC' rule.
#   - 'Hauling': a 12W or 22WD dump truck serving the internal RMC /
#     Asphalt Plant / CPS clients. Hauling is excluded from fleet
#     utilisation, so those trips are filed under the OT (Others)
#     truck category instead of the plate's own type.
_IMPORT_BACKLOAD_TOKENS = {'rmc', 'rmp', 'stockpile'}
_IMPORT_HAUL_TOKENS     = {'rmc', 'cps'}
_IMPORT_HAUL_TT_CODES   = {'12W', '22WD'}


def _imp_client_tokens(name):
    return set(re.findall(r'[a-z0-9]+', str(name or '').lower()))


@app.route('/api/schedule/import-xlsx', methods=['POST'])
@login_required
def api_schedule_import_xlsx():
    """Import a monthly 'Daily Sales and Logistics Materials Monitoring'
    workbook into Waves + TripRecords.

    Called twice by the UI with the same file: commit=0 returns a
    preview (nothing written); commit=1 performs the import.

    Rules (agreed with operations, July 2026):
      - Wave = the plate's Nth trip of that day: a plate's first trip
        lands in Wave 1 of its truck type, its second in Wave 2, etc.
      - Unknown clients/products/drivers/helpers/dispatchers/plates are
        auto-created (new plates get truck type OT — recategorise in
        Master Data afterwards; the preview lists everything new).
      - Only fields the app already has are imported; the rest of the
        workbook's 56 columns are ignored.
      - Re-uploading the same (or an updated) file never duplicates:
        a row is skipped when its DR No. already exists on that date,
        or when an identical trip already sits in the same wave.
    """
    up = request.files.get('file')
    if not up or not up.filename:
        return jsonify({'error': 'No file uploaded.'}), 400
    commit = request.form.get('commit') == '1'

    try:
        wb = openpyxl.load_workbook(up, data_only=True, read_only=True)
    except Exception as e:
        return jsonify({'error': f'Could not read the Excel file: {e}'}), 400

    # ── Locate the trip-record sheets ────────────────────────────────
    # Only the agreed tabs are read (2.0 Data Input, 2.1 Data_Rental,
    # 2.2 Data_CPS, 6.0 Waste Input) — matched on normalised sheet
    # name; the Data Input sheet is scanned first so cross-sheet
    # duplicates keep the primary sheet's copy.
    target_sheets = [n for n in wb.sheetnames
                     if any(k in n.lower().replace('_', ' ')
                            for k in _IMPORT_SHEET_KEYWORDS)]
    target_sheets.sort(key=lambda n:
                       0 if 'data input' in n.lower().replace('_', ' ') else 1)
    if not target_sheets:
        return jsonify({'error':
            'None of the expected sheets (2.0 Data Input, 2.1 '
            'Data_Rental, 2.2 Data_CPS, 6.0 Waste Input) were found in '
            'this workbook.'}), 400

    parsed, sheets_read, bad_dates = [], [], 0
    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        colmap, hdr_row_idx = None, 0
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15,
                                                  values_only=True), 1):
            found = {}
            for c_idx, cell in enumerate(row or ()):
                key = _imp_header_key(cell)
                if key in _IMPORT_HEADERS:
                    found.setdefault(_IMPORT_HEADERS[key], c_idx)
            if len(found) >= 5 and 'date' in found and 'plate' in found:
                colmap, hdr_row_idx = found, r_idx
                break
        if colmap is None:
            continue                       # tab exists but no data table

        def _cell(row, field):
            i = colmap.get(field)
            if i is None or i >= len(row):
                return None
            return row[i]

        def _text(row, field, limit=120):
            v = _cell(row, field)
            if v is None:
                return ''
            return ' '.join(str(v).split())[:limit]

        n_before = len(parsed)
        for row in ws.iter_rows(min_row=hdr_row_idx + 1, values_only=True):
            raw_date = _cell(row, 'date')
            if raw_date is None:
                continue
            if isinstance(raw_date, datetime):
                d = raw_date.date()
            elif isinstance(raw_date, date):
                d = raw_date
            else:
                # STRICT text-date parsing. parse_date() falls back to
                # ph_today() on garbage, which turned rows whose date
                # cell held the text 'Cancelled' into trips dated the
                # day of the import — never accept a non-date here.
                # Accepted shapes: ISO date/datetime text
                # ('2026-05-15', '2026-05-15 00:00:00'), m/d/y with or
                # without a time tail ('5/15/2026 8:00'), and month-name
                # dates ('September 3, 2026'). No %d/%m fallback — the
                # operation writes m/d/y, and a d/m guess would silently
                # month-swap instead of flagging the row.
                s = str(raw_date).strip()
                d = None
                try:                       # ISO prefix, time tail ignored
                    d = date.fromisoformat(s[:10])
                except ValueError:
                    tok = s.split()[0] if s.split() else s
                    for cand, fmt in ((tok, '%m/%d/%Y'), (tok, '%m-%d-%Y'),
                                      (s, '%B %d, %Y'), (s, '%b %d, %Y')):
                        try:
                            d = datetime.strptime(cand, fmt).date()
                            break
                        except ValueError:
                            continue
                if d is None:
                    bad_dates += 1
                    continue
            # A schedule row needs at least a client or a plate to mean
            # anything; header-noise and summary rows have neither.
            if not (_text(row, 'client') or _text(row, 'plate')):
                continue
            parsed.append({
                'sheet':      sheet_name,
                'date':       d,
                'client':     _text(row, 'client'),
                'product':    _text(row, 'product'),
                'plate':      _text(row, 'plate', 60),
                'driver':     _text(row, 'driver', 80),
                'helper':     _text(row, 'helper', 80),
                'dispatcher': _text(row, 'dispatcher', 80),
                'status':     _IMPORT_STATUS_MAP.get(
                                  _imp_norm(_text(row, 'status')), 'Pending'),
                'rs_no':      _text(row, 'rs_no', 60),
                'po_no':      _text(row, 'po_no', 60),
                'dr_no':      _text(row, 'dr_no', 60),
                'volume':     _text(row, 'volume', 30),
            })
        sheets_read.append(f'{sheet_name} ({len(parsed) - n_before})')
    wb.close()
    if not parsed:
        return jsonify({'error': 'No rows with a delivery date found in '
                                 'the expected sheets.'}), 400

    # ── Cross-sheet dedup ────────────────────────────────────────────
    # The workbook intentionally double-enters some trips (e.g. rental
    # rows mirrored between tabs). A row whose (date, plate, driver,
    # DR no) signature was already read from ANOTHER sheet is a mirror
    # and is skipped; repeats within the SAME sheet are genuine
    # multiple runs and stay.
    seen_sig, kept, skipped_cross = {}, [], 0
    for r in parsed:
        sig = (r['date'].isoformat(), _imp_norm(r['plate']),
               _normalize_driver_key(r['driver']) or _imp_norm(r['driver']),
               _imp_norm(r['dr_no']))
        first_sheet = seen_sig.setdefault(sig, r['sheet'])
        if first_sheet != r['sheet']:
            skipped_cross += 1
            continue
        kept.append(r)
    parsed = kept

    # ── Entity resolution (existing first, auto-create otherwise) ────
    caches = {
        'client':     {_imp_norm(x.name): x for x in Client.query.all()},
        'product':    {_imp_norm(x.name): x for x in Product.query.all()},
        'driver':     {_imp_norm(x.name): x for x in Driver.query.all()},
        'helper':     {_imp_norm(x.name): x for x in Helper.query.all()},
        'dispatcher': {_imp_norm(x.name): x for x in Dispatcher.query.all()},
    }
    model_for = {'client': Client, 'product': Product, 'driver': Driver,
                 'helper': Helper, 'dispatcher': Dispatcher}
    plates_by_no   = {_imp_norm(p.plate_no): p for p in Plate.query.all()
                      if p.plate_no}
    plates_by_body = {_imp_norm(p.body_no): p for p in Plate.query.all()
                      if p.body_no}
    ot_type = TruckTypeDef.query.filter_by(code='OT').first() \
              or TruckTypeDef.query.order_by(TruckTypeDef.sort_order.desc()).first()
    tt_code_by_id = {t.id: t.code for t in TruckTypeDef.query.all()}

    # Person-name variant matching: 'Arnel Eusebio' == 'A. Eusebio' ==
    # 'A.EUSEBIO' — reuse the Breakdown module's canonical initial +
    # surname key, plus a fuzzy-surname fallback (same initial,
    # SequenceMatcher >= 0.85) so 'R. Hagonoy' still finds 'R. Hagunoy'
    # instead of creating a double entry.
    from difflib import SequenceMatcher
    person_kinds = ('driver', 'helper', 'dispatcher')
    person_keys = {k: {} for k in person_kinds}
    for kind in person_kinds:
        for key, obj in caches[kind].items():
            pk = _normalize_driver_key(obj.name)
            if pk:
                person_keys[kind].setdefault(pk, obj)

    def _person_match(kind, name):
        pk = _normalize_driver_key(name)
        if not pk:
            return None
        obj = person_keys[kind].get(pk)
        if obj is not None:
            return obj
        initial, _, surname = pk.partition('.')
        best, best_ratio = None, 0.0
        for ek, ex in person_keys[kind].items():
            e_init, _, e_sur = ek.partition('.')
            if e_init != initial:
                continue
            ratio = SequenceMatcher(None, surname, e_sur).ratio()
            if ratio > best_ratio:
                best, best_ratio = ex, ratio
        return best if best_ratio >= 0.85 else None

    new_master = {k: [] for k in
                  ('client', 'product', 'driver', 'helper', 'dispatcher',
                   'plate')}
    created_ids = {k: [] for k in new_master}   # for the revert journal

    def resolve(kind, name):
        """Return the entity for `name`, creating it when committing.
        Returns None for blank names."""
        if not name:
            return None
        key = _imp_norm(name)
        obj = caches[kind].get(key)
        if obj is not None:
            return obj
        if kind in person_kinds:
            obj = _person_match(kind, name)
            if obj is not None:
                caches[kind][key] = obj    # remember the variant
                return obj
        if name not in new_master[kind]:
            new_master[kind].append(name)
        if not commit:
            return None                    # preview: record, don't create
        obj = model_for[kind](name=name, active=True)
        db.session.add(obj)
        db.session.flush()
        caches[kind][key] = obj
        created_ids[kind].append(obj.id)
        if kind in person_kinds:
            pk = _normalize_driver_key(name)
            if pk:
                person_keys[kind].setdefault(pk, obj)
        return obj

    def resolve_plate(raw):
        """'NIZ 1044_DT-32' -> Plate row (matched on plate no OR body
        no, either side of the underscore)."""
        if not raw:
            return None
        parts = [p.strip() for p in raw.replace('_', '|').split('|')
                 if p.strip()]
        for part in parts:
            hit = plates_by_no.get(_imp_norm(part)) \
                  or plates_by_body.get(_imp_norm(part))
            if hit:
                return hit
        if raw not in new_master['plate']:
            new_master['plate'].append(raw)
        if not commit:
            return None
        p = Plate(plate_no=parts[0] if parts else raw,
                  body_no=parts[1] if len(parts) > 1 else '',
                  truck_type_id=ot_type.id, active=True)
        db.session.add(p)
        db.session.flush()
        created_ids['plate'].append(p.id)
        plates_by_no[_imp_norm(p.plate_no)] = p
        if p.body_no:
            plates_by_body[_imp_norm(p.body_no)] = p
        return p

    # ── Assemble waves + trips ───────────────────────────────────────
    # Keep the workbook's row order within each date so a plate's
    # first-listed trip becomes Wave 1, its second Wave 2, and so on.
    parsed.sort(key=lambda r: r['date'])
    plate_seq   = {}     # (date, plate_key) -> how many trips seen so far
    wave_cache  = {}     # (date, tt_id, wave_number) -> Wave
    planned     = set()  # wave keys counted as to-create (preview mode)
    trip_counts = {}     # wave key -> next trip_number bookkeeping
    existing_dr = {}     # date -> set of dr_no already in the DB
    new_waves   = []     # Wave objects created (revert journal)
    new_trips   = []     # TripRecord objects created (revert journal)

    stats = {'total_rows': len(parsed), 'imported': 0, 'skipped_dupe': 0,
             'skipped_no_plate': 0, 'waves_created': 0,
             'skipped_cross_sheet': skipped_cross, 'bad_dates': bad_dates,
             'hauling_to_others': 0}
    per_date = {}

    for r in parsed:
        d = r['date']
        plate = resolve_plate(r['plate'])
        plate_key = _imp_norm(r['plate']) or '(none)'
        if plate is None and commit:
            # blank plate cell — a trip row needs a truck
            stats['skipped_no_plate'] += 1
            continue
        if not r['plate']:
            stats['skipped_no_plate'] += 1
            continue

        # Trip type from the client: internal stockpile/RMC/RMP
        # transfers are back loads, everything else is a front load.
        # Waste hauling (6.0 Waste Input sheet) flips the default:
        # only Eco Protect runs are front loads; every other waste
        # destination is a back load.
        c_toks = _imp_client_tokens(r['client'])
        if 'waste input' in r['sheet'].lower().replace('_', ' '):
            trip_type = ('Front Load'
                         if 'ecoprotect' in _imp_norm(r['client'])
                         else 'Back Load')
        else:
            trip_type = ('Back Load' if (c_toks & _IMPORT_BACKLOAD_TOKENS)
                         else 'Front Load')

        # Hauling ("hustling") detection: a 12W/22WD dump truck serving
        # RMC / Asphalt Plant / CPS files under OT so utilisation
        # ignores it, and its trip type is Hustling rather than the
        # client-derived front/back load.
        tt_id = plate.truck_type_id if plate else ot_type.id
        plate_tt_code = tt_code_by_id.get(tt_id, '')
        is_hauling = (plate_tt_code in _IMPORT_HAUL_TT_CODES
                      and bool((c_toks & _IMPORT_HAUL_TOKENS)
                               or 'asphalt' in (r['client'] or '').lower()))
        if is_hauling:
            tt_id = ot_type.id
            trip_type = 'Hustling'
            stats['hauling_to_others'] += 1

        # Wave number = the plate's Nth trip of the day WITHIN this
        # category. Hauling runs count against the OT tab and
        # deliveries against the plate's own truck type — with one
        # shared counter, a truck whose morning was spent hauling had
        # its first DELIVERY land in e.g. Wave 3 of its own tab,
        # scattering single-trip waves across the schedule.
        seq_key = (d, plate_key, tt_id)
        plate_seq[seq_key] = plate_seq.get(seq_key, 0) + 1
        wave_no = plate_seq[seq_key]
        wkey = (d, tt_id, wave_no)
        wave = wave_cache.get(wkey)
        if wave is None:
            wave = Wave.query.filter_by(date=d, truck_type_id=tt_id,
                                        wave_number=wave_no).first()
            if wave is None and wkey not in planned:
                # Count each missing wave once — in preview mode nothing
                # is created, so without `planned` every row that lands
                # in the same new wave would inflate the count.
                planned.add(wkey)
                stats['waves_created'] += 1
                if commit:
                    wave = Wave(date=d, truck_type_id=tt_id,
                                wave_number=wave_no)
                    db.session.add(wave)
                    db.session.flush()
                    new_waves.append(wave)
            if wave is not None:
                wave_cache[wkey] = wave

        # ── Dedup ────────────────────────────────────────────────────
        if d not in existing_dr:
            existing_dr[d] = {t.dr_no for t in
                              (TripRecord.query.join(Wave)
                               .filter(Wave.date == d,
                                       TripRecord.dr_no.isnot(None),
                                       TripRecord.dr_no != '').all())}
        if r['dr_no'] and r['dr_no'] in existing_dr[d]:
            stats['skipped_dupe'] += 1
            continue

        client     = resolve('client', r['client'])
        product    = resolve('product', r['product'])
        driver     = resolve('driver', r['driver'])
        helper     = resolve('helper', r['helper'])
        dispatcher = resolve('dispatcher', r['dispatcher'])

        if wave is not None and plate is not None:
            dup = (TripRecord.query
                   .filter_by(wave_id=wave.id, plate_id=plate.id,
                              client_id=client.id if client else None,
                              product_id=product.id if product else None,
                              driver_id=driver.id if driver else None)
                   .first())
            if dup is not None:
                stats['skipped_dupe'] += 1
                continue

        stats['imported'] += 1
        per_date[d.isoformat()] = per_date.get(d.isoformat(), 0) + 1

        if commit:
            tkey = wave.id
            if tkey not in trip_counts:
                last = (TripRecord.query.filter_by(wave_id=wave.id)
                        .order_by(TripRecord.trip_number.desc()).first())
                trip_counts[tkey] = last.trip_number if last else 0
            trip_counts[tkey] += 1
            trip = TripRecord(
                wave_id=wave.id, trip_number=trip_counts[tkey],
                plate_id=plate.id,
                client_id=client.id if client else None,
                product_id=product.id if product else None,
                driver_id=driver.id if driver else None,
                helper_id=helper.id if helper else None,
                dispatcher_id=dispatcher.id if dispatcher else None,
                trip_type=trip_type,
                status=r['status'],
                rs_no=r['rs_no'] or None, po_no=r['po_no'] or None,
                dr_no=r['dr_no'] or None, volume=r['volume'] or None,
            )
            db.session.add(trip)
            new_trips.append(trip)
            if r['dr_no']:
                existing_dr[d].add(r['dr_no'])

    batch_id = None
    if commit:
        db.session.commit()
        # Journal the batch so it can be reverted from the UI.
        batch_id = utc_now().strftime('%Y%m%d-%H%M%S')
        batches = _load_import_batches()
        batches.append({
            'id':      batch_id,
            'file':    up.filename,
            'at':      utc_now().isoformat(),
            'user':    session.get('user_name') or '',
            'trips':   [t.id for t in new_trips],
            'waves':   [w.id for w in new_waves],
            'master':  {k: v for k, v in created_ids.items() if v},
            'stats':   stats,
        })
        _save_import_batches(batches)
        log_change(
            f"Imported {stats['imported']} trips "
            f"({stats['waves_created']} new waves, "
            f"{stats['skipped_dupe']} duplicates skipped) from "
            f"{up.filename} [batch {batch_id}]", 'trip')
        db.session.commit()

    dates_sorted = dict(sorted(per_date.items()))
    return jsonify({
        'preview':    not commit,
        'batch_id':   batch_id,
        'sheets':     sheets_read,
        'stats':      stats,
        'dates':      dates_sorted,
        'first_date': next(iter(dates_sorted), None),
        'new_master': {k: v for k, v in new_master.items() if v},
    })


@app.route('/api/schedule/import-batches')
@login_required
def api_schedule_import_batches():
    """Recent import batches (newest first) for the Revert UI."""
    out = []
    for b in reversed(_load_import_batches()):
        out.append({'id': b['id'], 'file': b.get('file', ''),
                    'at': b.get('at', ''), 'user': b.get('user', ''),
                    'trips': len(b.get('trips', [])),
                    'reverted': bool(b.get('reverted'))})
    return jsonify({'batches': out[:5]})


@app.route('/api/schedule/import-revert', methods=['POST'])
@login_required
def api_schedule_import_revert():
    """Undo an import batch: delete the trips it created, then any of
    its waves that are now empty, then any master-data rows it created
    that nothing references anymore."""
    data = request.get_json(silent=True) or {}
    want = data.get('batch_id')
    batches = _load_import_batches()
    batch = next((b for b in reversed(batches)
                  if not b.get('reverted')
                  and (b['id'] == want if want else True)), None)
    if batch is None:
        return jsonify({'error': 'No import batch available to revert.'}), 404

    trip_ids = batch.get('trips', [])
    removed_trips = 0
    # Chunked: PythonAnywhere's SQLite caps bound parameters at 999,
    # and a 4-sheet month can exceed that.
    for i in range(0, len(trip_ids), 500):
        removed_trips += (TripRecord.query
                          .filter(TripRecord.id.in_(trip_ids[i:i + 500]))
                          .delete(synchronize_session=False))
    db.session.flush()

    removed_waves = 0
    for wid in batch.get('waves', []):
        w = db.session.get(Wave, wid)
        if w is not None and not TripRecord.query.filter_by(wave_id=wid).first():
            db.session.delete(w)
            removed_waves += 1

    # Master data created by this batch: remove only rows nothing
    # references anymore (a later manual trip may already be using a
    # driver the import created — those stay).
    fk_for = {'driver': TripRecord.driver_id, 'helper': TripRecord.helper_id,
              'product': TripRecord.product_id, 'client': TripRecord.client_id,
              'dispatcher': TripRecord.dispatcher_id,
              'plate': TripRecord.plate_id}
    mdl_for = {'driver': Driver, 'helper': Helper, 'product': Product,
               'client': Client, 'dispatcher': Dispatcher, 'plate': Plate}
    removed_master = {}
    for kind, ids in (batch.get('master') or {}).items():
        n = 0
        for mid in ids:
            obj = db.session.get(mdl_for[kind], mid)
            if obj is not None and not (TripRecord.query
                                        .filter(fk_for[kind] == mid).first()):
                db.session.delete(obj)
                n += 1
        if n:
            removed_master[kind] = n

    batch['reverted'] = True
    batch['reverted_at'] = utc_now().isoformat()
    db.session.commit()
    _save_import_batches(batches)
    log_change(
        f"Reverted schedule import batch {batch['id']} "
        f"({removed_trips} trips, {removed_waves} waves removed)", 'trip')
    db.session.commit()
    return jsonify({'ok': True, 'batch_id': batch['id'],
                    'removed_trips': removed_trips,
                    'removed_waves': removed_waves,
                    'removed_master': removed_master})


# ── DASHBOARD KPI REFRESH API ──────────────────────────────────────────────
@app.route('/api/dashboard/kpis')
@login_required
def api_dashboard_kpis():
    """Lightweight endpoint for live KPI card refresh (called by Firebase listener)."""
    from_date = request.args.get('trend_start', (ph_today() - timedelta(days=13)).isoformat())
    to_date   = request.args.get('trend_end',   ph_today().isoformat())
    truck     = request.args.get('truck',  'all')
    status    = request.args.get('status', 'all')
    from_d = parse_date(from_date)
    to_d   = parse_date(to_date)
    q = (db.session.query(TripRecord).join(Wave)
         .filter(Wave.date >= from_d, Wave.date <= to_d))
    if truck != 'all':
        tt = TruckTypeDef.query.filter_by(code=truck).first()
        if tt:
            q = q.filter(Wave.truck_type_id == tt.id)
    if status != 'all':
        q = q.filter(TripRecord.status == status)
    trips = q.all()
    # Total breakdown hours within range (uses started_at)
    bd_from_dt = datetime.combine(from_d, datetime.min.time())
    bd_to_dt   = datetime.combine(to_d,   datetime.max.time())
    bd_logs = (db.session.query(BreakdownLog)
               .filter(BreakdownLog.started_at.isnot(None),
                       BreakdownLog.ended_at.isnot(None),
                       BreakdownLog.started_at >= bd_from_dt,
                       BreakdownLog.started_at <= bd_to_dt)
               .all())
    total_breakdown_hours = round(sum(b.duration_hours for b in bd_logs), 1)

    # GPS-detected toll today — separate KPI sourced from CartrackEvent.
    # This is what the polling worker auto-detected from plaza transits;
    # it is INDEPENDENT of TripRecord.toll_fee (which is dispatcher's
    # manual entry from physical receipts). Comparing the two helps
    # Finance spot exemptions, RFID misses, or GPS-missed plazas.
    from sqlalchemy import func as _func
    today_pht_d = ph_today()
    # Convert PHT calendar day to its UTC bounds for the CartrackEvent
    # query — events are stored in UTC.
    day_start_utc = datetime.combine(today_pht_d, datetime.min.time()).replace(tzinfo=PH_TZ).astimezone(UTC_TZ).replace(tzinfo=None)
    day_end_utc   = datetime.combine(today_pht_d, datetime.max.time()).replace(tzinfo=PH_TZ).astimezone(UTC_TZ).replace(tzinfo=None)
    gps_toll = (db.session.query(
                    _func.coalesce(_func.sum(CartrackEvent.toll_fee), 0.0),
                    _func.count(CartrackEvent.id),
                )
                .filter(CartrackEvent.event_type == 'trip_closed',
                        CartrackEvent.toll_fee.isnot(None),
                        CartrackEvent.created_at >= day_start_utc,
                        CartrackEvent.created_at <= day_end_utc)
                .first())
    gps_toll_total  = float(gps_toll[0] or 0)
    gps_toll_count  = int(gps_toll[1] or 0)

    return jsonify({
        'total':                 len(trips),
        'by_status':             {s: sum(1 for t in trips if t.status == s) for s in STATUSES},
        'total_toll_fee':        sum((t.toll_fee or 0) for t in trips if t.status != 'Canceled'),
        'total_breakdown_hours': total_breakdown_hours,
        # GPS-detected toll (independent of TripRecord.toll_fee)
        'gps_toll_total':        gps_toll_total,
        'gps_toll_count':        gps_toll_count,
    })


# ── FLEET UTILIZATION API ──────────────────────────────────────────────────
@app.route('/api/dashboard/fleet-utilization')
@login_required
def api_dashboard_fleet_utilization():
    """
    Compute fleet utilization % per truck type per day for the requested range.

    Scoring rules (configurable per truck type):
      - point_per_leg       = points awarded per delivered trip leg (0.5 for 10W)
      - daily_target_points = daily 100% threshold (4.0 for 10W, 1.5 for others)

    Special-case: trips whose product has is_full_day_trip = True count for the
    truck type's full daily_target_points instead of point_per_leg
    (e.g., ASPHALT runs).

    Active truck count per type = active plates assigned to that type.
    """
    from_date = request.args.get('trend_start',
                                 (ph_today() - timedelta(days=13)).isoformat())
    to_date   = request.args.get('trend_end', ph_today().isoformat())
    from_d = parse_date(from_date)
    to_d   = parse_date(to_date)
    if from_d > to_d:
        from_d, to_d = to_d, from_d

    truck_types = TruckTypeDef.query.order_by(TruckTypeDef.sort_order).all()

    # Active truck count per type (active plates only)
    plate_counts = dict(
        db.session.query(Plate.truck_type_id, db.func.count(Plate.id))
                  .filter(Plate.active == True, Plate.truck_type_id.isnot(None))
                  .group_by(Plate.truck_type_id).all()
    )

    # Fetch all delivered trips in range with their wave + product info
    trips = (db.session.query(TripRecord, Wave)
             .join(Wave, TripRecord.wave_id == Wave.id)
             .filter(Wave.date >= from_d, Wave.date <= to_d,
                     TripRecord.status == 'Delivered')
             .all())

    # Build product → is_full_day map (only for products that appear in the trips).
    # A product is "full-day" if EITHER:
    #   - its is_full_day_trip flag is True, OR
    #   - its name contains a FULL_DAY_KEYWORDS substring (e.g., "asphalt",
    #     case-insensitive — covers "ASPHALT", "Asphalt Plant", "Asphalt Mix"...).
    product_ids = {t.product_id for (t, _w) in trips if t.product_id}
    full_day_products = set()
    if product_ids:
        for prod in Product.query.filter(Product.id.in_(product_ids)).all():
            if is_product_full_day(prod):
                full_day_products.add(prod.id)

    # Build day list
    days = []
    cur = from_d
    while cur <= to_d:
        days.append(cur)
        cur += timedelta(days=1)

    # Aggregate points per (truck_type_id, day)
    points_map = {tt.id: {d: 0.0 for d in days} for tt in truck_types}
    for trip, wave in trips:
        if not wave or wave.truck_type_id not in points_map:
            continue
        tt = next((x for x in truck_types if x.id == wave.truck_type_id), None)
        if not tt:
            continue
        if trip.product_id and trip.product_id in full_day_products:
            pts = float(tt.daily_target_points or 1.5)
        else:
            pts = float(tt.point_per_leg or 1.0)
        if wave.date in points_map[tt.id]:
            points_map[tt.id][wave.date] += pts

    # Build series per truck type. OT (Others) is EXCLUDED: it holds
    # the hustling/hauling runs (and catch-all units), which operations
    # deliberately keeps out of fleet utilisation.
    series = []
    for tt in truck_types:
        if tt.code == 'OT':
            continue
        truck_count   = plate_counts.get(tt.id, 0)
        target_per_truck = float(tt.daily_target_points or 1.5)
        max_per_day   = truck_count * target_per_truck if truck_count > 0 else 0
        daily_pts     = [points_map[tt.id][d] for d in days]
        daily_util    = [(p / max_per_day * 100.0) if max_per_day > 0 else 0.0
                         for p in daily_pts]
        series.append({
            'code':              tt.code,
            'name':              tt.name,
            'color':             tt.color,
            'point_per_leg':     float(tt.point_per_leg or 1.0),
            'daily_target':      target_per_truck,
            'truck_count':       truck_count,
            'max_per_day':       max_per_day,
            'points':            [round(v, 2) for v in daily_pts],
            'util_pct':          [round(v, 2) for v in daily_util],
        })

    return jsonify({
        'days':        [d.strftime('%b %d') for d in days],
        'days_iso':    [iso_ph(d) for d in days],
        'series':      series,
    })


# ── BREAKDOWN HOURS KPI API ────────────────────────────────────────────────
@app.route('/api/dashboard/breakdown-hours')
@login_required
def api_dashboard_breakdown_hours():
    """
    Total accumulated breakdown hours within the date range, grouped by plate.

    Inclusion rule: a BreakdownLog counts toward the KPI if it has BOTH
    started_at and ended_at populated, and started_at falls within the range.
    Older records without timestamps are excluded (you can't compute hours).
    """
    from_date = request.args.get('trend_start',
                                 (ph_today() - timedelta(days=13)).isoformat())
    to_date   = request.args.get('trend_end', ph_today().isoformat())
    from_d = parse_date(from_date)
    to_d   = parse_date(to_date)
    if from_d > to_d:
        from_d, to_d = to_d, from_d

    from_dt = datetime.combine(from_d, datetime.min.time())
    to_dt   = datetime.combine(to_d,   datetime.max.time())

    rows = (db.session.query(BreakdownLog)
            .filter(BreakdownLog.started_at.isnot(None),
                    BreakdownLog.ended_at.isnot(None),
                    BreakdownLog.started_at >= from_dt,
                    BreakdownLog.started_at <= to_dt)
            .all())

    by_plate = {}
    for log in rows:
        if not log.plate_id:
            continue
        hrs = log.duration_hours
        if hrs <= 0:
            continue
        plate = log.plate
        key = log.plate_id
        if key not in by_plate:
            by_plate[key] = {
                'plate_id':       log.plate_id,
                'plate_display':  plate.display if plate else 'Unknown',
                'truck_type':     plate.truck_type.name if (plate and plate.truck_type) else 'Unassigned',
                'truck_color':    plate.truck_type.color if (plate and plate.truck_type) else '#999',
                'total_hours':    0.0,
                'job_orders':     [],
            }
        by_plate[key]['total_hours'] += hrs
        by_plate[key]['job_orders'].append({
            'id':            log.id,
            'started_at':    log.started_at.strftime('%Y-%m-%d %H:%M'),
            'ended_at':      log.ended_at.strftime('%Y-%m-%d %H:%M') if log.ended_at else '',
            'hours':         round(hrs, 2),
            'description':   log.description or '',
            'status':        log.status or '',
            'remarks':       log.remarks or '',
        })

    # Sort plates by total_hours desc
    plates_list = sorted(by_plate.values(),
                         key=lambda p: p['total_hours'], reverse=True)
    for p in plates_list:
        p['total_hours'] = round(p['total_hours'], 2)
        # Sort each plate's job orders by start time desc (most recent first)
        p['job_orders'].sort(key=lambda jo: jo['started_at'], reverse=True)

    total_hours = round(sum(p['total_hours'] for p in plates_list), 2)
    total_jobs  = sum(len(p['job_orders']) for p in plates_list)

    return jsonify({
        'total_hours':    total_hours,
        'total_jobs':     total_jobs,
        'plate_count':    len(plates_list),
        'by_plate':       plates_list,
    })


# ── DRIVER / TRUCK RATIO API ───────────────────────────────────────────────
@app.route('/api/dashboard/driver-truck-ratio')
@login_required
def api_dashboard_driver_truck_ratio():
    """
    Daily breakdown of working trucks vs present drivers.

    Working Trucks  = active plates of selected truck type(s) - plates currently
                      under repair on that day (BreakdownLog with status='Under
                      Repair' covering the date, i.e., date <= D AND
                      (resolved_date IS NULL OR resolved_date > D)).
    Present Drivers = WHEN no truck-type filter: distinct drivers with
                      attendance.status='Present' that day (system-wide).
                      WHEN filtered: distinct drivers who ACTUALLY drove a truck
                      of one of the selected types that day (from TripRecord,
                      excluding 'Canceled' trips). This works even if a driver
                      is qualified for multiple types — they're counted under
                      the type they actually used that day. The API also returns
                      per-day driver→truck assignments so the UI can show which
                      driver drove which unit.
    Coverage Ratio  = Present Drivers / Working Trucks (0 if no working trucks).
    Driver Shortage = max(0, Working Trucks - Present Drivers).

    Query params:
        trend_start, trend_end – date range (defaults to last 14 days)
        truck_types – comma-separated list of truck-type codes, e.g., '10W,12W'
                      (omit or empty = all)
    """
    from_date = request.args.get('trend_start',
                                 (ph_today() - timedelta(days=13)).isoformat())
    to_date   = request.args.get('trend_end', ph_today().isoformat())
    raw_codes = request.args.get('truck_types', '').strip()
    selected_codes = [c.strip() for c in raw_codes.split(',') if c.strip()]

    from_d = parse_date(from_date)
    to_d   = parse_date(to_date)
    if from_d > to_d:
        from_d, to_d = to_d, from_d

    # Resolve truck-type codes -> ids (skip unknown codes silently)
    selected_tt_ids = set()
    if selected_codes:
        selected_tt_ids = {
            tt.id for tt in TruckTypeDef.query.filter(
                TruckTypeDef.code.in_(selected_codes)).all()
        }

    # Day list
    days = []
    cur = from_d
    while cur <= to_d:
        days.append(cur)
        cur += timedelta(days=1)

    # Active plates (filtered by truck type if requested)
    plates_q = Plate.query.filter_by(active=True)
    if selected_tt_ids:
        plates_q = plates_q.filter(Plate.truck_type_id.in_(selected_tt_ids))
    active_plate_ids = {p.id for p in plates_q.all()}
    total_active_plates = len(active_plate_ids)

    # Pre-fetch breakdowns for ONLY plates we care about.
    #
    # Two kinds of rows matter for the per-day window check below:
    #   1. Still open (status='Under Repair', resolved_date NULL) —
    #      the plate is broken from `date` through today.
    #   2. Already fixed (status='Fixed' WITH a resolved_date) — the
    #      plate was broken from `date` until `resolved_date`; needed
    #      so HISTORICAL days in the trend still subtract the plate.
    #
    # Filtering on status='Under Repair' alone (the original query)
    # silently dropped kind 2: the moment FixFlo flipped a row to
    # Fixed, past trend days started counting that plate as working
    # during the very period it sat in the shop. Fixed-without-
    # resolved_date rows (legacy manual entries) are excluded — with
    # no end date they would subtract the plate forever.
    breakdown_q = db.session.query(
            BreakdownLog.plate_id, BreakdownLog.date, BreakdownLog.resolved_date
        ).filter(
            BreakdownLog.date <= to_d,
            db.or_(
                BreakdownLog.status == 'Under Repair',
                db.and_(BreakdownLog.status == 'Fixed',
                        BreakdownLog.resolved_date.isnot(None)),
            ))
    if active_plate_ids:
        breakdown_q = breakdown_q.filter(BreakdownLog.plate_id.in_(active_plate_ids))
    breakdowns = breakdown_q.all()

    # Drivers present per day:
    # If no filter: system-wide attendance count.
    # If filtered: drivers who actually drove a truck of selected type(s) that day.
    drivers_per_day = {}
    assignments_per_day = {}   # date -> [{driver, plate, body, type_code, type_color}, ...]

    if not selected_tt_ids:
        attn_rows = (db.session.query(Attendance.date,
                                      db.func.count(db.distinct(Attendance.driver_id)))
                     .filter(Attendance.status == 'Present',
                             Attendance.date >= from_d,
                             Attendance.date <= to_d)
                     .group_by(Attendance.date).all())
        drivers_per_day = {row[0]: row[1] for row in attn_rows}
    else:
        # Detect ACTUAL usage from TripRecords. A driver who is qualified for
        # both 10W and 12W but drove a 12W truck that day counts under 12W,
        # not 10W (we look at what they actually used).
        trip_rows = (db.session.query(
                        Wave.date,
                        TripRecord.driver_id,
                        Driver.name,
                        Plate.id, Plate.plate_no, Plate.body_no,
                        TruckTypeDef.code, TruckTypeDef.color, TruckTypeDef.name,
                    )
                     .join(TripRecord, TripRecord.wave_id == Wave.id)
                     .join(Driver,    Driver.id    == TripRecord.driver_id)
                     .outerjoin(Plate, Plate.id    == TripRecord.plate_id)
                     .join(TruckTypeDef, TruckTypeDef.id == Wave.truck_type_id)
                     .filter(Wave.date >= from_d, Wave.date <= to_d,
                             Wave.truck_type_id.in_(selected_tt_ids),
                             TripRecord.driver_id.isnot(None),
                             TripRecord.status != 'Canceled')
                     .distinct()
                     .all())
        # Build per-day distinct-driver count + assignment list
        per_day_drivers = {}
        for d_date, drv_id, drv_name, plate_id, plate_no, body_no, tt_code, tt_color, tt_name in trip_rows:
            per_day_drivers.setdefault(d_date, set()).add(drv_id)
            display = (f"{body_no} / {plate_no}" if body_no else plate_no) if plate_no else '—'
            assignments_per_day.setdefault(d_date, []).append({
                'driver':    drv_name,
                'plate':     display,
                'type_code': tt_code,
                'type_name': tt_name,
                'type_color': tt_color,
            })
        drivers_per_day = {d: len(s) for d, s in per_day_drivers.items()}

    working = []
    present = []
    ratios  = []
    shortages = []
    for d in days:
        # A breakdown is "active on D" if start_date <= D AND (resolved_date IS NULL OR resolved_date > D)
        broken_plate_ids = {
            b.plate_id for b in breakdowns
            if b.date <= d and (b.resolved_date is None or b.resolved_date > d)
        }
        wt = max(0, total_active_plates - len(broken_plate_ids))
        pd = drivers_per_day.get(d, 0)
        rt = round((pd / wt), 3) if wt > 0 else 0
        sh = max(0, wt - pd)
        working.append(wt)
        present.append(pd)
        ratios.append(rt)
        shortages.append(sh)

    return jsonify({
        'days':            [d.strftime('%b %d') for d in days],
        'days_iso':        [iso_ph(d) for d in days],
        'working_trucks':  working,
        'present_drivers': present,
        'coverage_ratio':  ratios,
        'driver_shortage': shortages,
        'total_plates':    total_active_plates,
        'selected_codes':  selected_codes,
        'driver_source':   'attendance' if not selected_tt_ids else 'trips',
        # Per-day driver→truck assignments (only when filtered).
        # Deduped: a driver who did multiple trips on the same truck shows once.
        'assignments':     {
            iso_ph(d): _dedupe_assignments(rows)
            for d, rows in assignments_per_day.items()
        },
    })


def _dedupe_assignments(rows):
    """Collapse identical (driver, plate) pairs that come from multiple trips."""
    seen = set()
    out = []
    for r in rows:
        key = (r['driver'], r['plate'], r['type_code'])
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


# ── DASHBOARD ──────────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    from collections import defaultdict
    filter_date  = request.args.get('date',  ph_today().isoformat())
    filter_truck = request.args.get('truck', 'all')
    filter_status= request.args.get('status','all')
    trend_end_str   = request.args.get('trend_end',   ph_today().isoformat())
    trend_start_str = request.args.get('trend_start', (ph_today() - timedelta(days=13)).isoformat())
    trend_end_d     = parse_date(trend_end_str)
    trend_start_d   = parse_date(trend_start_str)
    if trend_start_d > trend_end_d:
        trend_start_d, trend_end_d = trend_end_d, trend_start_d
    d = parse_date(filter_date)

    truck_types = TruckTypeDef.query.order_by(TruckTypeDef.sort_order).all()

    # Base query for the selected date range
    q = (db.session.query(TripRecord)
         .join(Wave)
         .filter(Wave.date >= trend_start_d, Wave.date <= trend_end_d))
    if filter_truck != 'all':
        tt = TruckTypeDef.query.filter_by(code=filter_truck).first()
        if tt:
            q = q.filter(Wave.truck_type_id == tt.id)
    if filter_status != 'all':
        q = q.filter(TripRecord.status == filter_status)

    trips = q.all()

    # Stats
    total          = len(trips)
    by_status      = {s: sum(1 for t in trips if t.status == s) for s in STATUSES}
    total_toll_fee = sum((t.toll_fee or 0) for t in trips if t.status != 'Canceled')

    # GPS-detected toll — sourced from CartrackEvent (independent of
    # TripRecord.toll_fee). Separate KPI so dispatchers and Finance can
    # compare what the polling worker saw vs the manual entries. Computed
    # over the SELECTED date range (same as every other KPI) so the card
    # and its trend delta stay consistent with the rest of the dashboard.
    from sqlalchemy import func as _func

    def _gps_toll_for(start_d, end_d):
        """(sum, count) of GPS-detected tolls for a PHT date range."""
        s = datetime.combine(start_d, datetime.min.time()).replace(
            tzinfo=PH_TZ).astimezone(UTC_TZ).replace(tzinfo=None)
        e = datetime.combine(end_d, datetime.max.time()).replace(
            tzinfo=PH_TZ).astimezone(UTC_TZ).replace(tzinfo=None)
        r = (db.session.query(
                 _func.coalesce(_func.sum(CartrackEvent.toll_fee), 0.0),
                 _func.count(CartrackEvent.id))
             .filter(CartrackEvent.event_type == 'trip_closed',
                     CartrackEvent.toll_fee.isnot(None),
                     CartrackEvent.created_at >= s,
                     CartrackEvent.created_at <= e)
             .first())
        return float(r[0] or 0), int(r[1] or 0)

    gps_toll_total, gps_toll_count = _gps_toll_for(trend_start_d, trend_end_d)
    # Breakdown hours within the trend range (accumulate completed J.O.s)
    _bd_from_dt = datetime.combine(trend_start_d, datetime.min.time())
    _bd_to_dt   = datetime.combine(trend_end_d,   datetime.max.time())
    _bd_logs = (db.session.query(BreakdownLog)
                .filter(BreakdownLog.started_at.isnot(None),
                        BreakdownLog.ended_at.isnot(None),
                        BreakdownLog.started_at >= _bd_from_dt,
                        BreakdownLog.started_at <= _bd_to_dt)
                .all())
    total_breakdown_hours = round(sum(b.duration_hours for b in _bd_logs), 1)

    # ── Trend deltas: compare each KPI to the previous equal-length period ──
    # e.g. a 14-day window compares against the 14 days immediately before it.
    period_len   = (trend_end_d - trend_start_d).days + 1
    prev_end_d   = trend_start_d - timedelta(days=1)
    prev_start_d = prev_end_d - timedelta(days=period_len - 1)

    _tt_id = None
    if filter_truck != 'all':
        _ttf = TruckTypeDef.query.filter_by(code=filter_truck).first()
        _tt_id = _ttf.id if _ttf else None

    # Aggregate in SQL instead of loading every previous-period TripRecord.
    _pq = (db.session.query(TripRecord.status,
                            db.func.count(TripRecord.id),
                            db.func.coalesce(db.func.sum(TripRecord.toll_fee), 0.0))
           .join(Wave)
           .filter(Wave.date >= prev_start_d, Wave.date <= prev_end_d))
    if _tt_id:
        _pq = _pq.filter(Wave.truck_type_id == _tt_id)
    _prev_rows     = _pq.group_by(TripRecord.status).all()
    prev_total     = sum(r[1] for r in _prev_rows)
    prev_by_status = {s: 0 for s in STATUSES}
    prev_toll_fee  = 0.0
    for st, cnt, toll in _prev_rows:
        if st in prev_by_status:
            prev_by_status[st] = cnt
        if st != 'Canceled':          # NULL-status rows count too
            prev_toll_fee += float(toll or 0)
    prev_gps_total, _ = _gps_toll_for(prev_start_d, prev_end_d)

    _pbd = (db.session.query(BreakdownLog)
            .filter(BreakdownLog.started_at.isnot(None),
                    BreakdownLog.ended_at.isnot(None),
                    BreakdownLog.started_at >= datetime.combine(prev_start_d, datetime.min.time()),
                    BreakdownLog.started_at <= datetime.combine(prev_end_d, datetime.max.time()))
            .all())
    prev_breakdown_hours = round(sum(b.duration_hours for b in _pbd), 1)

    def _delta(cur, prev):
        """Percent change vs previous period, or None when there's no
        baseline to compare against (previous value is zero)."""
        if not prev:
            return None
        return round((cur - prev) / prev * 100, 1)

    # sentiment: is an INCREASE good, bad, or neutral? Drives the colour.
    kpi_deltas = {
        'Total Trips':     {'pct': _delta(total, prev_total),                        'sentiment': 'good_up'},
        'Delivered':       {'pct': _delta(by_status.get('Delivered', 0),  prev_by_status.get('Delivered', 0)),  'sentiment': 'good_up'},
        'In Transit':      {'pct': _delta(by_status.get('In Transit', 0), prev_by_status.get('In Transit', 0)), 'sentiment': 'neutral'},
        'Toll Fee':        {'pct': _delta(total_toll_fee, prev_toll_fee),            'sentiment': 'neutral'},
        'GPS Toll':        {'pct': _delta(gps_toll_total, prev_gps_total),           'sentiment': 'neutral'},
        'Breakdown Hours': {'pct': _delta(total_breakdown_hours, prev_breakdown_hours), 'sentiment': 'bad_up'},
        'Pending':         {'pct': _delta(by_status.get('Pending', 0),  prev_by_status.get('Pending', 0)),  'sentiment': 'bad_up'},
        'Canceled':        {'pct': _delta(by_status.get('Canceled', 0), prev_by_status.get('Canceled', 0)), 'sentiment': 'bad_up'},
    }
    prev_period_str = f"{prev_start_d.strftime('%b %d')} – {prev_end_d.strftime('%b %d')}"

    by_truck    = {}
    for tt in truck_types:
        cnt = sum(1 for t in trips
                  if t.wave and t.wave.truck_type_id == tt.id)
        by_truck[tt.code] = {'name': tt.name, 'color': tt.color, 'count': cnt}

    # Trend — one GROUP BY over the whole range instead of a COUNT query
    # per day per truck type (14 days x 8 types was ~112 round-trips; this
    # is a single query for the totals and one for the per-type series).
    day_list = []
    cur = trend_start_d
    while cur <= trend_end_d:
        day_list.append(cur)
        cur += timedelta(days=1)

    trend_days = [dd.strftime('%b %d') for dd in day_list]

    _tot_rows = (db.session.query(Wave.date, db.func.count(TripRecord.id))
                 .join(TripRecord, TripRecord.wave_id == Wave.id)
                 .filter(Wave.date >= trend_start_d, Wave.date <= trend_end_d)
                 .group_by(Wave.date).all())
    _tot_by_day = {r[0]: r[1] for r in _tot_rows}
    trend_counts = [_tot_by_day.get(dd, 0) for dd in day_list]

    _tt_rows = (db.session.query(Wave.truck_type_id, Wave.date,
                                 db.func.count(TripRecord.id))
                .join(TripRecord, TripRecord.wave_id == Wave.id)
                .filter(Wave.date >= trend_start_d, Wave.date <= trend_end_d)
                .group_by(Wave.truck_type_id, Wave.date).all())
    _tt_by_day = defaultdict(dict)
    for tt_id, dd, cnt in _tt_rows:
        _tt_by_day[tt_id][dd] = cnt

    trend_by_truck = []
    for tt in truck_types:
        per_day = _tt_by_day.get(tt.id, {})
        trend_by_truck.append({
            'code': tt.code, 'name': tt.name, 'color': tt.color,
            'counts': [per_day.get(dd, 0) for dd in day_list],
        })

    # Recent changes
    recent_changes = (ChangeLog.query.order_by(ChangeLog.timestamp.desc()).limit(8).all())

    # Top drivers by delivered trips — selected date range
    _drv = defaultdict(lambda: defaultdict(lambda: {'name': '', 'total': 0, 'delivered': 0}))
    all_tr = (db.session.query(TripRecord).join(Wave)
              .filter(TripRecord.driver_id.isnot(None),
                      Wave.date >= trend_start_d, Wave.date <= trend_end_d).all())
    for t in all_tr:
        if t.wave and t.driver:
            s = _drv[t.wave.truck_type_id][t.driver_id]
            s['name'] = t.driver.name
            s['total'] += 1
            if t.status == 'Delivered':
                s['delivered'] += 1

    top_drivers_by_truck = []
    for tt in truck_types:
        drivers = sorted(_drv.get(tt.id, {}).values(),
                         key=lambda x: x['delivered'], reverse=True)[:5]
        drivers = [d for d in drivers if d['total'] > 0 and d['name']]
        if drivers:
            top_drivers_by_truck.append({
                'truck': tt.name, 'code': tt.code,
                'color': tt.color, 'drivers': drivers
            })

    # Top absent drivers — selected date range
    absent_stats = (db.session.query(
                        Driver.name,
                        db.func.count(Attendance.id).label('cnt')
                    )
                    .join(Attendance, Attendance.driver_id == Driver.id)
                    .filter(Attendance.status == 'Absent',
                            Attendance.date >= trend_start_d,
                            Attendance.date <= trend_end_d)
                    .group_by(Driver.id, Driver.name)
                    .order_by(db.func.count(Attendance.id).desc())
                    .limit(10)
                    .all())
    absent_drivers = [{'name': r[0], 'absences': r[1]} for r in absent_stats]

    return render_template('dashboard_v2.html',
        d=d, filter_date=filter_date,
        filter_truck=filter_truck, filter_status=filter_status,
        trend_start_str=trend_start_str, trend_end_str=trend_end_str,
        truck_types=truck_types, trips=trips,
        total=total, by_status=by_status, by_truck=by_truck, total_toll_fee=total_toll_fee,
        total_breakdown_hours=total_breakdown_hours,
        gps_toll_total=gps_toll_total, gps_toll_count=gps_toll_count,
        kpi_deltas=kpi_deltas, prev_period_str=prev_period_str,
        trend_days=trend_days, trend_counts=trend_counts,
        trend_by_truck=trend_by_truck,
        recent_changes=recent_changes,
        top_drivers_by_truck=top_drivers_by_truck,
        absent_drivers=absent_drivers)


# ── MASTER DATA ────────────────────────────────────────────────────────────
@app.route('/master')
@login_required
def master():
    truck_types = TruckTypeDef.query.order_by(TruckTypeDef.sort_order).all()
    drivers     = Driver.query.order_by(Driver.name).all()
    helpers     = Helper.query.order_by(Helper.name).all()
    products    = Product.query.order_by(Product.name).all()
    clients     = Client.query.order_by(Client.name).all()
    dispatchers = Dispatcher.query.order_by(Dispatcher.name).all()
    plates      = Plate.query.order_by(Plate.truck_type_id, Plate.plate_no).all()
    return render_template('master/index.html',
        truck_types=truck_types,
        drivers=drivers, helpers=helpers, products=products,
        clients=clients, dispatchers=dispatchers, plates=plates,
        full_day_keywords=FULL_DAY_KEYWORDS)


@app.route('/api/master/<category>/add', methods=['POST'])
@login_required
def api_master_add(category):
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Name required'}), 400

    model_map = {'drivers': Driver, 'helpers': Helper,
                 'products': Product, 'clients': Client,
                 'dispatchers': Dispatcher}
    if category == 'plates':
        plate_no = name
        body_no  = (data.get('body_no') or '').strip()
        ttid     = data.get('truck_type_id') or None
        if ttid: ttid = int(ttid)
        # Validate toll_class on insert — defaults to Class 3 (heavy trucks)
        # if not provided. Accepts 'Class 1', 'Class 2', or 'Class 3' only;
        # silently coerces anything else to Class 3 so a typo doesn't end
        # up in the DB.
        toll_class = (data.get('toll_class') or 'Class 3').strip()
        if toll_class not in ('Class 1', 'Class 2', 'Class 3'):
            toll_class = 'Class 3'
        obj = Plate(plate_no=plate_no, body_no=body_no or None,
                    truck_type_id=ttid, toll_class=toll_class)
        db.session.add(obj)
        db.session.commit()
        log_change(f"Added plate {obj.display}", 'master')
        db.session.commit()
        tt = TruckTypeDef.query.get(ttid) if ttid else None
        return jsonify({'id': obj.id, 'display': obj.display,
                        'truck_type': tt.name if tt else '',
                        'toll_class': obj.toll_class})

    Model = model_map.get(category)
    if not Model:
        return jsonify({'error': 'Unknown category'}), 400
    obj = Model(name=name)
    # Drivers can be qualified for multiple truck types
    if category == 'drivers':
        tt_ids = data.get('truck_type_ids') or []
        if isinstance(tt_ids, list) and tt_ids:
            tts = TruckTypeDef.query.filter(TruckTypeDef.id.in_(tt_ids)).all()
            obj.truck_types = tts
    db.session.add(obj)
    db.session.commit()
    log_change(f"Added {category[:-1]} '{name}'", 'master')
    db.session.commit()
    resp = {'id': obj.id, 'name': obj.name}
    if category == 'drivers':
        resp['truck_type_ids'] = [t.id for t in obj.truck_types]
    return jsonify(resp)


@app.route('/api/master/<category>/<int:item_id>/update', methods=['POST'])
@login_required
def api_master_update(category, item_id):
    data = request.get_json()
    model_map = {'drivers': Driver, 'helpers': Helper,
                 'products': Product, 'clients': Client,
                 'dispatchers': Dispatcher, 'plates': Plate}
    Model = model_map.get(category)
    if not Model:
        return jsonify({'error': 'Unknown'}), 400
    obj = Model.query.get_or_404(item_id)
    if category == 'plates':
        obj.plate_no = (data.get('plate_no') or obj.plate_no).strip()
        obj.body_no  = (data.get('body_no') or '').strip() or None
        ttid = data.get('truck_type_id') or None
        obj.truck_type_id = int(ttid) if ttid else None
        # toll_class — only update if the client sent the field, so
        # legacy edit calls that don't include it preserve the existing
        # class. Same validation as the add endpoint.
        if 'toll_class' in data:
            new_class = (data.get('toll_class') or '').strip()
            if new_class in ('Class 1', 'Class 2', 'Class 3'):
                obj.toll_class = new_class
    else:
        obj.name = (data.get('name') or obj.name).strip()
        # Products: optional is_full_day_trip toggle
        if category == 'products' and 'is_full_day_trip' in data:
            obj.is_full_day_trip = bool(data.get('is_full_day_trip'))
        # Drivers: optional list of truck-type categories (many-to-many)
        if category == 'drivers' and 'truck_type_ids' in data:
            tt_ids = data.get('truck_type_ids') or []
            if isinstance(tt_ids, list):
                tts = (TruckTypeDef.query.filter(TruckTypeDef.id.in_(tt_ids)).all()
                       if tt_ids else [])
                obj.truck_types = tts
                # Keep legacy single-category column in sync (use first selection
                # as the "primary" — or NULL when none selected)
                obj.truck_type_id = tts[0].id if tts else None
    db.session.commit()
    log_change(f"Updated {category[:-1]} id={item_id}", 'master')
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/master/truck-type/<int:tt_id>/update', methods=['POST'])
@login_required
def api_truck_type_update(tt_id):
    """Update fleet-utilization scoring fields for a truck type."""
    data = request.get_json() or {}
    tt = TruckTypeDef.query.get_or_404(tt_id)
    if 'point_per_leg' in data:
        try:
            tt.point_per_leg = float(data['point_per_leg'])
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid point_per_leg'}), 400
    if 'daily_target_points' in data:
        try:
            tt.daily_target_points = float(data['daily_target_points'])
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid daily_target_points'}), 400
    db.session.commit()
    log_change(f"Updated truck type {tt.code} scoring (pts/leg={tt.point_per_leg}, target={tt.daily_target_points})", 'master')
    db.session.commit()
    return jsonify({'ok': True, 'point_per_leg': tt.point_per_leg,
                    'daily_target_points': tt.daily_target_points})


@app.route('/api/master/<category>/<int:item_id>/toggle', methods=['POST'])
@login_required
def api_master_toggle(category, item_id):
    if not check_can_delete():
        return jsonify({'error': 'You do not have permission to deactivate records.'}), 403
    model_map = {'drivers': Driver, 'helpers': Helper,
                 'products': Product, 'clients': Client,
                 'dispatchers': Dispatcher, 'plates': Plate}
    Model = model_map.get(category)
    if not Model:
        return jsonify({'error': 'Unknown'}), 400
    obj = Model.query.get_or_404(item_id)
    obj.active = not obj.active
    db.session.commit()
    return jsonify({'active': obj.active})


# ── REPORTS ────────────────────────────────────────────────────────────────
@app.route('/reports')
@login_required
def reports():
    year        = request.args.get('year',  ph_today().year,  type=int)
    month       = request.args.get('month', ph_today().month, type=int)
    filter_truck= request.args.get('truck', 'all')
    truck_types = TruckTypeDef.query.order_by(TruckTypeDef.sort_order).all()

    last_day    = calendar.monthrange(year, month)[1]
    mo_s        = date(year, month, 1)
    mo_e        = date(year, month, last_day)

    q = (db.session.query(TripRecord).join(Wave)
         .filter(Wave.date >= mo_s, Wave.date <= mo_e))
    if filter_truck != 'all':
        tt = TruckTypeDef.query.filter_by(code=filter_truck).first()
        if tt: q = q.filter(Wave.truck_type_id == tt.id)
    trips = q.order_by(Wave.date, Wave.wave_number, TripRecord.trip_number).all()

    by_status = {s: sum(1 for t in trips if t.status == s) for s in STATUSES}
    years     = list(range(2024, ph_today().year + 2))

    return render_template('reports/index.html',
        year=year, month=month, years=years,
        filter_truck=filter_truck, truck_types=truck_types,
        trips=trips, by_status=by_status, mo_s=mo_s)


@app.route('/reports/export')
@login_required
def export():
    year        = request.args.get('year',  ph_today().year,  type=int)
    month       = request.args.get('month', ph_today().month, type=int)
    filter_truck= request.args.get('truck', 'all')
    last_day    = calendar.monthrange(year, month)[1]
    mo_s        = date(year, month, 1)
    mo_e        = date(year, month, last_day)
    truck_types = TruckTypeDef.query.order_by(TruckTypeDef.sort_order).all()

    MAROON  = '8B1A2B'; WHITE = 'FFFFFF'; LGRAY = 'F5F0F0'
    GRN     = 'C6EFCE'; RED   = 'FFB3B3'; YEL   = 'FFF2CC'; BLU   = 'DDEEFF'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    def hdr_style(cell, bg=MAROON, fg=WHITE, bold=True, sz=10):
        cell.font  = Font(name='Calibri', bold=bold, color=fg, size=sz)
        cell.fill  = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def data_style(cell, bg=None):
        cell.font = Font(name='Calibri', size=9)
        cell.alignment = Alignment(vertical='center')
        if bg: cell.fill = PatternFill('solid', fgColor=bg)

    status_colors = {'Delivered': GRN, 'In Transit': YEL, 'Loading': BLU, 'Pending': None}
    col_hdrs = ['Date','Truck Type','Wave','#','Driver','Helper','Plate',
                'Product','Client','Dispatcher','RS No','PO No','Reference',
                'DR No','Volume','Status','Notes','Updated By']

    sheets_to_create = []
    if filter_truck == 'all':
        for tt in truck_types: sheets_to_create.append(tt)
        sheets_to_create.append(None)
    else:
        tt = TruckTypeDef.query.filter_by(code=filter_truck).first()
        if tt: sheets_to_create.append(tt)

    for tt_obj in sheets_to_create:
        sheet_name = tt_obj.name[:28] if tt_obj else 'All Trucks'
        ws = wb.create_sheet(title=sheet_name)

        ws.merge_cells(f'A1:{openpyxl.utils.get_column_letter(len(col_hdrs))}1')
        title_cell = ws['A1']
        mo_label   = mo_s.strftime('%B %Y')
        title_cell.value = f"DISPATCH SCHEDULE — {tt_obj.name if tt_obj else 'ALL TRUCKS'} — {mo_label}"
        hdr_style(title_cell, sz=13)
        ws.row_dimensions[1].height = 26

        for ci, h in enumerate(col_hdrs, 1):
            c = ws.cell(row=2, column=ci, value=h)
            hdr_style(c, bg='5D0E1B', sz=9)
        ws.row_dimensions[2].height = 22

        q2 = (db.session.query(TripRecord).join(Wave)
              .filter(Wave.date >= mo_s, Wave.date <= mo_e))
        if tt_obj: q2 = q2.filter(Wave.truck_type_id == tt_obj.id)
        trip_rows = q2.order_by(Wave.date, Wave.wave_number, TripRecord.trip_number).all()

        for ri, t in enumerate(trip_rows, 3):
            bg = status_colors.get(t.status or '', None)
            vals = [
                t.wave.date.strftime('%Y-%m-%d') if t.wave else '',
                t.wave.truck_type.name if t.wave else '',
                t.wave.label if t.wave else '',
                t.trip_number,
                t.driver.name if t.driver else '',
                t.helper.name if t.helper else '',
                t.plate.display if t.plate else '',
                t.product.name if t.product else '',
                t.client.name if t.client else '',
                t.dispatcher.name if t.dispatcher else '',
                t.rs_no or '', t.po_no or '', t.reference or '',
                t.dr_no or '', t.volume or '',
                t.status or '', t.notes or '',
                t.updated_by or '',
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                data_style(c, bg=bg)
            if ri % 2 == 0 and not bg:
                for ci in range(1, len(col_hdrs)+1):
                    ws.cell(row=ri, column=ci).fill = PatternFill('solid', fgColor=LGRAY)

        col_widths = [12,18,10,4,16,16,14,18,18,14,10,10,10,10,8,12,20,14]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"Dispatch_{year}_{month:02d}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── ATTENDANCE ─────────────────────────────────────────────────────────────
@app.route('/attendance')
@login_required
def attendance():
    year   = request.args.get('year',  ph_today().year,  type=int)
    month  = request.args.get('month', ph_today().month, type=int)
    years  = list(range(2024, ph_today().year + 2))

    last_day = calendar.monthrange(year, month)[1]
    mo_s     = date(year, month, 1)
    mo_e     = date(year, month, last_day)
    days     = list(range(1, last_day + 1))

    drivers = Driver.query.filter_by(active=True).order_by(Driver.name).all()

    # Build attendance map: {(driver_id, day): status}
    records = (Attendance.query
               .filter(Attendance.date >= mo_s, Attendance.date <= mo_e)
               .all())
    att_map = {}
    for r in records:
        att_map[(r.driver_id, r.date.day)] = r.status

    # Summary per driver
    summary = {}
    for drv in drivers:
        summary[drv.id] = {s: 0 for s in ATTENDANCE_STATUSES}
        for day in days:
            st = att_map.get((drv.id, day))
            if st and st in summary[drv.id]:
                summary[drv.id][st] += 1

    return render_template('attendance/index.html',
        year=year, month=month, years=years,
        mo_s=mo_s, days=days,
        drivers=drivers, att_map=att_map, summary=summary,
        att_statuses=ATTENDANCE_STATUSES)


@app.route('/api/attendance/set', methods=['POST'])
@login_required
def api_attendance_set():
    data      = request.get_json()
    driver_id = data.get('driver_id')
    date_str  = data.get('date')
    status    = data.get('status')  # None = clear

    if not driver_id or not date_str:
        return jsonify({'error': 'Missing fields'}), 400

    d = parse_date(date_str)
    record = Attendance.query.filter_by(driver_id=driver_id, date=d).first()

    if status is None or status == '':
        # Clear the record
        if record:
            db.session.delete(record)
            db.session.commit()
        return jsonify({'status': '', 'date': date_str, 'driver_id': driver_id})

    if not record:
        record = Attendance(driver_id=driver_id, date=d)
        db.session.add(record)

    record.status     = status
    record.updated_by = get_user()
    record.updated_at = utc_now()
    db.session.commit()

    drv = Driver.query.get(driver_id)
    log_change(f"Attendance {drv.name if drv else driver_id}: {status} on {d}", 'attendance')
    db.session.commit()

    return jsonify({'status': status, 'date': date_str, 'driver_id': driver_id})


# ── HELPER ATTENDANCE ──────────────────────────────────────────────────────
@app.route('/helper-attendance')
@login_required
def helper_attendance():
    year   = request.args.get('year',  ph_today().year,  type=int)
    month  = request.args.get('month', ph_today().month, type=int)
    years  = list(range(2024, ph_today().year + 2))

    last_day = calendar.monthrange(year, month)[1]
    mo_s     = date(year, month, 1)
    mo_e     = date(year, month, last_day)
    days     = list(range(1, last_day + 1))

    helpers = Helper.query.filter_by(active=True).order_by(Helper.name).all()

    records = (HelperAttendance.query
               .filter(HelperAttendance.date >= mo_s, HelperAttendance.date <= mo_e)
               .all())
    att_map = {}
    for r in records:
        att_map[(r.helper_id, r.date.day)] = r.status

    summary = {}
    for hlp in helpers:
        summary[hlp.id] = {s: 0 for s in ATTENDANCE_STATUSES}
        for day in days:
            st = att_map.get((hlp.id, day))
            if st and st in summary[hlp.id]:
                summary[hlp.id][st] += 1

    return render_template('attendance/helper_index.html',
        year=year, month=month, years=years,
        mo_s=mo_s, days=days,
        helpers=helpers, att_map=att_map, summary=summary,
        att_statuses=ATTENDANCE_STATUSES)


@app.route('/api/helper-attendance/set', methods=['POST'])
@login_required
def api_helper_attendance_set():
    data      = request.get_json()
    helper_id = data.get('helper_id')
    date_str  = data.get('date')
    status    = data.get('status')

    if not helper_id or not date_str:
        return jsonify({'error': 'Missing fields'}), 400

    d = parse_date(date_str)
    record = HelperAttendance.query.filter_by(helper_id=helper_id, date=d).first()

    if status is None or status == '':
        if record:
            db.session.delete(record)
            db.session.commit()
        return jsonify({'status': '', 'date': date_str, 'helper_id': helper_id})

    if not record:
        record = HelperAttendance(helper_id=helper_id, date=d)
        db.session.add(record)

    record.status     = status
    record.updated_by = get_user()
    record.updated_at = utc_now()
    db.session.commit()

    hlp = Helper.query.get(helper_id)
    log_change(f"Helper Attendance {hlp.name if hlp else helper_id}: {status} on {d}", 'attendance')
    db.session.commit()

    return jsonify({'status': status, 'date': date_str, 'helper_id': helper_id})


# ── BREAKDOWN ──────────────────────────────────────────────────────────────
# ── Driver-name normalisation (for the Breakdowns-by-Driver chart) ──
# FixFlo operators are typed freeform — same person ends up as
# 'JIM LAYAG', 'J.LAYAG', 'Jim layag', or even 'DANILO MIASCO TH10
# NHA 3948' (with the plate appended by accident). We derive a
# stable key per name so the chart aggregates spellings instead of
# fanning them out into one bar each.
#
# Heuristic: strip trailing plate-like patterns and common name
# suffixes, then build a key of '<first-initial>.<surname>'. This
# correctly merges 'JIM LAYAG' + 'J.LAYAG' (both J.LAYAG) and
# separates 'R.JARO' from 'R.DELA CRUZ' (different surnames). False
# merges are rare in a ~30-driver fleet; if they do happen, the
# tooltip exposes the merged aliases so the dispatcher can spot it.

_DRIVER_NAME_SKIP = {
    'NO OPERATOR', 'N/A', 'NA', 'NONE',
    'UNASSIGNED', 'TBA', 'TBD', '-',
}
_DRIVER_SUFFIX_TOKENS = {
    'JR', 'JR.', 'SR', 'SR.', 'JUNIOR', 'SENIOR',
    'II', 'III', 'IV',
}
# Matches trailing plate-like garbage on a name. Two shapes:
#   ' TH10 NHA 3948'  (truck-type code, plate prefix, plate number)
#   ' DT4 NET2693'    (same with prefix glued to number)
# Anchored to the end of the string with leading whitespace so a
# real first-name token like 'TH...' can't accidentally match (no
# real first name starts with 2+ uppercase letters followed by
# digits).
_DRIVER_PLATE_TAIL_RE = re.compile(
    r'\s+[A-Z]{2,4}\d{1,3}'        # truck-type code, e.g. TH10 / DT4
    r'(\s+[A-Z]{3}\s*\d{3,4})?'    # optional plate, e.g. NHA 3948
    r'\s*$',
    flags=re.IGNORECASE,
)


def _normalize_driver_key(name):
    """Return a canonical key for grouping driver name variants.

    Examples:
      'Judemar Salonga'                  -> 'J.SALONGA'
      'JUDEMAR SALONGA'                  -> 'J.SALONGA'
      'J.SALONGA'                        -> 'J.SALONGA'
      'JIM LAYAG'                        -> 'J.LAYAG'
      'J.LAYAG'                          -> 'J.LAYAG'
      'DANILO MIASCO TH10 NHA 3948'      -> 'D.MIASCO'
      'REYNALDO DELA CRUZ'               -> 'R.CRUZ'
      'No Operator'                      -> None  (skipped)

    Returns None when the input is empty or matches a placeholder
    in _DRIVER_NAME_SKIP — those should be excluded from the chart
    rather than rendered as an 'Unknown' bucket.
    """
    if not name:
        return None
    s = name.strip().upper()
    if not s or s in _DRIVER_NAME_SKIP:
        return None
    # Strip trailing plate-like tokens.
    s = _DRIVER_PLATE_TAIL_RE.sub('', s).strip()
    # Tokenise on whitespace, dots, commas. Drop empties and
    # name suffixes (JR, SR, III, ...).
    parts = [p for p in re.split(r'[\s.,]+', s)
             if p and p not in _DRIVER_SUFFIX_TOKENS]
    if not parts:
        return None
    # First-letter of the first word as the initial; last word as
    # the surname. Single-letter first tokens (e.g. the 'J' in
    # 'J.LAYAG') already collapse to themselves.
    surname = parts[-1]
    initial = parts[0][0] if parts[0] else ''
    if not initial or not surname:
        return None
    return f'{initial}.{surname}'


@app.route('/breakdown')
@login_required
def breakdown():
    year   = request.args.get('year',  ph_today().year,  type=int)
    month  = request.args.get('month', ph_today().month, type=int)
    filter_status = request.args.get('status', 'all')
    # Plate filter — value comes in as a string ID. Kept as string so
    # the template's selected check works uniformly (Plate.id|string vs
    # the querystring value). 'all' means no plate filter.
    filter_plate = request.args.get('plate', 'all')
    years  = list(range(2024, ph_today().year + 2))

    last_day = calendar.monthrange(year, month)[1]
    mo_s     = date(year, month, 1)
    mo_e     = date(year, month, last_day)

    q = BreakdownLog.query.filter(
        BreakdownLog.date >= mo_s,
        BreakdownLog.date <= mo_e
    )
    if filter_status != 'all':
        q = q.filter(BreakdownLog.status == filter_status)
    # Plate filter — coerce to int defensively so a stray non-numeric
    # value (manually edited URL) doesn't bubble up as a 500. Falls
    # back to no filter on bad input and resets filter_plate so the
    # dropdown shows 'All Plates' rather than a corrupt selection.
    if filter_plate != 'all':
        try:
            q = q.filter(BreakdownLog.plate_id == int(filter_plate))
        except (TypeError, ValueError):
            filter_plate = 'all'

    logs = q.order_by(BreakdownLog.date.desc(), BreakdownLog.id.desc()).all()

    plates      = Plate.query.filter_by(active=True).order_by(Plate.plate_no).all()
    truck_types = TruckTypeDef.query.order_by(TruckTypeDef.sort_order).all()

    # Summary counts
    all_logs_month = BreakdownLog.query.filter(
        BreakdownLog.date >= mo_s,
        BreakdownLog.date <= mo_e
    ).all()
    summary = {s: sum(1 for l in all_logs_month if l.status == s) for s in BREAKDOWN_STATUSES}
    summary['Total'] = len(all_logs_month)

    # Breakdowns by Plate — bar-chart data. Ranks every unit that
    # appears in the CURRENT filter window (year + month + status) by
    # breakdown-record count, descending. Drives the Chart.js bar
    # chart below the filter bar. Empty list when nothing matches.
    #
    # Two sources are merged into one ranking:
    #   1. Plate-matched rows  — keyed by plate_id, resolved via the
    #      preloaded `plates` lookup with a direct query fallback so
    #      retired (inactive) plates with rows in the window still
    #      appear instead of vanishing silently.
    #   2. Unmapped equipment  — rows where plate_id is null but the
    #      FixFlo sync captured an equipment.name (typically trailers
    #      and generators not registered as plates). Keyed by the
    #      equipment_name string itself so each unique label gets its
    #      own bar. Flagged is_equipment=True so the frontend can
    #      visually distinguish these from real plate bars.
    from collections import Counter
    plate_counts = Counter()
    equipment_counts = Counter()
    for l in logs:
        if l.plate_id is not None:
            plate_counts[l.plate_id] += 1
        elif getattr(l, 'equipment_name', None):
            equipment_counts[l.equipment_name] += 1

    plate_lookup = {p.id: p for p in plates}
    plate_breakdown_chart = []
    for plate_id, count in plate_counts.most_common():
        p = plate_lookup.get(plate_id) or Plate.query.get(plate_id)
        if p:
            plate_breakdown_chart.append({
                'label':        p.body_no or p.plate_no,
                'count':        count,
                'is_equipment': False,
                # plate_id surfaced so the frontend bar-click handler
                # can hit the unit-logs endpoint with the right key.
                # Stays absent for equipment rows (those identify by
                # name string, not id).
                'plate_id':     p.id,
            })
    for eq_name, count in equipment_counts.most_common():
        plate_breakdown_chart.append({
            'label':        eq_name,
            'count':        count,
            'is_equipment': True,
        })
    # Final re-sort so plates and equipment interleave correctly by
    # count, not by source group. Stable sort preserves the original
    # within-group order on ties (a plate and an equipment with the
    # same count stay in their relative order from the two pre-sorted
    # most_common() lists).
    plate_breakdown_chart.sort(key=lambda r: -r['count'])

    # Breakdowns by Driver — parallel chart, same filter window. Keyed
    # by operator_name from the FixFlo job order ('Operator Name' on
    # the request form, screenshot 06/2026). Helps spot drivers most
    # often associated with breakdown reports.
    #
    # Names from FixFlo are messy — the same driver appears as
    # 'JIM LAYAG', 'J.LAYAG', 'Judemar Salonga', 'JUDEMAR SALONGA',
    # 'J.SALONGA', sometimes with trailing plate junk like
    # 'DANILO MIASCO TH10 NHA 3948'. Without normalisation each
    # spelling forks into its own bar and the chart becomes useless
    # for finding chronic reporters.
    #
    # Strategy: derive a canonical key per name (first-initial +
    # surname) via _normalize_driver_key, group rows by that key,
    # display the longest variant as the chart label, and surface
    # the merged aliases in the tooltip so the dispatcher can audit
    # any false merges. Explicit placeholders like 'No Operator' /
    # 'N/A' are filtered out (they aren't real drivers).
    # Two-stage grouping:
    #   Stage 1 — exact-key merge via _normalize_driver_key
    #             ('J.SALONGA' / 'JUDEMAR SALONGA' / 'Judemar Salonga'
    #             all collapse to one key here).
    #   Stage 2 — fuzzy surname merge across keys with the same
    #             initial. Catches typos like 'R.HAGONOY' vs
    #             'Ronie Hagunoy' (surnames differ by one letter)
    #             that Stage 1 leaves on separate keys. Threshold
    #             tuned at 0.82 (SequenceMatcher) — empirically
    #             merges Hagonoy/Hagunoy without joining unrelated
    #             surnames in the test fleet.
    from difflib import SequenceMatcher
    SURNAME_SIMILARITY_THRESHOLD = 0.82

    operator_groups = {}   # key -> {'canonical': str, 'count': int, 'aliases': set[str]}
    for l in logs:
        nm = (getattr(l, 'operator_name', None) or '').strip()
        if not nm:
            continue
        key = _normalize_driver_key(nm)
        if key is None:
            continue
        # Stage 2 lookup: scan existing keys for a near-surname
        # match within the same initial bucket. If found, fold
        # this name into that existing group instead of creating
        # a new key. We deliberately compare against ORIGINAL keys
        # (not chained merges) so the threshold stays predictable.
        new_initial, _, new_surname = key.partition('.')
        merge_into = key
        for existing in operator_groups.keys():
            ex_initial, _, ex_surname = existing.partition('.')
            if ex_initial != new_initial:
                continue
            if ex_surname == new_surname:
                merge_into = existing
                break
            ratio = SequenceMatcher(None, ex_surname, new_surname).ratio()
            if ratio >= SURNAME_SIMILARITY_THRESHOLD:
                merge_into = existing
                break

        g = operator_groups.setdefault(merge_into, {
            'canonical': nm, 'count': 0, 'aliases': set(),
        })
        g['count'] += 1
        g['aliases'].add(nm)
        # Pick the longest spelling as the canonical display — it
        # almost always carries the most information (full first
        # name beats an initial). Ties keep the first one seen.
        if len(nm) > len(g['canonical']):
            g['canonical'] = nm

    operator_breakdown_chart = []
    for key, g in sorted(operator_groups.items(),
                          key=lambda kv: (-kv[1]['count'], kv[0])):
        aliases = sorted(g['aliases'])
        operator_breakdown_chart.append({
            'label':    g['canonical'],
            'count':    g['count'],
            # Always expose the full spelling list — the frontend
            # needs it to drive the click-to-modal lookup (the
            # endpoint filters by operator_name IN <aliases>).
            # Tooltip code only renders the 'Merged spellings'
            # block when length > 1.
            'aliases':  aliases,
        })

    return render_template('breakdown/index.html',
        year=year, month=month, years=years, mo_s=mo_s,
        logs=logs, plates=plates, truck_types=truck_types,
        filter_status=filter_status, filter_plate=filter_plate,
        bd_statuses=BREAKDOWN_STATUSES, summary=summary,
        plate_breakdown_chart=plate_breakdown_chart,
        operator_breakdown_chart=operator_breakdown_chart)


# ── PRINTABLE REPORTS ──────────────────────────────────────────────────────
# JO descriptions in FixFlo are free text ("Flat tire (rear left)",
# "Vulcanize tire", "CHANGE TIRE"...), so grouping by exact text is no
# summary at all. Each description is bucketed into a repair CATEGORY
# by keyword instead — first match wins, so more specific phrases come
# before generic ones. Edit this list to match shop vocabulary; anything
# unmatched lands in "Others".
_JO_CATEGORIES = [
    ('Tires',                  ['tire', 'gulong', 'vulcaniz', 'flat',
                                'interchange']),
    ('Change Oil / PMS',       ['change oil', 'pms', 'preventive',
                                'oil filter', 'lubric', 'grease']),
    ('Brakes',                 ['brake', 'preno']),
    ('Aircon',                 ['aircon', 'a/c', 'freon']),
    ('Electrical & Lights',    ['electrical', 'wiring', 'light', 'signal',
                                'battery', 'starter', 'alternator', 'horn',
                                'fuse', 'busted']),
    ('Engine & Cooling',       ['engine', 'overheat', 'radiator', 'coolant',
                                'fan belt', 'injector', 'turbo', 'makina',
                                'oil leak', 'gasket', 'fuel']),
    ('Clutch & Transmission',  ['clutch', 'transmission', 'gear']),
    ('Suspension & Chassis',   ['suspension', 'spring', 'shock', 'bushing',
                                'chassis', 'tie rod', 'axle', 'differential',
                                'bearing']),
    ('Hydraulics / Hoist',     ['hydraulic', 'hoist', 'cylinder']),
    ('Welding / Body Works',   ['weld', 'fabricat', 'body', 'siding',
                                'flooring', 'gate', 'bracket']),
]


def _jo_category(description):
    d = ' '.join(str(description or '').split()).lower()
    if not d:
        return '(No description)'
    for cat, keywords in _JO_CATEGORIES:
        if any(k in d for k in keywords):
            return cat
    return 'Others'


@app.route('/breakdown/print-report')
@login_required
def breakdown_print_report():
    """Printable summary of Job Orders grouped by KIND of work
    ("Change tire", "Repair brake", ...) for the same Year/Month/
    Status/Plate window the Breakdown page is filtered to. Opens in a
    new tab and auto-triggers the browser's print dialog."""
    year   = request.args.get('year',  ph_today().year,  type=int)
    month  = request.args.get('month', ph_today().month, type=int)
    filter_status = request.args.get('status', 'all')
    filter_plate  = request.args.get('plate', 'all')

    last_day = calendar.monthrange(year, month)[1]
    mo_s, mo_e = date(year, month, 1), date(year, month, last_day)

    q = BreakdownLog.query.filter(BreakdownLog.date >= mo_s,
                                  BreakdownLog.date <= mo_e)
    if filter_status != 'all':
        q = q.filter(BreakdownLog.status == filter_status)
    if filter_plate != 'all':
        try:
            q = q.filter(BreakdownLog.plate_id == int(filter_plate))
        except ValueError:
            pass
    logs = q.all()

    # Bucket every JO into a repair category (see _JO_CATEGORIES), and
    # keep a tally of the raw descriptions inside each bucket so the
    # print-out still shows what the category actually contained
    # ("Flat tire ×2 · Vulcanize tire ×1").
    from collections import Counter
    groups = {}
    for l in logs:
        cat = _jo_category(l.description)
        g = groups.get(cat)
        if g is None:
            g = groups[cat] = {
                'label': cat,
                'count': 0, 'under_repair': 0, 'fixed': 0,
                'hours': 0.0, 'units': set(), 'details': Counter(),
            }
        g['count'] += 1
        if l.status == 'Under Repair':
            g['under_repair'] += 1
        elif l.status == 'Fixed':
            g['fixed'] += 1
        g['hours'] += l.duration_hours
        unit = (l.plate.body_no or l.plate.plate_no) if l.plate \
               else (l.equipment_name or '—')
        g['units'].add(unit)
        # Sub-item key: drop parentheticals so "Flat tire (rear left)"
        # and "Flat tire (front)" tally together as "Flat tire".
        desc = re.sub(r'\([^)]*\)', '', l.description or '')
        desc = ' '.join(desc.split())
        if desc:
            g['details'][desc.capitalize()] += 1

    rows = sorted(groups.values(),
                  key=lambda g: (-g['count'], g['label'].lower()))
    for g in rows:
        g['units'] = ', '.join(sorted(g['units']))
        g['hours'] = round(g['hours'], 1)
        g['detail_items'] = [
            (f'{d} — {n}' if n > 1 else d)
            for d, n in g['details'].most_common()]

    totals = {
        'count':        sum(g['count'] for g in rows),
        'under_repair': sum(g['under_repair'] for g in rows),
        'fixed':        sum(g['fixed'] for g in rows),
        'hours':        round(sum(g['hours'] for g in rows), 1),
    }
    plate_label = 'All units'
    if filter_plate != 'all':
        p = db.session.get(Plate, int(filter_plate)) \
            if filter_plate.isdigit() else None
        if p:
            plate_label = p.display

    return render_template('reports/print_breakdown.html',
        rows=rows, totals=totals,
        period=mo_s.strftime('%B %Y'),
        status_label=('All statuses' if filter_status == 'all'
                      else filter_status),
        plate_label=plate_label,
        generated=datetime.now(PH_TZ).strftime('%b %d, %Y %I:%M %p'),
        generated_by=get_user())


@app.route('/schedule/print-report')
@login_required
def schedule_print_report():
    """Printable materials summary: every product hauled in the date
    range with trip counts and total volume. Cancelled trips are
    excluded. Opens in a new tab and auto-triggers printing."""
    from_d = parse_date(request.args.get('from', ph_today().isoformat()))
    to_d   = parse_date(request.args.get('to',   ph_today().isoformat()))
    if from_d > to_d:
        from_d, to_d = to_d, from_d

    trips = (db.session.query(TripRecord, Wave)
             .join(Wave, TripRecord.wave_id == Wave.id)
             .filter(Wave.date >= from_d, Wave.date <= to_d,
                     db.or_(TripRecord.status.is_(None),
                            TripRecord.status != 'Canceled'))
             .all())

    groups = {}
    for t, w in trips:
        name = t.product.name if t.product else '(No product)'
        g = groups.get(name)
        if g is None:
            g = groups[name] = {'label': name, 'trips': 0,
                                'delivered': 0, 'volume': 0.0,
                                'no_volume': 0}
        g['trips'] += 1
        if t.status == 'Delivered':
            g['delivered'] += 1
        try:
            g['volume'] += float(str(t.volume).replace(',', ''))
        except (TypeError, ValueError):
            g['no_volume'] += 1

    rows = sorted(groups.values(),
                  key=lambda g: (-g['volume'], g['label'].lower()))
    for g in rows:
        g['volume'] = round(g['volume'], 2)

    totals = {
        'trips':     sum(g['trips'] for g in rows),
        'delivered': sum(g['delivered'] for g in rows),
        'volume':    round(sum(g['volume'] for g in rows), 2),
        'no_volume': sum(g['no_volume'] for g in rows),
    }
    doc = {k: AppSetting.get(k, v) for k, v in DOC_HEADER_DEFAULTS.items()}
    period = (from_d.strftime('%B %d, %Y') if from_d == to_d else
              f"{from_d.strftime('%b %d, %Y')} – {to_d.strftime('%b %d, %Y')}")

    return render_template('reports/print_materials.html',
        rows=rows, totals=totals, period=period, doc=doc,
        generated=datetime.now(PH_TZ).strftime('%b %d, %Y %I:%M %p'),
        generated_by=get_user())


@app.route('/api/breakdown/unit-logs')
@login_required
def api_breakdown_unit_logs():
    """Modal-detail endpoint for the Breakdowns-by-Plate bar chart.

    Returns the list of breakdown rows for one specific unit within
    the same Year/Month/Status window the page is currently showing.
    The frontend invokes this when the dispatcher clicks a bar — the
    response populates a Bootstrap modal listing each Job Order so
    they can drill into trailers and other unmapped equipment that
    the existing Plate dropdown filter can't reach.

    Query string:
      type   — 'plate' or 'equipment' (required)
      id     — Plate.id (required when type=plate)
      name   — equipment_name string (required when type=equipment)
      year   — int, defaults to current PHT year
      month  — int, defaults to current PHT month
      status — 'all' or one of BREAKDOWN_STATUSES, defaults to 'all'

    Response shape:
      { label: str,
        count: int,
        logs:  [ {id, date, jo_ref_no, jo_url, description, status,
                  started_at, ended_at, duration_hours, remarks,
                  operator_name}, ... ] }
    """
    unit_type = (request.args.get('type') or '').strip().lower()
    if unit_type not in ('plate', 'equipment', 'driver'):
        return jsonify({'error': "type must be 'plate', 'equipment', or 'driver'"}), 400

    year   = request.args.get('year',  ph_today().year,  type=int)
    month  = request.args.get('month', ph_today().month, type=int)
    status = request.args.get('status', 'all')

    last_day = calendar.monthrange(year, month)[1]
    mo_s     = date(year, month, 1)
    mo_e     = date(year, month, last_day)

    q = BreakdownLog.query.filter(
        BreakdownLog.date >= mo_s,
        BreakdownLog.date <= mo_e
    )
    if status != 'all':
        q = q.filter(BreakdownLog.status == status)

    label = ''
    if unit_type == 'plate':
        pid = request.args.get('id', type=int)
        if not pid:
            return jsonify({'error': 'id is required for type=plate'}), 400
        q = q.filter(BreakdownLog.plate_id == pid)
        p = Plate.query.get(pid)
        label = p.display if p else f'Plate #{pid}'
    elif unit_type == 'equipment':
        name = (request.args.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'name is required for type=equipment'}), 400
        # Equipment bars come from rows with plate_id IS NULL AND a
        # specific equipment_name. Filter exactly that combo so we
        # don't accidentally surface plate-matched rows that share
        # the equipment label.
        q = q.filter(BreakdownLog.plate_id.is_(None),
                     BreakdownLog.equipment_name == name)
        label = name

    else:  # driver — operator_name IN <list of aliases>
        # The frontend passes every original spelling that the
        # server-side grouping merged into this bar (e.g.
        # 'JIM LAYAG' + 'J.LAYAG'). Filtering by IN keeps the
        # endpoint stateless — no need to re-run the normaliser
        # here. The canonical display label is sent separately
        # so the modal title reads as the dispatcher expects.
        names = [n.strip() for n in request.args.getlist('name')
                 if n and n.strip()]
        if not names:
            return jsonify({'error': 'name(s) required for type=driver'}), 400
        q = q.filter(BreakdownLog.operator_name.in_(names))
        # Optional 'label' arg lets the frontend pass the canonical
        # display so the modal title matches the bar label exactly
        # (otherwise we'd fall back to one of the aliases).
        label = (request.args.get('label') or names[0]).strip()

    rows = q.order_by(BreakdownLog.date.desc(),
                      BreakdownLog.id.desc()).all()

    def _fmt_dt(dt):
        return dt.strftime('%Y-%m-%d %I:%M %p') if dt else None

    logs = [{
        'id':              r.id,
        'date':            r.date.isoformat() if r.date else None,
        'jo_ref_no':       getattr(r, 'jo_ref_no', None),
        'jo_url':          getattr(r, 'jo_url',    None),
        'description':     r.description,
        'status':          r.status,
        'started_at':      _fmt_dt(r.started_at),
        'ended_at':        _fmt_dt(r.ended_at),
        'duration_hours':  round(r.duration_hours, 2) if r.duration_hours else None,
        'remarks':         r.remarks,
        'operator_name':   getattr(r, 'operator_name', None),
    } for r in rows]

    return jsonify({'label': label, 'count': len(logs), 'logs': logs})


def _parse_dt(s):
    """Accept 'YYYY-MM-DDTHH:MM' (HTML5 datetime-local) or 'YYYY-MM-DD HH:MM[:SS]'."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S',
                '%Y-%m-%d %H:%M',  '%Y-%m-%d %H:%M:%S'):
        try:    return datetime.strptime(s, fmt)
        except: continue
    return None

@app.route('/api/breakdown/add', methods=['POST'])
@login_required
def api_breakdown_add():
    data = request.get_json()
    plate_id    = data.get('plate_id') or None
    date_str    = data.get('date', ph_today().isoformat())
    description = (data.get('description') or '').strip()
    status      = data.get('status', 'Under Repair')
    remarks     = (data.get('remarks') or '').strip()
    started_at  = _parse_dt(data.get('started_at'))
    ended_at    = _parse_dt(data.get('ended_at'))

    if plate_id: plate_id = int(plate_id)

    log = BreakdownLog(
        plate_id    = plate_id,
        date        = parse_date(date_str),
        description = description or None,
        status      = status,
        remarks     = remarks or None,
        started_at  = started_at,
        ended_at    = ended_at,
        updated_by  = get_user(),
    )
    db.session.add(log)
    db.session.commit()

    plate = Plate.query.get(plate_id) if plate_id else None
    log_change(f"Breakdown logged: {plate.display if plate else 'Unknown'} on {date_str} — {status}", 'breakdown')
    db.session.commit()

    return jsonify(log.to_dict())


@app.route('/api/breakdown/<int:lid>/update', methods=['POST'])
@login_required
def api_breakdown_update(lid):
    data = request.get_json()
    log  = BreakdownLog.query.get_or_404(lid)

    if 'plate_id' in data:
        log.plate_id = int(data['plate_id']) if data['plate_id'] else None
    if 'date' in data:
        log.date = parse_date(data['date'])
    if 'description' in data:
        log.description = data['description'] or None
    if 'status' in data:
        log.status = data['status']
    if 'resolved_date' in data:
        rd = data['resolved_date']
        log.resolved_date = parse_date(rd) if rd else None
    if 'started_at' in data:
        log.started_at = _parse_dt(data['started_at'])
    if 'ended_at' in data:
        log.ended_at = _parse_dt(data['ended_at'])
    if 'remarks' in data:
        log.remarks = data['remarks'] or None

    log.updated_by = get_user()
    log.updated_at = utc_now()
    db.session.commit()

    log_change(f"Updated breakdown #{lid}", 'breakdown')
    db.session.commit()

    return jsonify(log.to_dict())


@app.route('/api/breakdown/<int:lid>/delete', methods=['POST'])
@login_required
def api_breakdown_delete(lid):
    if not check_can_delete():
        return jsonify({'error': 'You do not have permission to delete.'}), 403
    log = BreakdownLog.query.get_or_404(lid)
    info = f"breakdown #{lid}"
    db.session.delete(log)
    log_change(f"Deleted {info}", 'breakdown')
    db.session.commit()
    return jsonify({'ok': True})


# ── CARTRACK GPS INTEGRATION ──────────────────────────────────────────────
@app.route('/api/cartrack/vehicles')
@login_required
def api_cartrack_vehicles():
    """Return the list of vehicles from Cartrack (for plate-mapping UI)."""
    from cartrack_client import CartrackClient
    cc = CartrackClient.from_env()
    if not cc.configured:
        return jsonify({'error': 'Cartrack not configured. Set CARTRACK_USERNAME and CARTRACK_PASSWORD in environment variables.'}), 503
    vehicles, err = cc.list_vehicles()
    if err:
        return jsonify({'error': err}), 502
    # Trim down to just what the UI needs
    slim = [{
        'vehicle_id':    v.get('vehicle_id'),
        'registration':  v.get('registration') or v.get('vehicle_name'),
        'vehicle_name':  v.get('vehicle_name'),
        'description':   v.get('client_vehicle_description'),
        'type':          v.get('vehicle_type'),
    } for v in (vehicles or [])]
    return jsonify({'vehicles': slim, 'total': len(slim)})


@app.route('/api/cartrack/plate/<int:plate_id>/map', methods=['POST'])
@login_required
def api_cartrack_plate_map(plate_id):
    """Map a Plate to a Cartrack vehicle_id (or clear the mapping)."""
    plate = Plate.query.get_or_404(plate_id)
    data = request.get_json() or {}
    cartrack_id = data.get('cartrack_vehicle_id')
    if cartrack_id in (None, '', 0):
        plate.cartrack_vehicle_id = None
        action = f"Unmapped {plate.display} from Cartrack"
    else:
        try:
            plate.cartrack_vehicle_id = int(cartrack_id)
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid cartrack_vehicle_id'}), 400
        action = f"Mapped {plate.display} -> Cartrack #{cartrack_id}"
    db.session.commit()
    log_change(action, 'plates')
    db.session.commit()
    return jsonify({'ok': True, 'plate_id': plate.id,
                    'cartrack_vehicle_id': plate.cartrack_vehicle_id})


@app.route('/api/cartrack/auto-map', methods=['POST'])
@login_required
def api_cartrack_auto_map():
    """Try to auto-match all unmapped active plates to Cartrack vehicles by
    matching the plate number against the vehicle's registration field."""
    from cartrack_client import CartrackClient
    cc = CartrackClient.from_env()
    if not cc.configured:
        return jsonify({'error': 'Cartrack not configured'}), 503

    vehicles, err = cc.list_vehicles()
    if err:
        return jsonify({'error': err}), 502

    matched, ambiguous, unmatched = [], [], []
    plates = Plate.query.filter(Plate.active == True,
                                Plate.cartrack_vehicle_id.is_(None)).all()
    from cartrack_client import _norm_plate
    for plate in plates:
        target = _norm_plate(plate.plate_no)
        if not target:
            continue
        hits = []
        for v in vehicles:
            for field in ('registration', 'vehicle_name'):
                v_norm = _norm_plate(v.get(field) or '')
                if target and v_norm and target in v_norm:
                    hits.append(v)
                    break
        if len(hits) == 1:
            plate.cartrack_vehicle_id = hits[0].get('vehicle_id')
            matched.append({'plate': plate.display,
                            'cartrack_id': plate.cartrack_vehicle_id,
                            'cartrack_name': hits[0].get('vehicle_name')})
        elif len(hits) > 1:
            ambiguous.append({'plate': plate.display,
                              'candidates': [h.get('vehicle_name') for h in hits]})
        else:
            unmatched.append(plate.display)
    db.session.commit()
    log_change(f"Cartrack auto-map: {len(matched)} matched, {len(ambiguous)} ambiguous, {len(unmatched)} unmatched", 'plates')
    db.session.commit()
    return jsonify({
        'matched_count':   len(matched),
        'ambiguous_count': len(ambiguous),
        'unmatched_count': len(unmatched),
        'matched':         matched,
        'ambiguous':       ambiguous,
        'unmatched':       unmatched,
    })


@app.route('/api/cartrack/poll-now', methods=['POST'])
@login_required
def api_cartrack_poll_now():
    """Manually trigger the polling worker. Returns the summary dict."""
    try:
        from cartrack_poll import run_poll
        summary = run_poll(app=app)
        return jsonify(summary)
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/cartrack/poll-cron', methods=['GET', 'POST'])
def api_cartrack_poll_cron():
    """External-cron entrypoint for Cartrack polling.

    Bypasses @login_required so services like cron-job.org or GitHub Actions
    can hit this endpoint on a schedule. Auth is via a shared secret token
    set in the CRON_SECRET environment variable.

    Usage:
        GET/POST /api/cartrack/poll-cron?token=YOUR_SECRET
        — OR —
        GET/POST /api/cartrack/poll-cron
        Header: X-Cron-Token: YOUR_SECRET

    Returns the same summary dict as /api/cartrack/poll-now on success,
    or 401 if the token is missing/invalid, 500 on poll exceptions.
    """
    # Resolve the expected secret
    expected = os.environ.get('CRON_SECRET', '').strip()
    if not expected:
        return jsonify({
            'error': 'CRON_SECRET env var not configured on server. '
                     'Set it in the WSGI config to enable this endpoint.'
        }), 503

    # Accept token from either query string or header (cron-job.org friendly)
    provided = (request.args.get('token') or
                request.headers.get('X-Cron-Token') or '').strip()
    if not provided or provided != expected:
        return jsonify({'error': 'Unauthorized — invalid or missing token'}), 401

    # Run the same polling worker as the "Poll Now" button
    try:
        from cartrack_poll import run_poll
        summary = run_poll(app=app)
        # Augment with caller info for cron dashboard visibility
        summary['triggered_by'] = 'cron'
        summary['ts'] = ph_now().isoformat()
        return jsonify(summary)
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'trace': traceback.format_exc(),
        }), 500


@app.route('/api/cartrack/status')
@login_required
def api_cartrack_status():
    """Diagnostic endpoint: how many plates are mapped, when last poll ran, etc."""
    total_plates = Plate.query.filter_by(active=True).count()
    mapped_plates = Plate.query.filter(
        Plate.active == True, Plate.cartrack_vehicle_id.isnot(None)
    ).count()
    # Most recent CartrackEvent
    last_event = CartrackEvent.query.order_by(CartrackEvent.created_at.desc()).first()
    # CartrackTruckState count + open trips count
    state_total = CartrackTruckState.query.count()
    open_trips  = CartrackTruckState.query.filter(
        CartrackTruckState.entry_plaza.isnot(None)
    ).count()
    # Cartrack config
    cartrack_configured = bool(os.environ.get('CARTRACK_USERNAME') and
                               os.environ.get('CARTRACK_PASSWORD'))
    # Toll fills today
    from sqlalchemy import func
    today = ph_today()
    fills_today = (db.session.query(func.count(CartrackEvent.id))
                   .filter(CartrackEvent.event_type == 'trip_closed',
                           CartrackEvent.toll_fee.isnot(None),
                           func.date(CartrackEvent.created_at) == today)
                   .scalar() or 0)
    return jsonify({
        'cartrack_configured': cartrack_configured,
        'plates_total':        total_plates,
        'plates_mapped':       mapped_plates,
        'plates_unmapped':     total_plates - mapped_plates,
        'tracked_trucks':      state_total,
        'open_trips':          open_trips,
        'last_event': {
            'when':       iso_ph(last_event.created_at),
            'plate_id':   last_event.plate_id if last_event else None,
            'type':       last_event.event_type if last_event else None,
            'plaza':      last_event.plaza_name if last_event else None,
        } if last_event else None,
        'auto_fills_today': fills_today,
    })


# ── JOB ORDERS / REPAIR REQUEST INTEGRATION ──────────────────────────────
# Mirrors the Cartrack admin endpoints: a status check, a manual sync
# trigger, and a token-protected cron entrypoint. All three share the
# same env vars (JOBORDERS_BASE_URL, JOBORDERS_TOKEN) and the same sync
# worker (joborders_sync.run_sync).

@app.route('/api/joborders/status')
@login_required
def api_joborders_status():
    """Diagnostic snapshot of the ERP Repair Request integration.

    Returns enough state for an admin to tell at a glance:
      - Are the env vars set?
      - When did we last successfully pull a record?
      - How many ERP-sourced breakdowns do we have locally?
      - Are there any unlinked rows (no matching plate)?
    """
    from sqlalchemy import func
    configured = bool(os.environ.get('JOBORDERS_TOKEN'))
    base_url   = os.environ.get('JOBORDERS_BASE_URL',
                                 'https://erp-api.gainersand.ph/api')

    # ERP-sourced BreakdownLog stats
    erp_total = (db.session.query(func.count(BreakdownLog.id))
                  .filter(BreakdownLog.jo_external_id.isnot(None))
                  .scalar() or 0)
    erp_unlinked = (db.session.query(func.count(BreakdownLog.id))
                     .filter(BreakdownLog.jo_external_id.isnot(None),
                             BreakdownLog.plate_id.is_(None))
                     .scalar() or 0)
    last_sync = (db.session.query(func.max(BreakdownLog.last_synced_at))
                  .filter(BreakdownLog.jo_external_id.isnot(None))
                  .scalar())
    return jsonify({
        'configured':           configured,
        'base_url':             base_url,
        'erp_sourced_total':    erp_total,
        'unlinked_no_plate':    erp_unlinked,
        'last_sync':            iso_ph(last_sync) if last_sync else None,
    })


@app.route('/api/joborders/sync-now', methods=['POST'])
@login_required
def api_joborders_sync_now():
    """Manually trigger one sync iteration. Returns the summary dict.

    Useful for first-time testing (before scheduling the always-on task)
    and for one-off "I just added a repair request, pull it in now"
    workflows. Body params (optional, all default sensibly):

        { "filter":    "" | "pending" | "approved" | "rejected",
          "from_date": "YYYY-MM-DD",
          "to_date":   "YYYY-MM-DD" }
    """
    try:
        from joborders_sync import run_sync
        body = request.get_json(silent=True) or {}
        summary = run_sync(
            app=app,
            filter=body.get('filter', ''),
            from_date=body.get('from_date'),
            to_date=body.get('to_date'),
        )
        return jsonify(summary)
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/joborders/sync-cron', methods=['GET', 'POST'])
def api_joborders_sync_cron():
    """External-cron entrypoint for the ERP sync. Mirrors the Cartrack
    poll-cron endpoint — same shared-secret auth pattern via CRON_SECRET
    env var.

    Use this instead of an always-on task if you want simpler scheduling
    (cron-job.org, GitHub Actions, etc.).
    """
    expected = os.environ.get('CRON_SECRET', '').strip()
    if not expected:
        return jsonify({
            'error': 'CRON_SECRET env var not configured on server.',
        }), 503

    provided = (request.args.get('token', '').strip()
                or request.headers.get('X-Cron-Token', '').strip())
    if provided != expected:
        return jsonify({'error': 'Invalid or missing cron token.'}), 401

    try:
        from joborders_sync import run_sync
        summary = run_sync(app=app)
        return jsonify(summary)
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


# ── TOLL CALCULATOR ───────────────────────────────────────────────────────
import json as _json_mod
from collections import deque
_TOLL_DATA = None

def get_toll_data():
    global _TOLL_DATA
    if _TOLL_DATA is None:
        try:
            p = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'toll_rates.json')
            # Force UTF-8 — file contains special chars like ñ (Parañaque, etc.)
            # Default encoding on some Linux locales is ASCII, which would fail.
            with open(p, 'r', encoding='utf-8') as f:
                _TOLL_DATA = _json_mod.load(f)
        except Exception as e:
            import sys
            err = f'[Toll] Could not load toll_rates.json: {e}'
            print(err, flush=True)
            sys.stderr.write(err + '\n')
            sys.stderr.flush()
            _TOLL_DATA = {}
    return _TOLL_DATA

# Expressway connection points — where two expressways physically meet.
# Each pair is bidirectional: (expA, stationA) <-> (expB, stationB)
_TOLL_CONNECTIONS = [
    # NLEX south end ↔ Skyway Stage 3 north end (Balintawak) — free transfer, same physical point
    (('NLEX_SCTEX', 'Balintawak'),        ('Skyway_Stage3', 'Balintawak')),
    (('NLEX_SCTEX', 'Mindanao Avenue'),   ('Skyway_Stage3', 'Balintawak')),
    # Skyway Stage 3 south end ↔ Skyway / SLEX north end (Buendia) — free transfer
    (('Skyway_Stage3', 'Buendia'),        ('Skyway_SLEX_MCX', 'Skyway / Buendia')),
    # Skyway/SLEX south end ↔ STAR north end (Sto. Tomas / Calamba)
    (('Skyway_SLEX_MCX', 'Sto. Tomas'),   ('STAR', 'Sto. Tomas')),
    (('Skyway_SLEX_MCX', 'Calamba'),      ('STAR', 'Calamba')),
    # NLEX/SCTEX north end ↔ TPLEX south end
    (('NLEX_SCTEX', 'Tarlac'),            ('TPLEX', 'La Paz')),
    # NLEX/SCTEX terminus alias — Sta. Ines ≈ Mabalacat area
    (('NLEX_SCTEX', 'Sta. Ines'),         ('NLEX_SCTEX', 'Mabalacat')),
    # Skyway/SLEX ↔ CALAX
    (('Skyway_SLEX_MCX', 'Mamplasan'),    ('CALAX', 'Laguna Boulevard')),
    # Skyway ↔ NAIAX (Skyway entry from NAIAX side maps to Skyway/Buendia area)
    (('Skyway_SLEX_MCX', 'Skyway / Buendia'), ('NAIAX', 'Skyway')),
    (('Skyway_SLEX_MCX', 'Magallanes'),       ('NAIAX', 'Skyway')),
    # NAIAX south end ↔ CAVITEX (the two systems meet near the airport-Cavite corridor)
    (('NAIAX', 'CAVITEX'),                ('CAVITEX', 'Parañaque')),
    (('NAIAX', 'CAVITEX'),                ('CAVITEX', 'C5 Rd. Ext./C.P. Garcia')),

    # NLEX Connector — 8-km elevated bypass linking NLEX (Caloocan side) to
    # Skyway Stage 3 (España side) without going through EDSA.
    # North end: C-3 Road / 5th Ave., same physical area as NLEX Balintawak/Karuhatan.
    (('NLEX_Connector', 'C-3 Road/5th Ave.'), ('NLEX_SCTEX', 'Balintawak')),
    (('NLEX_Connector', 'C-3 Road/5th Ave.'), ('NLEX_SCTEX', 'Karuhatan')),
    # South end: España area, drops onto Skyway Stage 3 near Quirino / Plaza Dilao.
    (('NLEX_Connector', 'España'),             ('Skyway_Stage3', 'Quirino')),
    (('NLEX_Connector', 'España'),             ('Skyway_Stage3', 'Plaza Dilao')),

    # Harbor Link / NLEX Segment 10 — spur of NLEX heading to Manila Port.
    # Karuhatan booth is the same physical location on both NLEX and Harbor Link.
    (('Harbor_Link', 'Karuhatan/Valenzuela'), ('NLEX_SCTEX', 'Karuhatan')),
    # Harbor Link Interchange = the NLEX junction (already-free transfer point).
    (('Harbor_Link', 'Harbor Link Interchange'), ('NLEX_SCTEX', 'Balintawak')),
    # Harbor Link and NLEX Connector share a C-3 Road exit (same physical booth).
    (('Harbor_Link', 'C-3 Road/5th Ave.'), ('NLEX_Connector', 'C-3 Road/5th Ave.')),
]

def _toll_lookup(matrix, a, b):
    """Symmetric rate lookup in a class matrix."""
    return matrix.get(a, {}).get(b) or matrix.get(b, {}).get(a)

def find_toll_route(entry, exit_point, toll_class, data):
    """
    BFS across expressway connection points to find the cheapest multi-hop route.
    Returns (total_amount, segments_list) or (None, None).
    segments_list = [{'expressway': key, 'from': stn, 'to': stn, 'amount': float}, ...]
    """
    # Build bidirectional neighbour map for connection stations
    conn_neighbours = {}   # (exp, stn) -> [(exp, stn), ...]
    for (e1, s1), (e2, s2) in _TOLL_CONNECTIONS:
        conn_neighbours.setdefault((e1, s1), []).append((e2, s2))
        conn_neighbours.setdefault((e2, s2), []).append((e1, s1))

    # BFS state: (exp_key, station, accumulated_cost, segments)
    # Seed with every expressway that contains the entry station
    queue = deque()
    for exp_key, exp_data in data.items():
        matrix = exp_data.get(toll_class, {})
        if entry in matrix or any(entry in v for v in matrix.values()):
            queue.append((exp_key, entry, 0.0, []))

    visited   = set()
    best_cost = None
    best_segs = None

    while queue:
        cur_exp, cur_stn, cost_so_far, segs = queue.popleft()

        state = (cur_exp, cur_stn)
        if state in visited:
            continue
        visited.add(state)

        matrix = data.get(cur_exp, {}).get(toll_class, {})

        # ── Can we reach the exit from current expressway? ──────────────────
        amt = _toll_lookup(matrix, cur_stn, exit_point)
        if amt is not None:
            total = cost_so_far + amt
            if best_cost is None or total < best_cost:
                best_cost = total
                best_segs = segs + [{'expressway': cur_exp,
                                      'from': cur_stn, 'to': exit_point,
                                      'amount': amt}]
            continue   # don't expand further — already reached destination

        # ── Try every connection point reachable from current expressway ────
        for (c_exp, c_stn), neighbours in conn_neighbours.items():
            if c_exp != cur_exp:
                continue
            conn_amt = _toll_lookup(matrix, cur_stn, c_stn) if cur_stn != c_stn else 0.0
            if conn_amt is None:
                continue
            new_segs = segs + [{'expressway': cur_exp,
                                 'from': cur_stn, 'to': c_stn,
                                 'amount': conn_amt}]
            for next_exp, next_stn in neighbours:
                if (next_exp, next_stn) not in visited:
                    queue.append((next_exp, next_stn,
                                  cost_so_far + conn_amt, new_segs))

    return best_cost, best_segs

@app.route('/api/toll/expressways')
@login_required
def api_toll_expressways():
    data = get_toll_data()
    result = [{'key': k, 'name': v.get('name', k)} for k, v in data.items()]
    return jsonify(result)

@app.route('/api/toll/all-stations')
@login_required
def api_toll_all_stations():
    """Return every station across all expressways with which expressway(s) it belongs to."""
    data = get_toll_data()
    station_map = {}   # station_name -> set of expressway keys
    for exp_key, exp_data in data.items():
        matrix = exp_data.get('Class 3', exp_data.get('Class 1', exp_data.get('Class 2', {})))
        for stn in matrix.keys():
            station_map.setdefault(stn, set()).add(exp_key)
    result = [
        {'station': s, 'expressways': sorted(exps)}
        for s, exps in sorted(station_map.items())
    ]
    return jsonify(result)

@app.route('/api/toll/stations/<expressway>')
@login_required
def api_toll_stations(expressway):
    data = get_toll_data()
    exp = data.get(expressway, {})
    matrix = exp.get('Class 3', exp.get('Class 1', exp.get('Class 2', {})))
    stations = sorted(matrix.keys())
    return jsonify(stations)

@app.route('/api/toll/calculate', methods=['POST'])
@login_required
def api_toll_calculate():
    req = request.get_json()
    expressway  = req.get('expressway', '')   # optional – auto-detected if blank
    entry       = req.get('entry', '')
    exit_point  = req.get('exit', '')
    toll_class  = req.get('toll_class', 'Class 3')
    data = get_toll_data()

    # Auto-detect expressway: try every expressway until a rate is found
    if not expressway:
        for exp_key, exp_data in data.items():
            matrix = exp_data.get(toll_class, {})
            amt = (matrix.get(entry, {}).get(exit_point) or
                   matrix.get(exit_point, {}).get(entry))
            if amt is not None:
                expressway = exp_key
                break

    exp = data.get(expressway, {})
    matrix = exp.get(toll_class, {})
    amount = (matrix.get(entry, {}).get(exit_point) or
              matrix.get(exit_point, {}).get(entry))
    if amount is None:
        # ── Fallback: BFS multi-expressway routing ──────────────────────────
        best_cost, best_segs = find_toll_route(entry, exit_point, toll_class, data)
        if best_cost is not None:
            return jsonify({'amount': best_cost, 'segments': best_segs,
                            'expressway': 'multi', 'entry': entry,
                            'exit': exit_point, 'toll_class': toll_class})
        return jsonify({'error': 'Rate not found', 'amount': 0})
    return jsonify({'amount': amount, 'expressway': expressway,
                    'segments': [{'expressway': expressway, 'from': entry,
                                  'to': exit_point, 'amount': amount}],
                    'entry': entry, 'exit': exit_point, 'toll_class': toll_class})

@app.route('/toll-calculator')
@login_required
def toll_calculator():
    return render_template('toll_calculator.html')


# ── TOLL LOG (Cartrack GPS Events) ────────────────────────────────────────
@app.route('/toll-log')
@login_required
def toll_log():
    """Render the Toll Log page — shows all Cartrack plaza events and auto-fills."""
    return render_template('toll_log/index.html')


@app.route('/api/toll-log/summary')
@login_required
def api_toll_log_summary():
    """Return summary stats for the Toll Log page: today/week/month counts + fees."""
    today = ph_today()
    # CartrackEvent.created_at is stored in UTC. PHT midnight = 16:00
    # UTC the previous day — combining the PHT date with min.time()
    # and comparing it raw against UTC made "today" start at 8:00 AM
    # PHT, so all early-morning transits (the bulk of trucking) were
    # missing from the Today KPIs.
    today_start = (datetime.combine(today, datetime.min.time())
                   .replace(tzinfo=PH_TZ).astimezone(UTC_TZ)
                   .replace(tzinfo=None))
    week_start  = today_start - timedelta(days=7)
    month_start = today_start - timedelta(days=30)

    def stats_for(since):
        events = CartrackEvent.query.filter(CartrackEvent.created_at >= since)
        total = events.count()
        enters = events.filter_by(event_type='enter').count()
        exits  = events.filter_by(event_type='exit').count()
        closed = events.filter_by(event_type='trip_closed').all()
        total_fee = sum((e.toll_fee or 0) for e in closed)
        unique_trucks = events.with_entities(CartrackEvent.plate_id).distinct().count()
        return {
            'total_events': total,
            'enters': enters,
            'exits': exits,
            'trip_closed': len(closed),
            'total_toll_fee': round(total_fee, 2),
            'unique_trucks': unique_trucks,
        }

    return jsonify({
        'today': stats_for(today_start),
        'week':  stats_for(week_start),
        'month': stats_for(month_start),
    })


@app.route('/api/toll-log/events')
@login_required
def api_toll_log_events():
    """Return filtered list of CartrackEvent rows for the Toll Log table."""
    # Parse filters
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    plate_id  = request.args.get('plate_id', '')
    event_type= request.args.get('event_type', '')
    expressway= request.args.get('expressway', '')
    limit     = int(request.args.get('limit', 200))

    q = CartrackEvent.query
    if date_from:
        try:
            d = pht_filter_to_utc(date_from)
            q = q.filter(CartrackEvent.created_at >= d)
        except ValueError:
            pass
    if date_to:
        try:
            d = pht_filter_to_utc(date_to, plus_days=1)
            q = q.filter(CartrackEvent.created_at < d)
        except ValueError:
            pass
    if plate_id:
        try:
            q = q.filter(CartrackEvent.plate_id == int(plate_id))
        except ValueError:
            pass
    if event_type:
        q = q.filter(CartrackEvent.event_type == event_type)
    if expressway:
        q = q.filter(CartrackEvent.expressway == expressway)

    events = q.order_by(CartrackEvent.created_at.desc()).limit(limit).all()

    rows = []
    for e in events:
        plate = e.plate
        rows.append({
            'id':         e.id,
            'created_at': iso_ph(e.created_at),
            'plate_id':   e.plate_id,
            # Display form ("DT06 / LAK8098") so the UI doesn't need the
            # truck code memorized — falls back to bare plate_no if
            # body_no isn't set.
            'plate_no':   plate.display if plate else 'N/A',
            'event_type': e.event_type,
            'plaza_name': e.plaza_name or '',
            'expressway': e.expressway or '',
            'lat':        e.lat,
            'lng':        e.lng,
            'trip_id':    e.trip_id,
            'toll_fee':   e.toll_fee,
            'toll_entry': e.toll_entry,
            'toll_exit':  e.toll_exit,
            'notes':      e.notes or '',
        })
    return jsonify({'events': rows, 'count': len(rows), 'limit': limit})


@app.route('/api/toll-log/filters')
@login_required
def api_toll_log_filters():
    """Return dropdown options for filters: plates with events, expressways."""
    # Plates that have at least one event
    plate_ids = [pid for (pid,) in
                 db.session.query(CartrackEvent.plate_id).distinct().all()
                 if pid is not None]
    plates = Plate.query.filter(Plate.id.in_(plate_ids))\
                        .order_by(Plate.plate_no).all() if plate_ids else []

    # Distinct expressways
    expressways = sorted(set(
        (e or '') for (e,) in
        db.session.query(CartrackEvent.expressway).distinct().all()
        if e
    ))

    return jsonify({
        'plates':      [{'id': p.id, 'plate_no': p.plate_no} for p in plates],
        'expressways': expressways,
        'event_types': ['enter', 'exit', 'trip_closed'],
    })


@app.route('/api/toll-log/export')
@login_required
def api_toll_log_export():
    """Export filtered events to Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl not installed on server'}), 500

    # Same filters as events endpoint
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    plate_id  = request.args.get('plate_id', '')
    event_type= request.args.get('event_type', '')
    expressway= request.args.get('expressway', '')

    q = CartrackEvent.query
    if date_from:
        try:
            d = pht_filter_to_utc(date_from)
            q = q.filter(CartrackEvent.created_at >= d)
        except ValueError:
            pass
    if date_to:
        try:
            d = pht_filter_to_utc(date_to, plus_days=1)
            q = q.filter(CartrackEvent.created_at < d)
        except ValueError:
            pass
    if plate_id:
        try:
            q = q.filter(CartrackEvent.plate_id == int(plate_id))
        except ValueError:
            pass
    if event_type:
        q = q.filter(CartrackEvent.event_type == event_type)
    if expressway:
        q = q.filter(CartrackEvent.expressway == expressway)

    events = q.order_by(CartrackEvent.created_at.desc()).all()

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Toll Log"

    # Header row with styling
    headers = ['Date/Time', 'Plate No', 'Event Type', 'Plaza',
               'Expressway', 'Toll Fee', 'Entry Plaza', 'Exit Plaza',
               'Trip ID', 'Lat', 'Lng', 'Notes']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='8B1A2B')
    center = Alignment(horizontal='center', vertical='center')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Data rows
    for row_idx, e in enumerate(events, 2):
        plate = e.plate
        plate_no = plate.display if plate else 'N/A'
        ws.cell(row=row_idx, column=1,  value=e.created_at.strftime('%Y-%m-%d %H:%M:%S') if e.created_at else '')
        ws.cell(row=row_idx, column=2,  value=plate_no)
        ws.cell(row=row_idx, column=3,  value=e.event_type or '')
        ws.cell(row=row_idx, column=4,  value=e.plaza_name or '')
        ws.cell(row=row_idx, column=5,  value=e.expressway or '')
        ws.cell(row=row_idx, column=6,  value=float(e.toll_fee) if e.toll_fee else None)
        ws.cell(row=row_idx, column=7,  value=e.toll_entry or '')
        ws.cell(row=row_idx, column=8,  value=e.toll_exit or '')
        ws.cell(row=row_idx, column=9,  value=e.trip_id)
        ws.cell(row=row_idx, column=10, value=e.lat)
        ws.cell(row=row_idx, column=11, value=e.lng)
        ws.cell(row=row_idx, column=12, value=e.notes or '')

    # Summary row at bottom (total toll fees)
    if events:
        last_row = len(events) + 1
        summary_row = last_row + 2
        ws.cell(row=summary_row, column=1, value='TOTAL TOLL FEES:').font = Font(bold=True)
        ws.cell(row=summary_row, column=6,
                value=f'=SUM(F2:F{last_row})').font = Font(bold=True)

    # Auto-size columns
    column_widths = [20, 15, 14, 22, 12, 12, 22, 22, 10, 12, 12, 30]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Send file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'toll_log_{ph_today().isoformat()}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ── TRUCK CYCLE TIME (round-trip analytics + idling) ──────────────────────
@app.route('/truck-cycle-time')
@login_required
def truck_cycle_time():
    """Render the Truck Cycle Time analytics page."""
    return render_template('truck_cycle_time/index.html')


@app.route('/api/cycle-time/summary')
@login_required
def api_cycle_time_summary():
    """KPI summary for the Truck Cycle Time page.

    Returns counts + averages for today, last 7 days, last 30 days,
    plus a snapshot of currently-open cycles.
    """
    from sqlalchemy import func
    now = ph_now()
    today_start = datetime.combine(ph_today(), datetime.min.time())
    week_start  = today_start - timedelta(days=7)
    month_start = today_start - timedelta(days=30)

    def stats_for(since):
        cycles = TruckCycle.query.filter(
            TruckCycle.started_at >= since,
            TruckCycle.ended_at.isnot(None),
        ).all()
        closed_count = len(cycles)
        durations = [c.duration_minutes for c in cycles if c.duration_minutes]
        avg_minutes = round(sum(durations) / len(durations), 1) if durations else 0
        by_cat = {'short': 0, 'standard': 0, 'long': 0}
        for c in cycles:
            if c.category in by_cat:
                by_cat[c.category] += 1
        return {
            'closed_cycles':    closed_count,
            'avg_minutes':      avg_minutes,
            'avg_hours':        round(avg_minutes / 60, 2) if avg_minutes else 0,
            'by_category':      by_cat,
        }

    open_cycles = TruckCycle.query.filter(TruckCycle.ended_at.is_(None)).all()

    # Pre-fetch live state rows for all open cycles in one query (avoids N+1).
    plate_ids = [c.plate_id for c in open_cycles]
    state_map = {}
    if plate_ids:
        states = CartrackTruckState.query.filter(
            CartrackTruckState.plate_id.in_(plate_ids)
        ).all()
        state_map = {s.plate_id: s for s in states}

    # Pre-fetch geofence names so the "Currently at: X" badge can resolve
    # UUID -> name without one query per cycle.
    all_gf_uuids = set()
    for s in state_map.values():
        for u in (s.last_geofence_uuids or '').split(','):
            if u:
                all_gf_uuids.add(u)
    gf_name_map = {}
    if all_gf_uuids:
        for g in CartrackGeofence.query.filter(
            CartrackGeofence.cartrack_id.in_(all_gf_uuids)
        ).all():
            gf_name_map[g.cartrack_id] = g.name

    open_summary = []
    for c in open_cycles[:50]:   # cap to keep payload small
        elapsed_min = int((utc_now() - c.started_at).total_seconds() / 60)
        state = state_map.get(c.plate_id)
        # Resolve current geofence name(s) — usually just one, but the truck
        # could be in nested zones (e.g., customer site inside an industrial park).
        current_gf_names = []
        if state and state.last_geofence_uuids:
            for u in state.last_geofence_uuids.split(','):
                if u and u in gf_name_map:
                    current_gf_names.append(gf_name_map[u])
        # Compute "last seen" age in minutes so UI can flag stale data
        last_seen_age_min = None
        if state and state.last_position_at:
            last_seen_age_min = int(
                (utc_now() - state.last_position_at).total_seconds() / 60)
        open_summary.append({
            'cycle_id':    c.id,
            'plate_id':    c.plate_id,
            'plate_no':    c.plate.display if c.plate else 'N/A',
            'started_at':  iso_ph(c.started_at),
            'elapsed_minutes': elapsed_min,
            'elapsed_hours':   round(elapsed_min / 60, 1),
            # NEW: live state (from CartrackTruckState, refreshed every poll)
            'location':         (state.last_position_description if state else '') or '—',
            'current_geofences': current_gf_names,
            'status':           (state.live_status if state else 'UNKNOWN'),
            'speed':            (state.last_speed if state else 0),
            'last_seen_min_ago': last_seen_age_min,
        })

    # Expose the threshold so the UI can show "Min dwell: 5 min"
    try:
        from cartrack_poll import MIN_VISIT_MINUTES, TRACK_HOME_AS_VISIT
    except ImportError:
        MIN_VISIT_MINUTES, TRACK_HOME_AS_VISIT = 5, False

    home_row = CartrackGeofence.query.filter_by(is_home=True).first()
    return jsonify({
        'today':       stats_for(today_start),
        'week':        stats_for(week_start),
        'month':       stats_for(month_start),
        'open_count':  len(open_cycles),
        'open':        open_summary,
        'home_geofence': home_row.name if home_row else None,
        'min_visit_minutes':  MIN_VISIT_MINUTES,
        'track_home_as_visit': TRACK_HOME_AS_VISIT,
    })


@app.route('/api/cycle-time/cycles')
@login_required
def api_cycle_time_cycles():
    """Return list of cycles with filters.

    Query params:
        date_from, date_to  — ISO dates, filter by started_at
        plate_id            — single plate
        truck_type_id       — filter by truck type (DT / TH / MDT / ...)
        category            — short / standard / long / ongoing
        status              — 'open' / 'closed' / 'all'
        limit               — max rows (default 200)
    """
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    plate_id  = request.args.get('plate_id', '')
    truck_type_id = request.args.get('truck_type_id', '')
    category  = request.args.get('category', '')
    status    = request.args.get('status', 'all')
    limit     = min(int(request.args.get('limit', 200)), 1000)

    q = TruckCycle.query
    if date_from:
        try:
            q = q.filter(TruckCycle.started_at >= pht_filter_to_utc(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(TruckCycle.started_at < pht_filter_to_utc(date_to, plus_days=1))
        except ValueError:
            pass
    if plate_id:
        try:
            q = q.filter(TruckCycle.plate_id == int(plate_id))
        except ValueError:
            pass
    if truck_type_id:
        try:
            # TruckCycle has no direct truck_type_id — join through Plate.
            q = q.join(Plate, TruckCycle.plate_id == Plate.id) \
                 .filter(Plate.truck_type_id == int(truck_type_id))
        except ValueError:
            pass
    if category:
        q = q.filter(TruckCycle.category == category)
    if status == 'open':
        q = q.filter(TruckCycle.ended_at.is_(None))
    elif status == 'closed':
        q = q.filter(TruckCycle.ended_at.isnot(None))

    cycles = q.order_by(TruckCycle.started_at.desc()).limit(limit).all()

    # Pre-fetch live state for OPEN cycles only — no point joining for closed
    # cycles since the truck has long since moved on.
    open_plate_ids = [c.plate_id for c in cycles if c.ended_at is None]
    state_map = {}
    gf_name_map = {}
    if open_plate_ids:
        states = CartrackTruckState.query.filter(
            CartrackTruckState.plate_id.in_(open_plate_ids)
        ).all()
        state_map = {s.plate_id: s for s in states}
        # Resolve geofence names for the "Currently at" column
        all_gf_uuids = set()
        for s in states:
            for u in (s.last_geofence_uuids or '').split(','):
                if u:
                    all_gf_uuids.add(u)
        if all_gf_uuids:
            for g in CartrackGeofence.query.filter(
                CartrackGeofence.cartrack_id.in_(all_gf_uuids)
            ).all():
                gf_name_map[g.cartrack_id] = g.name

    rows = []
    for c in cycles:
        # Count only REAL visits (exclude drive-bys) in this cycle.
        visit_count = (SiteVisit.query
                       .filter_by(cycle_id=c.id)
                       .filter(SiteVisit.is_drive_by == False)   # noqa: E712
                       .count())
        is_open = c.ended_at is None
        # Resolve live fields (only meaningful for open cycles).
        live = {'location': '', 'current_geofences': [], 'status': '', 'speed': 0}
        if is_open:
            state = state_map.get(c.plate_id)
            if state:
                live['location'] = state.last_position_description or ''
                live['status']   = state.live_status
                live['speed']    = state.last_speed or 0
                if state.last_geofence_uuids:
                    for u in state.last_geofence_uuids.split(','):
                        if u and u in gf_name_map:
                            live['current_geofences'].append(gf_name_map[u])
        rows.append({
            'id':          c.id,
            'plate_id':    c.plate_id,
            'plate_no':    c.plate.display if c.plate else 'N/A',
            'started_at':  iso_ph(c.started_at),
            'ended_at':    iso_ph(c.ended_at),
            'duration_minutes': c.duration_minutes,
            'duration_hours':   (round(c.duration_minutes / 60, 1) if c.duration_minutes else None),
            'category':    c.category,
            'visit_count': visit_count,
            'is_open':     is_open,
            # Live state (populated only when is_open=True)
            'location':         live['location'],
            'current_geofences': live['current_geofences'],
            'status':           live['status'],
            'speed':            live['speed'],
        })
    return jsonify({'cycles': rows, 'count': len(rows), 'limit': limit})


@app.route('/api/cycle-time/plates')
@login_required
def api_cycle_time_plates():
    """Plate-centric live view. One entry per active plate, with:
      - live status (DRIVING / IDLING / STOPPED / OFF / NO DATA)
      - currently at  (list of geofence names; empty = in transit)
      - last GPS position address
      - open cycle info (if any) — id, started_at, elapsed_hours
      - visit_count (real visits, not drive-bys, in the open cycle)
      - last_departure / last_arrival  (recent SiteVisit pair)

    Sorting:
      1. Plates with open cycles first, longest-elapsed first.
      2. Then parked plates (no open cycle), alphabetically.

    Powers the new TCT Plate Status table that replaces the per-cycle
    history view. Closed cycles per plate are fetched on demand by
    /api/cycle-time/plate/<id>/cycles.

    Filters:
        truck_type_id  — restrict to one truck type
    """
    truck_type_id = request.args.get('truck_type_id', '')

    plates_q = Plate.query.filter(Plate.active == True)   # noqa: E712
    if truck_type_id:
        try:
            plates_q = plates_q.filter(Plate.truck_type_id == int(truck_type_id))
        except ValueError:
            pass
    plates = plates_q.order_by(Plate.body_no, Plate.plate_no).all()

    if not plates:
        return jsonify({'plates': [], 'open_count': 0, 'parked_count': 0})

    plate_ids = [p.id for p in plates]

    # Batch-fetch CartrackTruckState rows for these plates
    states = CartrackTruckState.query.filter(
        CartrackTruckState.plate_id.in_(plate_ids)).all()
    state_map = {s.plate_id: s for s in states}

    # Batch-fetch open cycles (one per plate at most)
    open_cycles = (TruckCycle.query
                   .filter(TruckCycle.plate_id.in_(plate_ids),
                           TruckCycle.ended_at.is_(None))
                   .all())
    open_cycle_map = {c.plate_id: c for c in open_cycles}

    # Build a set of geofence UUIDs referenced by any of the live states,
    # so we can resolve names in a single batch query.
    all_gf_uuids = set()
    for s in states:
        for u in (s.last_geofence_uuids or '').split(','):
            if u:
                all_gf_uuids.add(u)
    gf_name_map = {}
    if all_gf_uuids:
        for g in CartrackGeofence.query.filter(
            CartrackGeofence.cartrack_id.in_(all_gf_uuids)).all():
            gf_name_map[g.cartrack_id] = g.name

    # For each open cycle, fetch the most recent closed SiteVisit (last
    # arrival/departure) and the count of real visits.
    visit_aggregates = {}
    last_arrival_map = {}
    last_departure_map = {}
    for c in open_cycles:
        # Real-visit count
        from sqlalchemy import func
        cnt = (db.session.query(func.count(SiteVisit.id))
               .filter(SiteVisit.cycle_id == c.id,
                       SiteVisit.is_drive_by == False)   # noqa: E712
               .scalar() or 0)
        visit_aggregates[c.plate_id] = int(cnt)

        # Most recent SiteVisit in this cycle (closed or open)
        latest = (SiteVisit.query
                  .filter(SiteVisit.cycle_id == c.id,
                          SiteVisit.is_drive_by == False)   # noqa: E712
                  .order_by(SiteVisit.enter_at.desc())
                  .first())
        if latest:
            gf = (CartrackGeofence.query
                  .filter_by(id=latest.geofence_id).first())
            gf_name = gf.name if gf else None
            last_arrival_map[c.plate_id] = {
                'time': iso_ph(latest.enter_at),
                'location': gf_name,
                'still_here': latest.exit_at is None,
            }
            if latest.exit_at:
                last_departure_map[c.plate_id] = {
                    'time': iso_ph(latest.exit_at),
                    'location': gf_name,
                }

    rows = []
    now = utc_now()
    for p in plates:
        state = state_map.get(p.id)
        oc = open_cycle_map.get(p.id)
        # Live status
        live_status = state.live_status if state else 'UNKNOWN'
        speed = state.last_speed if state else 0
        location = (state.last_position_description if state else '') or ''
        # Currently-at geofence names
        current_gfs = []
        if state and state.last_geofence_uuids:
            for u in state.last_geofence_uuids.split(','):
                if u and u in gf_name_map:
                    current_gfs.append(gf_name_map[u])
        # Open cycle details
        open_cycle_info = None
        if oc:
            elapsed_s = (now - oc.started_at).total_seconds() if oc.started_at else 0
            open_cycle_info = {
                'id':           oc.id,
                'started_at':   iso_ph(oc.started_at),
                'elapsed_hours': round(elapsed_s / 3600.0, 2),
            }

        # In-progress ad-hoc stop — only reported when truck has an
        # open stop tracker AND is not currently inside any known
        # geofence. Lets the UI render "STOPPED at <address> (Xm)"
        # without waiting for the truck to resume moving.
        ad_hoc_stop = None
        if (state and state.last_stop_started_at
                and not current_gfs):
            stop_elapsed = int((now - state.last_stop_started_at).total_seconds() / 60)
            ad_hoc_stop = {
                'address':         state.last_stop_address or location,
                'lat':             state.last_stop_lat,
                'lng':             state.last_stop_lng,
                'started_at':      iso_ph(state.last_stop_started_at),
                'duration_minutes': stop_elapsed,
            }

        rows.append({
            'plate_id':      p.id,
            'plate_no':      p.plate_no,
            'body_no':       p.body_no or '',
            'display':       p.display,
            'truck_type':    p.truck_type.code if p.truck_type else '',
            'truck_type_name': p.truck_type.name if p.truck_type else '',
            'live_status':   live_status,
            'speed':         speed,
            'location':      location,
            'current_geofences': current_gfs,
            'ad_hoc_stop':   ad_hoc_stop,
            'is_open':       oc is not None,
            'open_cycle':    open_cycle_info,
            'visit_count':   visit_aggregates.get(p.id, 0) if oc else 0,
            'last_arrival':  last_arrival_map.get(p.id) if oc else None,
            'last_departure': last_departure_map.get(p.id) if oc else None,
        })

    # Sort: open cycles first (longest elapsed first), then parked alphabetically
    def sort_key(r):
        if r['is_open']:
            return (0, -(r['open_cycle']['elapsed_hours'] or 0), r['display'])
        return (1, 0, r['display'])
    rows.sort(key=sort_key)

    open_count = sum(1 for r in rows if r['is_open'])
    return jsonify({
        'plates':       rows,
        'count':        len(rows),
        'open_count':   open_count,
        'parked_count': len(rows) - open_count,
    })


@app.route('/api/cycle-time/plate/<int:plate_id>/cycles')
@login_required
def api_cycle_time_plate_cycles(plate_id):
    """Return all cycles (open + closed) for a single plate, ordered
    most-recent-first. Used by the cycle-picker dropdown when a plate
    row is expanded on the TCT page.

    Query params:
        date_from, date_to  — ISO dates, filter by started_at
        limit               — max rows (default 50)
    """
    plate = Plate.query.get_or_404(plate_id)
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    limit     = min(int(request.args.get('limit', 50)), 200)

    q = TruckCycle.query.filter(TruckCycle.plate_id == plate.id)
    if date_from:
        try:
            q = q.filter(TruckCycle.started_at >= pht_filter_to_utc(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(TruckCycle.started_at < pht_filter_to_utc(date_to, plus_days=1))
        except ValueError:
            pass

    # Order: open cycles first (ended_at IS NULL → sorted as "newest"),
    # then closed cycles by started_at desc.
    cycles = (q.order_by(TruckCycle.ended_at.is_(None).desc(),
                          TruckCycle.started_at.desc())
              .limit(limit).all())

    from sqlalchemy import func
    rows = []
    now = utc_now()
    for c in cycles:
        # Real-visit count (exclude drive-bys; INCLUDE ad-hoc stops)
        vc = (db.session.query(func.count(SiteVisit.id))
              .filter(SiteVisit.cycle_id == c.id,
                      SiteVisit.is_drive_by == False)   # noqa: E712
              .scalar() or 0)
        is_open = c.ended_at is None
        duration_min = c.duration_minutes
        if is_open and c.started_at:
            duration_min = int((now - c.started_at).total_seconds() / 60)
        rows.append({
            'id':              c.id,
            'started_at':      iso_ph(c.started_at),
            'ended_at':        iso_ph(c.ended_at),
            'duration_minutes': duration_min,
            'duration_hours':  round(duration_min / 60, 2) if duration_min else None,
            'category':        c.category,
            'visit_count':     int(vc),
            'is_open':         is_open,
        })

    return jsonify({
        'plate_id':   plate.id,
        'plate_no':   plate.plate_no,
        'display':    plate.display,
        'cycles':     rows,
        'count':      len(rows),
    })


@app.route('/api/cycle-time/cycle/<int:cycle_id>/timeline')
@login_required
def api_cycle_time_cycle_timeline(cycle_id):
    """Return the full chronological audit trail of a single cycle.

    Merges two data sources:
      1. SiteVisit rows linked via cycle_id (geofence visits +
         ad-hoc stops). Each row produces TWO events: ARRIVED (enter_at)
         and DEPARTED (exit_at if set).
      2. CartrackEvent rows within the cycle's time range (plaza
         enter/exit/trip_closed events from the toll-tracking layer).

    Events are sorted by timestamp ascending so the user reads the
    truck's journey top-to-bottom.

    Each event has:
        ts          — ISO timestamp
        kind        — 'departed' | 'arrived' | 'stopped' | 'plaza_enter'
                       | 'plaza_exit' | 'trip_closed'
        location    — display name (geofence name / plaza / address)
        duration_minutes  — for visit close events
        idling_pct        — for visit close events (if computed)
        is_ad_hoc         — for stops outside any geofence
        notes             — free-text extras
    """
    cycle = TruckCycle.query.get_or_404(cycle_id)

    # ── Source 1: SiteVisits within this cycle ──
    visits = (SiteVisit.query
              .filter(SiteVisit.cycle_id == cycle.id,
                      SiteVisit.is_drive_by == False)   # noqa: E712
              .order_by(SiteVisit.enter_at.asc())
              .all())

    events = []
    # The cycle's own start/end events
    if cycle.started_at:
        events.append({
            'ts':       iso_ph(cycle.started_at),
            'kind':     'departed',
            'location': 'Home base',
            'is_ad_hoc': False,
            'notes':    'Cycle started — truck left home',
        })
    if cycle.ended_at:
        events.append({
            'ts':       iso_ph(cycle.ended_at),
            'kind':     'arrived',
            'location': 'Home base',
            'is_ad_hoc': False,
            'duration_minutes': cycle.duration_minutes,
            'notes':    'Cycle ended — truck returned home',
        })

    for v in visits:
        gf_name = v.geofence.name if v.geofence else (v.address or '(unknown location)')
        is_ad_hoc = v.geofence_id is None
        # Arrived event (enter)
        if v.enter_at:
            events.append({
                'ts':         iso_ph(v.enter_at),
                'kind':       'stopped' if is_ad_hoc else 'arrived',
                'location':   gf_name,
                'is_ad_hoc':  is_ad_hoc,
                'lat':        v.lat,
                'lng':        v.lng,
                'notes':      '',
            })
        # Departed event (exit, only if closed)
        if v.exit_at:
            events.append({
                'ts':              iso_ph(v.exit_at),
                'kind':            'departed',
                'location':        gf_name,
                'is_ad_hoc':       is_ad_hoc,
                'duration_minutes': (v.duration_seconds or 0) // 60,
                'idling_pct':      v.idling_pct,
                'notes':           '',
            })

    # ── Source 1b: in-progress ad-hoc stop ──
    # If the truck is currently in the middle of an ad-hoc stop that
    # started during this cycle, surface it as a synthetic 'stopped'
    # event so the timeline shows the live state without waiting for
    # the truck to resume moving. Only relevant for OPEN cycles.
    if cycle.ended_at is None:
        state = (CartrackTruckState.query
                 .filter_by(plate_id=cycle.plate_id).first())
        if (state and state.last_stop_started_at
                and (cycle.started_at is None
                     or state.last_stop_started_at >= cycle.started_at)):
            _now_utc = utc_now()
            duration_min = int((_now_utc - state.last_stop_started_at).total_seconds() / 60)
            events.append({
                'ts':          iso_ph(state.last_stop_started_at),
                'kind':        'stopped',
                'location':    state.last_stop_address or '(unknown location)',
                'is_ad_hoc':   True,
                'lat':         state.last_stop_lat,
                'lng':         state.last_stop_lng,
                'duration_minutes': duration_min,
                'in_progress': True,
                'notes':       f'currently here — {duration_min}m so far',
            })

    # ── Source 2: CartrackEvents within the cycle's time range ──
    plate_id = cycle.plate_id
    start_ts = cycle.started_at
    end_ts   = cycle.ended_at or utc_now()
    cev_q = (CartrackEvent.query
             .filter(CartrackEvent.plate_id == plate_id,
                     CartrackEvent.created_at >= start_ts,
                     CartrackEvent.created_at <= end_ts)
             .order_by(CartrackEvent.created_at.asc()))
    for e in cev_q.all():
        kind = {
            'enter':       'plaza_enter',
            'exit':        'plaza_exit',
            'trip_closed': 'trip_closed',
        }.get(e.event_type, e.event_type)
        events.append({
            'ts':       iso_ph(e.created_at),
            'kind':     kind,
            'location': e.plaza_name or '?',
            'is_ad_hoc': False,
            'expressway': e.expressway,
            'toll_fee':   e.toll_fee,
            'notes':      '',
        })

    # Final chronological sort
    events.sort(key=lambda x: x['ts'])

    return jsonify({
        'cycle_id':   cycle.id,
        'plate_id':   cycle.plate_id,
        'started_at': iso_ph(cycle.started_at),
        'ended_at':   iso_ph(cycle.ended_at),
        'duration_minutes': cycle.duration_minutes,
        'is_open':    cycle.ended_at is None,
        'events':     events,
        'count':      len(events),
    })


@app.route('/api/cycle-time/idling')
@login_required
def api_cycle_time_idling():
    """Return aggregated idling stats per truck.

    Computed across the date range from SiteVisit records:
      total_visits, total_duration_minutes, total_idling_minutes,
      idling_pct (overall), avg_idling_pct_per_visit.

    Used by the idling-rate bar chart on the Truck Cycle Time page.
    """
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    truck_type_id = request.args.get('truck_type_id', '')

    # Build query — excludes drive-by visits by default since they're
    # transient touches, not real delivery stops. Use ?include_drive_by=1
    # to see them.
    include_drive_by = request.args.get('include_drive_by', '0') in ('1', 'true', 'True')

    from sqlalchemy import func
    q = (db.session.query(
            SiteVisit.plate_id,
            func.count(SiteVisit.id).label('visits'),
            func.coalesce(func.sum(SiteVisit.duration_seconds), 0).label('total_dur'),
            func.coalesce(func.sum(SiteVisit.idling_seconds), 0).label('total_idle'),
            func.avg(SiteVisit.idling_pct).label('avg_pct'),
         )
         .filter(SiteVisit.exit_at.isnot(None))   # only closed visits
         .group_by(SiteVisit.plate_id))
    if not include_drive_by:
        q = q.filter(SiteVisit.is_drive_by == False)   # noqa: E712

    if date_from:
        try:
            q = q.filter(SiteVisit.enter_at >= pht_filter_to_utc(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(SiteVisit.enter_at < pht_filter_to_utc(date_to, plus_days=1))
        except ValueError:
            pass
    if truck_type_id:
        try:
            q = q.join(Plate, SiteVisit.plate_id == Plate.id) \
                 .filter(Plate.truck_type_id == int(truck_type_id))
        except ValueError:
            pass

    results = q.all()
    plate_map = {p.id: p for p in Plate.query.all()}

    rows = []
    for r in results:
        plate = plate_map.get(r.plate_id)
        total_dur = int(r.total_dur or 0)
        total_idle = int(r.total_idle or 0)
        pct_overall = round(100.0 * total_idle / total_dur, 1) if total_dur else 0
        rows.append({
            'plate_id':         r.plate_id,
            'plate_no':         plate.display if plate else 'N/A',
            'visits':           int(r.visits or 0),
            'total_minutes':    round(total_dur / 60, 1),
            'idling_minutes':   round(total_idle / 60, 1),
            'idling_pct':       pct_overall,
            'avg_idling_pct':   round(r.avg_pct or 0, 1),
        })
    # Sort by idling % desc
    rows.sort(key=lambda x: x['idling_pct'], reverse=True)
    return jsonify({'rows': rows, 'count': len(rows)})


@app.route('/api/cycle-time/filters')
@login_required
def api_cycle_time_filters():
    """Dropdown options: plates with cycle data, truck types, cycle categories."""
    plate_ids = [pid for (pid,) in
                 db.session.query(TruckCycle.plate_id).distinct().all()
                 if pid is not None]
    plates = (Plate.query.filter(Plate.id.in_(plate_ids))
                       .order_by(Plate.body_no, Plate.plate_no).all() if plate_ids else [])
    # Only surface truck types that actually have plates with cycle history —
    # avoids cluttering the dropdown with empty options.
    truck_type_ids = {p.truck_type_id for p in plates if p.truck_type_id}
    truck_types = (TruckTypeDef.query.filter(TruckTypeDef.id.in_(truck_type_ids))
                                    .order_by(TruckTypeDef.sort_order,
                                              TruckTypeDef.code).all()
                   if truck_type_ids else [])
    return jsonify({
        'plates': [{'id': p.id,
                    'plate_no': p.display,
                    'truck_type_id': p.truck_type_id}
                   for p in plates],
        'truck_types': [{'id': t.id, 'code': t.code, 'name': t.name,
                         'color': t.color} for t in truck_types],
        'categories': ['short', 'standard', 'long', 'ongoing'],
    })


@app.route('/api/cycle-time/export')
@login_required
def api_cycle_time_export():
    """Export cycles to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500

    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    plate_id  = request.args.get('plate_id', '')
    truck_type_id = request.args.get('truck_type_id', '')
    category  = request.args.get('category', '')

    q = TruckCycle.query
    if date_from:
        try:
            q = q.filter(TruckCycle.started_at >= pht_filter_to_utc(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(TruckCycle.started_at < pht_filter_to_utc(date_to, plus_days=1))
        except ValueError:
            pass
    if plate_id:
        try:
            q = q.filter(TruckCycle.plate_id == int(plate_id))
        except ValueError:
            pass
    if truck_type_id:
        try:
            q = q.join(Plate, TruckCycle.plate_id == Plate.id) \
                 .filter(Plate.truck_type_id == int(truck_type_id))
        except ValueError:
            pass
    if category:
        q = q.filter(TruckCycle.category == category)

    cycles = q.order_by(TruckCycle.started_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Truck Cycles'

    headers = ['Started', 'Ended', 'Plate', 'Duration (hrs)', 'Category',
               'Visit Count', 'Total Idling (min)', 'Status']
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='8B1A2B')
    ca = Alignment(horizontal='center', vertical='center')
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ca

    for row_idx, c in enumerate(cycles, 2):
        plate = c.plate
        ws.cell(row=row_idx, column=1, value=c.started_at.strftime('%Y-%m-%d %H:%M') if c.started_at else '')
        ws.cell(row=row_idx, column=2, value=c.ended_at.strftime('%Y-%m-%d %H:%M') if c.ended_at else '')
        ws.cell(row=row_idx, column=3, value=plate.display if plate else 'N/A')
        ws.cell(row=row_idx, column=4, value=round(c.duration_minutes / 60, 2) if c.duration_minutes else None)
        ws.cell(row=row_idx, column=5, value=c.category)
        ws.cell(row=row_idx, column=6, value=SiteVisit.query.filter_by(cycle_id=c.id).count())
        ws.cell(row=row_idx, column=7, value=c.total_idling_minutes)
        ws.cell(row=row_idx, column=8, value='Open' if c.ended_at is None else 'Closed')

    column_widths = [18, 18, 14, 14, 12, 12, 18, 10]
    for col_idx, w in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'truck_cycles_{ph_today().isoformat()}.xlsx'
    )


@app.route('/api/cycle-time/sync-geofences', methods=['POST'])
@login_required
def api_cycle_time_sync_geofences():
    """Admin trigger: pull all geofences from Cartrack into local cache.

    Should be called once after setup, then periodically (e.g., weekly)
    to pick up new geofences admins create in Cartrack. The polling
    worker only reads from the local cache, so newly-created geofences
    in Cartrack won't be tracked until this sync runs.
    """
    try:
        from cartrack_poll import sync_geofences
        summary = sync_geofences(app=app)
        return jsonify(summary)
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


# ── Truck Cycle Time settings (configurable thresholds) ──────────────────
# Two operator-tunable thresholds backed by AppSetting:
#   MIN_VISIT_MINUTES     — geofence dwell shorter than this is drive-by
#   STOP_DETECTION_MINUTES — ad-hoc stops shorter than this are ignored
#
# Polling worker reads these on every iteration, so changes take effect
# on the next poll (no restart needed).
SETTING_MIN_VISIT = 'TCT_MIN_VISIT_MINUTES'
SETTING_STOP_MIN  = 'TCT_STOP_DETECTION_MINUTES'
DEFAULT_MIN_VISIT = 5
DEFAULT_STOP_MIN  = 10


def _get_int_setting(key, default):
    raw = AppSetting.get(key, '')
    try:
        return max(1, int(raw)) if raw else default
    except (TypeError, ValueError):
        return default


@app.route('/api/cycle-time/settings', methods=['GET'])
@login_required
def api_cycle_time_settings_get():
    """Return current operator-tunable thresholds. Read-only for any
    authenticated user — only admins can update via the POST below."""
    is_admin = session.get('user_role') == 'admin'
    return jsonify({
        'min_visit_minutes':       _get_int_setting(SETTING_MIN_VISIT, DEFAULT_MIN_VISIT),
        'stop_detection_minutes':  _get_int_setting(SETTING_STOP_MIN,  DEFAULT_STOP_MIN),
        'is_admin':                is_admin,
    })


@app.route('/api/cycle-time/settings', methods=['POST'])
@admin_required
def api_cycle_time_settings_post():
    """Admin-only: update operator thresholds.

    Body (JSON):
        {
          "min_visit_minutes":      5,
          "stop_detection_minutes": 10
        }
    """
    data = request.get_json(silent=True) or {}
    saved = {}
    for key, setting in (
        ('min_visit_minutes',      SETTING_MIN_VISIT),
        ('stop_detection_minutes', SETTING_STOP_MIN),
    ):
        if key in data:
            try:
                n = max(1, int(data[key]))
                AppSetting.set(setting, str(n))
                saved[key] = n
            except (TypeError, ValueError):
                return jsonify({'error': f'Invalid value for {key}'}), 400
    db.session.commit()
    return jsonify({'ok': True, 'saved': saved})


@app.route('/api/cycle-time/clear-logs', methods=['POST'])
@admin_required
def api_cycle_time_clear_logs():
    """Admin-only: wipe trial-period tracking data.

    Deletes:
      - SiteVisit         (geofence visits + ad-hoc stops)
      - TruckCycle        (cycle round-trip records)
      - CartrackEvent     (plaza ENTER / EXIT / trip_closed events)

    Preserves:
      - Plate, Driver, Helper, Product, Client, Dispatcher, TruckTypeDef
      - TripRecord, Wave  (all real operational data, including auto-filled
                           toll fees — those stay on the trip record)
      - BreakdownLog
      - CartrackGeofence  (geofence cache — synced from Cartrack)
      - CartrackTruckState (per-plate live state — reset position/stop
                            tracking so the next poll starts clean)

    Required body params (both — for safety, no full-wipe is allowed):
      {
        "before":  "2026-05-20",   # ISO date; only delete rows before this
        "confirm": "CLEAR"         # exact string match — case sensitive
      }

    The confirm token and date are both required at the API layer, so
    even a misclick in the UI or a stray curl can't accidentally wipe
    everything. Rows on or after `before` are preserved.
    """
    from sqlalchemy import func
    data = request.get_json(silent=True) or {}

    # Safety gate 1 — explicit confirmation token
    if data.get('confirm') != 'CLEAR':
        return jsonify({
            'error': 'Confirmation token required',
            'hint':  'Send {"confirm": "CLEAR", "before": "YYYY-MM-DD"}.',
        }), 400

    # Safety gate 2 — explicit date cutoff (no full wipe)
    if not data.get('before'):
        return jsonify({
            'error': 'Cutoff date required',
            'hint':  'Send a "before" ISO date so post-cutoff records are kept.',
        }), 400

    try:
        cutoff = datetime.strptime(data['before'], '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'before must be ISO date YYYY-MM-DD'}), 400

    # Count before delete (for reporting)
    def count(model, ts_attr):
        return (db.session.query(func.count(model.id))
                .filter(getattr(model, ts_attr) < cutoff)
                .scalar() or 0)

    sv_count  = count(SiteVisit,    'enter_at')
    cyc_count = count(TruckCycle,   'started_at')
    ev_count  = count(CartrackEvent,'created_at')

    # Delete in dependency-safe order: SiteVisit (refs cycle, geofence)
    # first, then TruckCycle, then CartrackEvent. All scoped to `before`
    # so post-cutoff records are preserved.
    (db.session.query(SiteVisit)
     .filter(SiteVisit.enter_at < cutoff)
     .delete(synchronize_session=False))
    (db.session.query(TruckCycle)
     .filter(TruckCycle.started_at < cutoff)
     .delete(synchronize_session=False))
    (db.session.query(CartrackEvent)
     .filter(CartrackEvent.created_at < cutoff)
     .delete(synchronize_session=False))

    # Per-plate live state is NOT reset (partial wipe leaves the live
    # tracking intact — we're only purging history before the cutoff).

    db.session.commit()
    return jsonify({
        'ok': True,
        'deleted': {
            'site_visits':    sv_count,
            'truck_cycles':   cyc_count,
            'cartrack_events': ev_count,
        },
        'cutoff': data['before'],
    })


# ── GOOGLE SHEETS SYNC ────────────────────────────────────────────────────
@app.route('/api/sync-to-sheets', methods=['POST'])
@login_required
def api_sync_to_sheets():
    import urllib.request, json as _json

    webhook_url = AppSetting.get(SHEETS_WEBHOOK_KEY, SHEETS_WEBHOOK_DEFAULT)
    if not webhook_url:
        return jsonify({'error': 'Google Sheets webhook URL not configured.'}), 400

    # ── Build payload ──────────────────────────────────────────────────────
    # TRIPS
    trip_headers = ['Date','Truck Type','Wave','Trip #','Driver','Helper',
                    'Plate','Product','Client','Dispatcher','Trip Type',
                    'RS No','PO No','Reference','DR No','Volume','Status','Notes']
    trips_rows = []
    for w in Wave.query.order_by(Wave.date.desc()).all():
        for t in w.trips:
            trips_rows.append([
                iso_ph(w.date),
                w.truck_type.name if w.truck_type else '',
                w.label,
                t.trip_number,
                t.driver.name     if t.driver     else '',
                t.helper.name     if t.helper     else '',
                t.plate.display   if t.plate      else '',
                t.product.name    if t.product    else '',
                t.client.name     if t.client     else '',
                t.dispatcher.name if t.dispatcher else '',
                t.trip_type or '',
                t.rs_no or '', t.po_no or '', t.reference or '',
                t.dr_no or '', t.volume or '',
                t.status or '', t.notes or '',
            ])

    # ATTENDANCE
    att_headers = ['Driver','Date','Status','Remarks']
    att_rows = []
    for a in Attendance.query.order_by(Attendance.date.desc()).all():
        att_rows.append([
            a.driver.name if a.driver else '',
            iso_ph(a.date),
            a.status or '',
            a.remarks or '',
        ])

    # BREAKDOWN
    bd_headers = ['Plate','Date','Description','Status','Resolved Date','Remarks']
    bd_rows = []
    for b in BreakdownLog.query.order_by(BreakdownLog.date.desc()).all():
        bd_rows.append([
            b.plate.display if b.plate else '',
            iso_ph(b.date),
            b.description or '',
            b.status or '',
            iso_ph(b.resolved_date) if b.resolved_date else '',
            b.remarks or '',
        ])

    # DRIVERS
    drv_headers = ['Name','Active']
    drv_rows = [[d.name, 'Yes' if d.active else 'No']
                for d in Driver.query.order_by(Driver.name).all()]

    # PLATES
    plt_headers = ['Plate No','Body No','Truck Type','Active']
    plt_rows = [[p.plate_no, p.body_no or '',
                 p.truck_type.name if p.truck_type else '',
                 'Yes' if p.active else 'No']
                for p in Plate.query.order_by(Plate.plate_no).all()]

    payload = {
        'action': 'sync_all',
        'trips':      {'headers': trip_headers, 'rows': trips_rows},
        'attendance': {'headers': att_headers,  'rows': att_rows},
        'breakdown':  {'headers': bd_headers,   'rows': bd_rows},
        'drivers':    {'headers': drv_headers,  'rows': drv_rows},
        'plates':     {'headers': plt_headers,  'rows': plt_rows},
    }

    # ── POST to Google Apps Script ─────────────────────────────────────────
    try:
        body = _json.dumps(payload).encode('utf-8')
        req  = urllib.request.Request(
            webhook_url,
            data=body,
            headers={'Content-Type': 'application/json'},
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = _json.loads(resp.read().decode('utf-8'))

        if result.get('ok'):
            # Save last sync timestamp
            AppSetting.set('last_sheets_sync', ph_now().strftime('%b %d, %Y %I:%M %p'))
            db.session.commit()
            log_change('Synced data to Google Sheets', 'backup')
            db.session.commit()
            return jsonify({'ok': True,
                            'synced': {
                                'trips': len(trips_rows),
                                'attendance': len(att_rows),
                                'breakdown': len(bd_rows),
                                'drivers': len(drv_rows),
                                'plates': len(plt_rows),
                            }})
        else:
            return jsonify({'error': result.get('error', 'Unknown error from Google Sheets')}), 500

    except Exception as ex:
        # If the failure happened between AppSetting.set() and commit, the
        # session is dirty — roll back or every later request on this
        # worker dies with PendingRollbackError.
        db.session.rollback()
        return jsonify({'error': str(ex)}), 500


@app.route('/api/restore-from-sheets', methods=['POST'])
@login_required
def api_restore_from_sheets():
    import urllib.request, json as _json
    from auth.routes import check_can_delete

    # Admin only
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Admin access required.'}), 403

    webhook_url = AppSetting.get(SHEETS_WEBHOOK_KEY, SHEETS_WEBHOOK_DEFAULT)
    if not webhook_url:
        return jsonify({'error': 'Webhook URL not configured.'}), 400

    # Fetch data from Google Sheets via Apps Script doGet
    try:
        get_url = webhook_url + '?action=export'
        req = urllib.request.Request(get_url, method='GET')
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read().decode('utf-8')
            data = _json.loads(raw)
    except Exception as ex:
        return jsonify({'error': f'Could not fetch from Google Sheets: {str(ex)}'}), 500

    if not data.get('ok'):
        return jsonify({'error': data.get('error', 'Unknown error from Google Sheets')}), 500

    restored = {'trips': 0, 'drivers': 0, 'plates': 0, 'attendance': 0, 'breakdown': 0}

    try:
        # ── Restore Drivers ───────────────────────────────────────────────
        for row in data.get('drivers', []):
            name = (row.get('Name') or '').strip()
            if not name: continue
            if not Driver.query.filter_by(name=name).first():
                db.session.add(Driver(name=name, active=True))
                restored['drivers'] += 1
        db.session.commit()

        # ── Restore Plates ────────────────────────────────────────────────
        for row in data.get('plates', []):
            plate_no = (row.get('Plate No') or '').strip()
            if not plate_no: continue
            if not Plate.query.filter_by(plate_no=plate_no).first():
                db.session.add(Plate(plate_no=plate_no,
                                     body_no=row.get('Body No') or None,
                                     active=True))
                restored['plates'] += 1
        db.session.commit()

        # ── Restore Trips ─────────────────────────────────────────────────
        for row in data.get('trips', []):
            date_str = (row.get('Date') or '').strip()
            if not date_str: continue
            try:
                d = date.fromisoformat(date_str)
            except Exception:
                continue

            truck_name = (row.get('Truck Type') or '').strip()
            tt = TruckTypeDef.query.filter_by(name=truck_name).first()
            if not tt: continue

            wave_label = (row.get('Wave') or '').strip()
            wave_num = 1
            for num, label in [(1,'1st Wave'),(2,'2nd Wave'),(3,'3rd Wave'),
                               (4,'4th Wave'),(5,'5th Wave'),(6,'6th Wave'),
                               (7,'7th Wave'),(8,'8th Wave')]:
                if label == wave_label:
                    wave_num = num; break

            wave = Wave.query.filter_by(date=d, truck_type_id=tt.id,
                                        wave_number=wave_num).first()
            if not wave:
                wave = Wave(date=d, truck_type_id=tt.id, wave_number=wave_num)
                db.session.add(wave)
                db.session.commit()

            # Use original trip number from sheet
            try:
                trip_num = int(float(str(row.get('Trip #') or 1)))
            except Exception:
                trip_num = 1

            # Skip if this trip already exists in the wave (prevent duplicates)
            if TripRecord.query.filter_by(wave_id=wave.id, trip_number=trip_num).first():
                continue

            drv  = Driver.query.filter_by(name=(row.get('Driver') or '').strip()).first()
            prod = Product.query.filter_by(name=(row.get('Product') or '').strip()).first()
            cli  = Client.query.filter_by(name=(row.get('Client') or '').strip()).first()

            # Plate display may be "body_no / plate_no" — extract plate_no
            plate_display = (row.get('Plate') or '').strip()
            plt = None
            if plate_display:
                if ' / ' in plate_display:
                    plate_no_part = plate_display.split(' / ', 1)[1].strip()
                    plt = Plate.query.filter_by(plate_no=plate_no_part).first()
                else:
                    plt = Plate.query.filter_by(plate_no=plate_display).first()

            trip = TripRecord(
                wave_id      = wave.id,
                trip_number  = trip_num,
                driver_id    = drv.id  if drv  else None,
                plate_id     = plt.id  if plt  else None,
                product_id   = prod.id if prod else None,
                client_id    = cli.id  if cli  else None,
                trip_type    = row.get('Trip Type') or None,
                rs_no        = row.get('RS No') or None,
                po_no        = row.get('PO No') or None,
                dr_no        = row.get('DR No') or None,
                volume       = row.get('Volume') or None,
                status       = row.get('Status') or 'Pending',
                notes        = row.get('Notes') or None,
            )
            db.session.add(trip)
            restored['trips'] += 1
        db.session.commit()

        log_change('Restored data from Google Sheets', 'backup')
        db.session.commit()
        return jsonify({'ok': True, 'restored': restored})

    except Exception as ex:
        db.session.rollback()
        return jsonify({'error': f'Restore failed: {str(ex)}'}), 500


@app.route('/api/sync-to-sheets/last')
@login_required
def api_last_sync():
    return jsonify({'last_sync': AppSetting.get('last_sheets_sync', None)})


@app.route('/api/sync-to-sheets/webhook', methods=['POST'])
@login_required
def api_save_webhook():
    data = request.get_json() or {}
    url  = (data.get('url') or '').strip()
    if not url:
        return jsonify({'error': 'URL required'}), 400
    AppSetting.set(SHEETS_WEBHOOK_KEY, url)
    db.session.commit()
    return jsonify({'ok': True})


# ── COLLAB API ─────────────────────────────────────────────────────────────
@app.route('/api/settings/save', methods=['POST'])
@login_required
def api_settings_save():
    data = request.get_json() or {}
    for key in DOC_HEADER_DEFAULTS:
        if key in data:
            AppSetting.set(key, data[key].strip())
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/search')
@login_required
def api_search():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify({'results': []})
    like = f'%{q}%'
    results = []

    for d in Driver.query.filter(Driver.name.ilike(like), Driver.active == True).limit(6).all():
        results.append({'type': 'Driver', 'label': d.name, 'sub': '', 'url': None})

    for p in Plate.query.filter(
        db.or_(Plate.plate_no.ilike(like), Plate.body_no.ilike(like)),
        Plate.active == True
    ).limit(6).all():
        results.append({'type': 'Plate', 'label': p.display, 'sub': p.truck_type.name if p.truck_type else '', 'url': None})

    for c in Client.query.filter(Client.name.ilike(like), Client.active == True).limit(4).all():
        results.append({'type': 'Client', 'label': c.name, 'sub': '', 'url': None})

    for pr in Product.query.filter(Product.name.ilike(like), Product.active == True).limit(4).all():
        results.append({'type': 'Product', 'label': pr.name, 'sub': '', 'url': None})

    for tr in (TripRecord.query
               .filter(db.or_(
                   TripRecord.rs_no.ilike(like),
                   TripRecord.po_no.ilike(like),
                   TripRecord.dr_no.ilike(like),
                   TripRecord.reference.ilike(like),
               ))
               .join(Wave)
               .order_by(Wave.date.desc())
               .limit(6).all()):
        date_str = iso_ph(tr.wave.date) if tr.wave else ''
        parts = [x for x in [tr.rs_no, tr.po_no, tr.dr_no] if x]
        results.append({
            'type': 'Trip',
            'label': ' / '.join(parts) or f'Trip #{tr.id}',
            'sub': date_str,
            'url': f'/schedule/{date_str}' if date_str else None,
        })

    return jsonify({'results': results, 'query': q})


@app.route('/api/activity')
@login_required
def api_activity():
    since_ts = request.args.get('since', 0, type=float)
    since_dt = datetime.fromtimestamp(since_ts) if since_ts else None

    logs = (ChangeLog.query.order_by(ChangeLog.timestamp.desc()).limit(12).all())
    new_count = 0
    if since_dt:
        new_count = ChangeLog.query.filter(ChangeLog.timestamp > since_dt).count()

    latest_ts = logs[0].timestamp.timestamp() if logs else 0
    return jsonify({
        'latest_ts': latest_ts,
        'new_count': new_count,
        'logs': [{
            'user':   l.user_name,
            'action': l.action,
            'time':   l.timestamp.replace(tzinfo=timezone.utc).astimezone(PH_TZ).strftime('%b %d, %I:%M %p'),
            'ts':     l.timestamp.timestamp(),
        } for l in logs]
    })


@app.route('/api/set-user', methods=['POST'])
@login_required
def api_set_user():
    name = (request.get_json() or {}).get('name', '').strip()
    if name:
        session['user_name'] = name
        return jsonify({'ok': True, 'name': name})
    return jsonify({'error': 'Name required'}), 400


# ── INIT ───────────────────────────────────────────────────────────────────
def init_db():
    with app.app_context():
        db.create_all()

        # ─────────────────────────────────────────────────────────────────────
        # SCHEMA MIGRATIONS — must run FIRST, before any ORM .query.all() calls.
        # SQLAlchemy generates SELECTs that include EVERY model column, so a
        # query against an unmigrated table will crash with "no such column".
        # All migrations are idempotent (column-existence checks).
        # ─────────────────────────────────────────────────────────────────────
        from sqlalchemy import inspect, text
        import sys
        inspector = inspect(db.engine)

        def has_table(name):
            try:
                return name in inspector.get_table_names()
            except Exception:
                return False

        def cols_of(table):
            try:
                return {c['name'] for c in inspector.get_columns(table)}
            except Exception:
                return set()

        def add_col(table, col, ddl):
            if not has_table(table):
                return
            if col in cols_of(table):
                return
            try:
                with db.engine.connect() as conn:
                    conn.execute(text(ddl))
                    conn.commit()
                msg = f"  Migrated: added {col} column to {table}"
                print(msg, flush=True)
                sys.stderr.write(msg + "\n")   # also write to error log so it's visible
                sys.stderr.flush()
            except Exception as e:
                err = f"  Migration FAILED for {table}.{col}: {e}"
                print(err, flush=True)
                sys.stderr.write(err + "\n")
                sys.stderr.flush()

        # trip_records — toll calculator + trip_type
        add_col('trip_records', 'trip_type',       'ALTER TABLE trip_records ADD COLUMN trip_type VARCHAR(30)')
        add_col('trip_records', 'toll_fee',        'ALTER TABLE trip_records ADD COLUMN toll_fee FLOAT')
        add_col('trip_records', 'toll_expressway', 'ALTER TABLE trip_records ADD COLUMN toll_expressway VARCHAR(50)')
        add_col('trip_records', 'toll_entry',      'ALTER TABLE trip_records ADD COLUMN toll_entry VARCHAR(80)')
        add_col('trip_records', 'toll_exit',       'ALTER TABLE trip_records ADD COLUMN toll_exit VARCHAR(80)')
        add_col('trip_records', 'toll_class',      'ALTER TABLE trip_records ADD COLUMN toll_class VARCHAR(10)')

        # clients — legacy toll fee column
        add_col('clients', 'toll_fee', 'ALTER TABLE clients ADD COLUMN toll_fee FLOAT DEFAULT 0')

        # truck_type_defs — fleet utilization scoring
        add_col('truck_type_defs', 'point_per_leg',       'ALTER TABLE truck_type_defs ADD COLUMN point_per_leg FLOAT DEFAULT 1.0')
        add_col('truck_type_defs', 'daily_target_points', 'ALTER TABLE truck_type_defs ADD COLUMN daily_target_points FLOAT DEFAULT 1.5')

        # products — full-day trip flag
        add_col('products', 'is_full_day_trip', 'ALTER TABLE products ADD COLUMN is_full_day_trip BOOLEAN DEFAULT 0')

        # breakdown_log — precise start/end timestamps for hours calculation
        add_col('breakdown_log', 'started_at', 'ALTER TABLE breakdown_log ADD COLUMN started_at DATETIME')
        add_col('breakdown_log', 'ended_at',   'ALTER TABLE breakdown_log ADD COLUMN ended_at DATETIME')

        # breakdown_log — ERP Repair Request integration columns
        # Populated by joborders_sync.py when a breakdown is sourced from
        # the gainersand.ph ERP. Indexed on jo_external_id so the upsert
        # lookup stays fast even as the table grows.
        add_col('breakdown_log', 'jo_external_id',
                'ALTER TABLE breakdown_log ADD COLUMN jo_external_id INTEGER')
        add_col('breakdown_log', 'jo_ref_no',
                'ALTER TABLE breakdown_log ADD COLUMN jo_ref_no VARCHAR(30)')
        add_col('breakdown_log', 'equipment_name',
                'ALTER TABLE breakdown_log ADD COLUMN equipment_name VARCHAR(200)')
        add_col('breakdown_log', 'equipment_brand',
                'ALTER TABLE breakdown_log ADD COLUMN equipment_brand VARCHAR(100)')
        add_col('breakdown_log', 'operator_name',
                'ALTER TABLE breakdown_log ADD COLUMN operator_name VARCHAR(100)')
        add_col('breakdown_log', 'requested_by',
                'ALTER TABLE breakdown_log ADD COLUMN requested_by VARCHAR(100)')
        add_col('breakdown_log', 'approved_by_dispatcher',
                'ALTER TABLE breakdown_log ADD COLUMN approved_by_dispatcher VARCHAR(100)')
        add_col('breakdown_log', 'approved_by_maintenance',
                'ALTER TABLE breakdown_log ADD COLUMN approved_by_maintenance VARCHAR(100)')
        add_col('breakdown_log', 'jo_url',
                'ALTER TABLE breakdown_log ADD COLUMN jo_url VARCHAR(300)')
        add_col('breakdown_log', 'last_synced_at',
                'ALTER TABLE breakdown_log ADD COLUMN last_synced_at DATETIME')
        # Index on the upsert key — created idempotently (CREATE INDEX
        # IF NOT EXISTS is supported in SQLite 3.8+; PA's SQLite is much
        # newer).
        try:
            with db.engine.connect() as conn:
                conn.execute(text(
                    'CREATE INDEX IF NOT EXISTS ix_breakdown_log_jo_external_id '
                    'ON breakdown_log(jo_external_id)'
                ))
                conn.commit()
        except Exception as _mig_err:
            print(f"  Migration warning (jo_external_id index): {_mig_err}")

        # plates — Cartrack GPS provider linkage (which Cartrack vehicle this plate maps to)
        add_col('plates', 'cartrack_vehicle_id',
                'ALTER TABLE plates ADD COLUMN cartrack_vehicle_id INTEGER')

        # plates — per-plate NLEX/SCTEX toll class for GPS auto-fill rate
        # lookup. Default to 'Class 3' (heavy trucks) since that's the
        # majority of the fleet — admins flip to Class 1 / Class 2 for
        # vans and light trucks via the Master Data Plates UI.
        add_col('plates', 'toll_class',
                "ALTER TABLE plates ADD COLUMN toll_class VARCHAR(10) DEFAULT 'Class 3'")

        # drivers — truck type category (which type they're trained/assigned to drive)
        # Legacy single-category column. Source of truth moved to driver_truck_types
        # association table below; this column is kept and backfilled for compat.
        add_col('drivers', 'truck_type_id', 'ALTER TABLE drivers ADD COLUMN truck_type_id INTEGER')

        # driver_truck_types — many-to-many association so a driver can be qualified
        # for multiple truck types. Created by db.create_all() above; here we
        # backfill it from the legacy Driver.truck_type_id column on first run.
        if has_table('driver_truck_types') and has_table('drivers'):
            existing_pairs = set()
            try:
                with db.engine.connect() as conn:
                    rows = conn.execute(text(
                        'SELECT driver_id, truck_type_id FROM driver_truck_types'
                    )).fetchall()
                    existing_pairs = {(r[0], r[1]) for r in rows}
            except Exception:
                pass
            try:
                with db.engine.connect() as conn:
                    legacy = conn.execute(text(
                        'SELECT id, truck_type_id FROM drivers '
                        'WHERE truck_type_id IS NOT NULL'
                    )).fetchall()
                    inserted = 0
                    for drv_id, tt_id in legacy:
                        if (drv_id, tt_id) in existing_pairs:
                            continue
                        try:
                            conn.execute(text(
                                'INSERT INTO driver_truck_types (driver_id, truck_type_id) '
                                'VALUES (:d, :t)'
                            ), {'d': drv_id, 't': tt_id})
                            inserted += 1
                        except Exception:
                            pass
                    if inserted:
                        conn.commit()
                        msg = f"  Migrated: backfilled {inserted} driver-to-truck-type associations"
                        print(msg, flush=True)
                        sys.stderr.write(msg + "\n")
                        sys.stderr.flush()
            except Exception as e:
                err = f"  Migration FAILED for driver_truck_types backfill: {e}"
                print(err, flush=True)
                sys.stderr.write(err + "\n")
                sys.stderr.flush()

        # ─────────────────────────────────────────────────────────────────────
        # ORM-based seeding/backfill — safe now that schema is up to date.
        # ─────────────────────────────────────────────────────────────────────
        # Add any missing truck types (works for fresh AND existing databases)
        existing_codes = {t.code for t in TruckTypeDef.query.all()}
        added = []
        for t in TRUCK_TYPES_SEED:
            if t['code'] not in existing_codes:
                db.session.add(TruckTypeDef(**t))
                added.append(t['code'])
        if added:
            db.session.commit()
            print(f"  Truck types added: {', '.join(added)}")
        # Seed default document header settings
        for key, val in DOC_HEADER_DEFAULTS.items():
            if not AppSetting.query.filter_by(key=key).first():
                db.session.add(AppSetting(key=key, value=val))
        # Seed Google Sheets webhook URL
        if not AppSetting.query.filter_by(key=SHEETS_WEBHOOK_KEY).first():
            db.session.add(AppSetting(key=SHEETS_WEBHOOK_KEY, value=SHEETS_WEBHOOK_DEFAULT))
        db.session.commit()

        # Backfill defaults for existing truck types based on code (10W gets 0.5/4.0)
        for tt in TruckTypeDef.query.all():
            seed = next((s for s in TRUCK_TYPES_SEED if s['code'] == tt.code), None)
            if seed:
                if tt.point_per_leg is None or (tt.point_per_leg == 1.0 and tt.code == '10W'):
                    tt.point_per_leg = seed['point_per_leg']
                if tt.daily_target_points is None or (tt.daily_target_points == 1.5 and tt.code == '10W'):
                    tt.daily_target_points = seed['daily_target_points']
        # Auto-flag known full-day products (case-insensitive name match)
        for prod_name in FULL_DAY_PRODUCTS_SEED:
            prod = Product.query.filter(db.func.upper(Product.name) == prod_name.upper()).first()
            if prod and not prod.is_full_day_trip:
                prod.is_full_day_trip = True
        db.session.commit()
        # Migrate: add can_delete column to users if missing
        from auth.models import User
        try:
            ucols = [c['name'] for c in inspector.get_columns('users')]
            if 'can_delete' not in ucols:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE users ADD COLUMN can_delete BOOLEAN DEFAULT 0'))
                    conn.commit()
                print("  Migrated: added can_delete column to users")
        except Exception:
            pass  # users table may not exist yet on first run

        # Migrate: add is_drive_by column to site_visits if missing.
        # site_visits was first introduced without this column (commit 3052dbf);
        # is_drive_by was added in 644c005. db.create_all() doesn't ALTER
        # existing tables, so this guarded ALTER is needed for the polling
        # worker to query SiteVisit successfully on accounts that picked up
        # the first version of the table.
        try:
            if 'site_visits' in inspector.get_table_names():
                svcols = [c['name'] for c in inspector.get_columns('site_visits')]
                if 'is_drive_by' not in svcols:
                    with db.engine.connect() as conn:
                        conn.execute(text(
                            'ALTER TABLE site_visits ADD COLUMN is_drive_by BOOLEAN DEFAULT 0'
                        ))
                        conn.commit()
                    print("  Migrated: added is_drive_by column to site_visits")
        except Exception as _mig_err:
            print(f"  Migration warning (site_visits.is_drive_by): {_mig_err}")

        # Migrate: add live-status columns to cartrack_truck_state.
        # These power the "where is each truck right now" view on the
        # Truck Cycle Time page. Older databases will have cartrack_truck_state
        # without these columns and the polling worker will crash on update.
        try:
            if 'cartrack_truck_state' in inspector.get_table_names():
                ctscols = [c['name'] for c in inspector.get_columns('cartrack_truck_state')]
                migrations = [
                    ('last_position_description', 'VARCHAR(300)'),
                    ('last_idling',               'BOOLEAN DEFAULT 0'),
                    ('last_ignition',             'BOOLEAN DEFAULT 0'),
                    ('last_speed',                'INTEGER DEFAULT 0'),
                    ('last_geofence_uuids',       'TEXT DEFAULT \'\''),
                ]
                added = []
                for col_name, col_type in migrations:
                    if col_name not in ctscols:
                        with db.engine.connect() as conn:
                            conn.execute(text(
                                f'ALTER TABLE cartrack_truck_state ADD COLUMN {col_name} {col_type}'
                            ))
                            conn.commit()
                        added.append(col_name)
                if added:
                    print(f"  Migrated: added {len(added)} columns to cartrack_truck_state: {', '.join(added)}")
        except Exception as _mig_err:
            print(f"  Migration warning (cartrack_truck_state live fields): {_mig_err}")

        # Migrate: add stop-tracking columns to cartrack_truck_state +
        # ad-hoc stop fields to site_visits (lat/lng/address). Powers
        # detection of stops outside any known geofence.
        try:
            if 'cartrack_truck_state' in inspector.get_table_names():
                ctscols = [c['name'] for c in inspector.get_columns('cartrack_truck_state')]
                stop_migrations = [
                    ('last_stop_started_at', 'DATETIME'),
                    ('last_stop_lat',        'FLOAT'),
                    ('last_stop_lng',        'FLOAT'),
                    ('last_stop_address',    'VARCHAR(300)'),
                ]
                added = []
                for col_name, col_type in stop_migrations:
                    if col_name not in ctscols:
                        with db.engine.connect() as conn:
                            conn.execute(text(
                                f'ALTER TABLE cartrack_truck_state ADD COLUMN {col_name} {col_type}'
                            ))
                            conn.commit()
                        added.append(col_name)
                if added:
                    print(f"  Migrated: added {len(added)} stop-tracking columns to cartrack_truck_state")
            if 'site_visits' in inspector.get_table_names():
                svcols = [c['name'] for c in inspector.get_columns('site_visits')]
                # outside_poll_count powers the hysteresis guard on
                # non-toll exits (see cartrack_poll.py exit-handling
                # block). Pre-existing rows default to 0, which means
                # "no outside polls accumulated" — identical to fresh
                # visit behaviour.
                sv_migrations = [
                    ('address',            'VARCHAR(300)'),
                    ('lat',                'FLOAT'),
                    ('lng',                'FLOAT'),
                    ('outside_poll_count', 'INTEGER DEFAULT 0'),
                ]
                added = []
                for col_name, col_type in sv_migrations:
                    if col_name not in svcols:
                        with db.engine.connect() as conn:
                            conn.execute(text(
                                f'ALTER TABLE site_visits ADD COLUMN {col_name} {col_type}'
                            ))
                            conn.commit()
                        added.append(col_name)
                if added:
                    print(f"  Migrated: added {len(added)} columns to site_visits ({', '.join(added)})")
        except Exception as _mig_err:
            print(f"  Migration warning (stop-detection columns): {_mig_err}")

        # Migrate: relax site_visits.geofence_id from NOT NULL -> nullable.
        # Required so ad-hoc stops (truck stopped outside any geofence)
        # can be logged as SiteVisit rows with geofence_id=NULL. SQLite
        # does not support ALTER COLUMN to change a NOT NULL constraint,
        # so we use the standard table-rebuild pattern: create new table
        # with the relaxed schema, copy data, drop old, rename new.
        # Idempotent — checks the current nullability before acting.
        try:
            if 'site_visits' in inspector.get_table_names():
                cols = {c['name']: c for c in inspector.get_columns('site_visits')}
                gf_col = cols.get('geofence_id')
                # column metadata uses 'nullable' (bool). Older SQLAlchemy
                # may also expose it under 'notnull' (int); handle both.
                is_not_null = (gf_col and
                               ((gf_col.get('nullable') is False)
                                or gf_col.get('notnull') == 1))
                if is_not_null:
                    print('  Migrating site_visits.geofence_id NOT NULL '
                           '-> nullable (table rebuild)...')
                    with db.engine.connect() as conn:
                        conn.execute(text('PRAGMA foreign_keys=OFF'))
                        conn.execute(text('''
                            CREATE TABLE site_visits_new (
                                id INTEGER PRIMARY KEY,
                                plate_id INTEGER NOT NULL,
                                geofence_id INTEGER,
                                enter_at DATETIME NOT NULL,
                                exit_at DATETIME,
                                duration_seconds INTEGER DEFAULT 0,
                                idling_seconds INTEGER DEFAULT 0,
                                idling_pct FLOAT,
                                is_drive_by BOOLEAN DEFAULT 0,
                                address VARCHAR(300),
                                lat FLOAT,
                                lng FLOAT,
                                trip_id INTEGER,
                                cycle_id INTEGER,
                                FOREIGN KEY (plate_id)    REFERENCES plates(id),
                                FOREIGN KEY (geofence_id) REFERENCES cartrack_geofences(id),
                                FOREIGN KEY (trip_id)     REFERENCES trip_records(id),
                                FOREIGN KEY (cycle_id)    REFERENCES truck_cycles(id)
                            )
                        '''))
                        # Copy data — column list explicit so it tolerates
                        # any ordering difference between the old and new
                        # tables. address/lat/lng exist on both at this
                        # point (the earlier migration created them).
                        conn.execute(text('''
                            INSERT INTO site_visits_new
                                (id, plate_id, geofence_id, enter_at, exit_at,
                                 duration_seconds, idling_seconds, idling_pct,
                                 is_drive_by, address, lat, lng, trip_id, cycle_id)
                            SELECT
                                 id, plate_id, geofence_id, enter_at, exit_at,
                                 duration_seconds, idling_seconds, idling_pct,
                                 is_drive_by, address, lat, lng, trip_id, cycle_id
                            FROM site_visits
                        '''))
                        conn.execute(text('DROP TABLE site_visits'))
                        conn.execute(text('ALTER TABLE site_visits_new '
                                           'RENAME TO site_visits'))
                        # Recreate the indexes that the model declared.
                        for idx_sql in [
                            'CREATE INDEX ix_site_visits_plate_id     ON site_visits(plate_id)',
                            'CREATE INDEX ix_site_visits_geofence_id  ON site_visits(geofence_id)',
                            'CREATE INDEX ix_site_visits_enter_at     ON site_visits(enter_at)',
                            'CREATE INDEX ix_site_visits_is_drive_by  ON site_visits(is_drive_by)',
                            'CREATE INDEX ix_site_visits_trip_id      ON site_visits(trip_id)',
                            'CREATE INDEX ix_site_visits_cycle_id     ON site_visits(cycle_id)',
                        ]:
                            try:
                                conn.execute(text(idx_sql))
                            except Exception:
                                pass   # index may already exist
                        conn.execute(text('PRAGMA foreign_keys=ON'))
                        conn.commit()
                    print('  Migrated: site_visits.geofence_id is now nullable')
        except Exception as _mig_err:
            print(f"  Migration warning (site_visits.geofence_id nullable): {_mig_err}")
        # Import DeleteRequest so db.create_all() picks it up
        from auth.models import DeleteRequest  # noqa: F401
        db.create_all()  # create delete_requests table if new
        # Create default admin account if none exists
        if not User.query.filter_by(username='admin').first():
            admin = User(username='admin', full_name='Administrator', role='admin')
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print("  Default admin created — username: admin  password: admin123")


# Initialize DB at module level so gunicorn (cloud) can start the app correctly
init_db()


if __name__ == '__main__':
    local_ip = get_local_ip()
    print(f"\n  ╔══════════════════════════════════════════╗")
    print(f"  ║   DISPATCH SCHEDULER is running          ║")
    print(f"  ║   Local  : http://localhost:5001         ║")
    print(f"  ║   Network: http://{local_ip}:5001        ║")
    print(f"  ║   Share the Network URL with your team   ║")
    print(f"  ╚══════════════════════════════════════════╝\n")
    port = int(os.environ.get('PORT', 5001))
    app.run(host='0.0.0.0', port=port, debug=False)
